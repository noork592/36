from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, status, Request, Response
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import io
import re
import json
import base64
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
from rapidfuzz import fuzz, process as rf_process, utils as rf_utils
import backup as backup_mod

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_SECRET = os.environ['JWT_SECRET']
JWT_ALG = os.environ.get('JWT_ALG', 'HS256')
JWT_EXPIRES_HOURS = int(os.environ.get('JWT_EXPIRES_HOURS', '24'))

app = FastAPI(title="Factory Order Management System")
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


# Once an order line is dispatched at or above this fraction of its
# originally ordered qty, the remaining tail is considered fulfilled and
# the line is dropped from the order. Other lines in the same order keep
# their pending qty until they cross the same threshold (or fully clear).
LINE_CLEAR_THRESHOLD = 0.85


def _apply_line_clear_threshold(
    order_items: List[Dict[str, Any]],
    dispatched_map: Dict[str, int],
    original_items: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """Subtract `dispatched_map` from `order_items` and drop any line whose
    cumulative dispatched qty has reached LINE_CLEAR_THRESHOLD of its
    original qty (looked up in `original_items`). Lines whose remaining
    qty hits 0 are also dropped. Other lines are kept with the reduced qty.
    """
    orig_qty_by_iid: Dict[str, int] = {
        oi.get("item_id"): int(oi.get("quantity") or 0)
        for oi in (original_items or [])
        if oi.get("item_id")
    }
    new_items: List[Dict[str, Any]] = []
    for it in order_items:
        iid = it.get("item_id")
        cur_qty = int(it.get("quantity") or 0)
        give = int(dispatched_map.get(iid, 0) or 0)
        new_qty = cur_qty - give
        if new_qty <= 0:
            continue  # fully dispatched — drop
        # Baseline for the 85% rule: the original ordered qty if known,
        # otherwise fall back to (current + dispatched) which equals the
        # qty this line had right before any dispatch ever touched it.
        orig_qty = orig_qty_by_iid.get(iid) or (cur_qty + give)
        if orig_qty > 0 and (orig_qty - new_qty) / orig_qty >= LINE_CLEAR_THRESHOLD:
            continue  # 85%+ shipped — auto-clear the residual tail
        new_items.append({**it, "quantity": new_qty})
    return new_items


# ======================== Synonyms & Product Defaults ========================
SYNONYM_MAP = {
    # Center Stand With Kit
    "center stand with kit": "Center Stand with Kit",
    "main stand with kit": "Center Stand with Kit",
    "double stand with kit": "Center Stand with Kit",
    "kit": "Center Stand with Kit",
    # Center Stand Without Kit
    "center stand without kit": "Center Stand without Kit",
    "main stand without kit": "Center Stand without Kit",
    "double stand without kit": "Center Stand without Kit",
    "center stand": "Center Stand without Kit",
    "main stand": "Center Stand without Kit",
    # Center Stand Pin
    "center stand pin": "Center Stand Pin",
    "stand pin": "Center Stand Pin",
    "pin": "Center Stand Pin",
    # Seat Kunda
    "seat kunda": "Seat Kunda",
    "side seat handle": "Seat Kunda",
    # Lady footrest synonyms (user said lady footrest = side footrest but provided separate bag limits)
    "lady footrest": "Lady Footrest",
    # Side Stand
    "side stand": "Side Stand",
    "side stand splender": "Side Stand",
    # Other
    "footrest rod": "Footrest Rod",
    "front footrest rod": "Footrest Rod",
    "handlebar": "Handlebar",
    "handle bar": "Handlebar",
    "number plate": "Number Plate",
    "front number plate": "Number Plate",
    "rear number plate": "Number Plate",
    "no plate": "Number Plate",
    "engine plate": "Engine Plate",
    "side footrest": "Side Footrest",
    # New master products
    "v-bracket": "V-Bracket",
    "v bracket": "V-Bracket",
    "bracket": "V-Bracket",
    "luggage rod": "Luggage Rod",
    "luggage": "Luggage Rod",
    "side mirror clump": "Side Mirror Clump",
    "mirror clump": "Side Mirror Clump",
    "side mirror": "Side Mirror Clump",
    "rear seat handle": "Rear Seat Handle",
    "back seat handle": "Rear Seat Handle",
    "seat handle": "Rear Seat Handle",
    # Hindi / Devanagari product hints
    "साइड स्टैंड": "Side Stand",
    "सेंटर स्टैंड": "Center Stand without Kit",
    "सेंटर स्टैंड किट": "Center Stand with Kit",
    "मेन स्टैंड": "Center Stand without Kit",
    "पिन": "Center Stand Pin",
    "सीट कुंडा": "Seat Kunda",
    "फुटरेस्ट रॉड": "Footrest Rod",
    "लेडी फुटरेस्ट": "Lady Footrest",
    "हैंडलबार": "Handlebar",
    "हैंडल बार": "Handlebar",
    "नंबर प्लेट": "Number Plate",
    "इंजन प्लेट": "Engine Plate",
    "ब्रैकेट": "V-Bracket",
    "लगेज रॉड": "Luggage Rod",
    "मिरर": "Side Mirror Clump",
}

DEFAULT_PRODUCTS = [
    {"name": "Side Stand", "min_per_bag": 180, "max_per_bag": 200, "variants": ["Type A", "Type B", "Type C"], "variant_field": "side_stand_type"},
    {"name": "Center Stand with Kit", "min_per_bag": 55, "max_per_bag": 55, "variants": [], "variant_field": "center_stand_kit"},
    {"name": "Center Stand without Kit", "min_per_bag": 60, "max_per_bag": 60, "variants": [], "variant_field": "center_stand_kit"},
    {"name": "Center Stand Pin", "min_per_bag": 50, "max_per_bag": 100, "variants": [], "variant_field": None},
    {"name": "Footrest Rod", "min_per_bag": 70, "max_per_bag": 70, "variants": [], "variant_field": None},
    {"name": "Seat Kunda", "min_per_bag": 250, "max_per_bag": 300, "variants": ["Fix", "Folding"], "variant_field": "seat_kunda_type"},
    {"name": "Lady Footrest", "min_per_bag": 250, "max_per_bag": 300, "variants": [], "variant_field": None},
    {"name": "Handlebar", "min_per_bag": 90, "max_per_bag": 100, "variants": [], "variant_field": None},
    {"name": "Number Plate", "min_per_bag": 300, "max_per_bag": 400, "variants": [], "variant_field": None},
    {"name": "Engine Plate", "min_per_bag": 200, "max_per_bag": 250, "variants": [], "variant_field": None},
    {"name": "Side Footrest", "min_per_bag": 50, "max_per_bag": 50, "variants": [], "variant_field": None},
    # New master products (Feb 2026) — bag limits default to 50/100, editable later
    {"name": "V-Bracket", "min_per_bag": 50, "max_per_bag": 100, "variants": [], "variant_field": None},
    {"name": "Luggage Rod", "min_per_bag": 50, "max_per_bag": 100, "variants": [], "variant_field": None},
    {"name": "Side Mirror Clump", "min_per_bag": 50, "max_per_bag": 100, "variants": [], "variant_field": None},
    {"name": "Rear Seat Handle", "min_per_bag": 50, "max_per_bag": 100, "variants": [], "variant_field": None},
]

# Map item-sheet category headers → master product name in DB.
# User's explicit rules:
#  KIT → Center Stand with Kit, CENTER STAND (WITHOUT KIT) → Center Stand without Kit,
#  PIN → Center Stand Pin, FRONT/REAR NUMBER PLATE → Number Plate,
#  FRONT FOOTREST ROD → Footrest Rod, HANDLE BAR → Handlebar,
#  ENGINE PLATE → Engine Plate, V-BRACKET → V-Bracket,
#  LUGGAGE ROD → Luggage Rod, SIDE MIRROR CLUMP → Side Mirror Clump,
#  REAR SEAT HANDLE → Rear Seat Handle.
CATEGORY_TO_PRODUCT: Dict[str, str] = {
    "JK SIDE STAND": "Side Stand",
    "JK CENTER STAND": "Center Stand without Kit",
    "JK CENTER STAND KIT": "Center Stand with Kit",
    "JK CENTER STAND PIN": "Center Stand Pin",
    "JK FRONT NUMBER PLATE": "Number Plate",
    "JK REAR NUMBER PLATE": "Number Plate",
    "JK FRONT FOOTREST ROD": "Footrest Rod",
    "JK V-BRACKET": "V-Bracket",
    "JK HANDLE BAR": "Handlebar",
    "MOTER CYCLE ENGINE PLATE": "Engine Plate",
    "JK SEAT KUNDA": "Seat Kunda",
    "JK LADY FOOTREST": "Lady Footrest",
    "JK SIDE FOOTREST": "Side Footrest",
    "JK LUGGAGE ROD": "Luggage Rod",
    "JK SIDE MIRROR CLUMP": "Side Mirror Clump",
    "JK REAR SEAT HANDLE": "Rear Seat Handle",
}


# ======================== Models ========================
class UserIn(BaseModel):
    email: str  # accepts either email or username (validated downstream)
    password: str


class TokenOut(BaseModel):
    token: str
    user: Dict[str, Any]


class OtpVerifyIn(BaseModel):
    challenge_id: str
    code: str


class ProductIn(BaseModel):
    name: str
    min_per_bag: int
    max_per_bag: int
    variants: List[str] = []
    variant_field: Optional[str] = None


class ProductUpdate(BaseModel):
    min_per_bag: Optional[int] = None
    max_per_bag: Optional[int] = None
    variants: Optional[List[str]] = None


class ItemBagUpdate(BaseModel):
    # Per-SKU bag override. When set, overrides master product's bag limits
    # during dispatch bag calculation.
    min_per_bag: int
    max_per_bag: int


class CustomerIn(BaseModel):
    name: str
    phone: Optional[str] = ""
    address: Optional[str] = ""
    city: Optional[str] = ""
    location: Optional[str] = ""
    preferences: Dict[str, str] = {}
    price_list_id: Optional[str] = None  # assigned price list (per-party pricing)
    transport_name: Optional[str] = ""   # transport company / vehicle name
    private_mark: Optional[str] = ""     # stenciled mark on packages for this party
    bill_number_mode: Optional[bool] = False  # if True, this party uses a manually-typed Bill Number at dispatch instead of a Private Marka
    blocked_items: List[str] = []        # item_ids that must never be ordered/dispatched for this party


class CustomerUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    location: Optional[str] = None
    preferences: Optional[Dict[str, str]] = None
    price_list_id: Optional[str] = None
    transport_name: Optional[str] = None
    private_mark: Optional[str] = None
    bill_number_mode: Optional[bool] = None
    blocked_items: Optional[List[str]] = None


class OrderItemIn(BaseModel):
    product_name: str
    quantity: int
    variant: Optional[str] = None  # e.g. "Type A" for side stand
    # Strict item-wise: every order line MUST identify a specific SKU.
    item_id: str
    item_name: str


class OrderIn(BaseModel):
    customer_id: str
    items: List[OrderItemIn]
    order_date: Optional[str] = None
    delivery_date: Optional[str] = None
    notes: Optional[str] = ""
    merge_with_pending: bool = False
    clear_previous_pending: bool = False


class OrderStatusUpdate(BaseModel):
    status: str  # Pending / Dispatched / Cleared


class OrderUpdate(BaseModel):
    """Admin-only full-order edit. Any subset of fields can be supplied."""
    customer_id: Optional[str] = None
    items: Optional[List[OrderItemIn]] = None
    order_date: Optional[str] = None
    delivery_date: Optional[str] = None
    notes: Optional[str] = None
    status: Optional[str] = None


class AdminUserCreate(BaseModel):
    email: EmailStr
    name: str
    password: str
    role: str = "user"  # "admin" or "user"
    username: Optional[str] = None  # admin-supplied short login id (defaults to email local-part)
    otp_login: bool = False  # require email OTP as a second login step
    permissions: Optional[List[str]] = None  # explicit nav-key allowlist (optional)


class AdminPasswordReset(BaseModel):
    password: str


class UserOtpUpdate(BaseModel):
    otp_login: bool


class UserPermissionsUpdate(BaseModel):
    # None = clear stored permissions, fall back to role-based defaults.
    # List of nav keys = explicit allowlist (overrides defaults; may include admin-only items).
    permissions: Optional[List[str]] = None


# ======================== Permission catalog ========================
# Every nav key that can be granted or revoked. Frontend mirrors this list
# verbatim — keep the two in sync if you add a new page / sidebar entry.
ALL_PERMISSION_KEYS: List[str] = [
    "dashboard", "orders", "newOrder", "dispatch", "purchaseCenter", "dispatchLedger", "vendorLedger", "dailyReport",
    "estimates",
    "customers", "products", "rawMaterials", "suppliers",
    "priceLists", "vendorPriceLists",
    "adminUsers", "adminSettings", "loginAudit",
]

# Default allowlist for a non-admin user when `permissions` is unset.
# Matches the historical behaviour (non-adminOnly sidebar items).
DEFAULT_USER_PERMISSIONS: List[str] = [
    "dashboard", "orders", "dispatch", "dispatchLedger", "dailyReport",
    "estimates",
    "customers", "products",
]

# ---- Action (edit / delete) permission keys ----------------------------
# Fine-grained grants, kept SEPARATE for edit vs delete, one pair per module.
# An admin can give or take these away from any user via Admin → Users →
# Access. Stored in the SAME `permissions` array as nav keys; the distinct
# `edit:` / `delete:` prefixes keep them from colliding with view/nav keys.
# Non-admins have NONE by default (view-only) — must be granted explicitly.
# Admins always have all of them.
ACTION_PERMISSION_KEYS: List[str] = [
    "edit:customers", "delete:customers",
    "edit:products", "delete:products",
    "edit:rawMaterials", "delete:rawMaterials",
    "edit:suppliers", "delete:suppliers",
    "edit:vendorLedger", "delete:vendorLedger",
    "edit:customerLedger", "delete:customerLedger",
    "edit:orders", "delete:orders",
    "edit:dispatch", "delete:dispatch",
    "edit:priceLists", "delete:priceLists",
    "edit:vendorPriceLists", "delete:vendorPriceLists",
]

# Everything an admin may grant / revoke (nav access + edit/delete actions).
ALL_GRANTABLE_KEYS: List[str] = ALL_PERMISSION_KEYS + ACTION_PERMISSION_KEYS


class SettingsUpdate(BaseModel):
    overdue_days: Optional[int] = None
    edit_window_days: Optional[int] = None


class ItemCreate(BaseModel):
    name: str
    product_id: str
    min_per_bag: Optional[int] = None  # SKU bag override (optional)
    max_per_bag: Optional[int] = None


class ItemEdit(BaseModel):
    """Admin edit of a single item SKU: name / product mapping / bag override.
    Distinct from ItemBagUpdate which strictly enforces both bag fields."""
    name: Optional[str] = None
    product_id: Optional[str] = None
    min_per_bag: Optional[int] = None
    max_per_bag: Optional[int] = None


class DispatchStockIn(BaseModel):
    # Strict item-wise: keys are item_id (SKU), values are qty available
    items: Dict[str, int]


class DispatchAllocationIn(BaseModel):
    item_id: str
    quantity: int
    # Optional free-text description / note attached to this line. Shows
    # on the printed dispatch slip directly under the item name.
    description: Optional[str] = ""


class DispatchExecuteIn(BaseModel):
    """Partial / lot-wise dispatch: subtract the given quantities from one
    pending order. If everything in the order reaches 0, mark Dispatched.
    Otherwise the order stays Pending with reduced quantities so the
    remaining lot can be dispatched later."""
    order_id: str
    allocations: List[DispatchAllocationIn]
    notes: Optional[str] = ""
    # Task 1 — Operator may pick a predefined price list for this customer
    # while dispatching. The choice is persisted on the customer doc so the
    # next dispatch suggests it automatically. `""` (empty string) means
    # the operator explicitly cleared the price list for this customer.
    price_list_id: Optional[str] = None
    # Optional backdate — operator may stamp a slip with an earlier date
    # (e.g. catching up on yesterday's paperwork). Accepts an ISO date
    # ("YYYY-MM-DD") or full ISO datetime; the helper `_resolve_dispatch_ts`
    # normalises it to a UTC ISO timestamp pinned at noon IST so the day
    # bucket and report filters line up correctly.
    dispatched_at: Optional[str] = None


# ======================== Price List Models ========================
class PriceListIn(BaseModel):
    name: str
    description: Optional[str] = ""
    # When False, parties on this list are considered "no-bill" — the
    # Dispatch Report will not prompt for a Bill Amount for their
    # dispatches and completeness is judged without it. Defaults to True
    # (existing behaviour).
    bill_amount_required: bool = True


class PriceListUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    bill_amount_required: Optional[bool] = None


class PriceListCloneIn(BaseModel):
    """Clone an existing price list (items + category discounts) under a new
    name. `description` is optional — falls back to the source list's
    description with a "(Copy)" suffix when omitted."""
    name: str
    description: Optional[str] = None


class PriceListItemIn(BaseModel):
    """Set/update the price for one item inside a price list."""
    item_id: str
    price: float


class CategoryDiscountIn(BaseModel):
    """Per-category (master product) discount within a price list.
    `discount_type` is either '₹' (flat rupees off) or '%' (percentage off)."""
    product_name: str
    discount_value: float
    discount_type: str  # '₹' or '%'


# ---- Login attestation (consent-based security capture) ----
class LoginAttestationIn(BaseModel):
    """Captured at login with the user's explicit consent.

    All fields are optional — the client posts whatever the user granted.
    If the user clicked Skip or denied permissions, set `consent=False`
    and the relevant `*_skipped` flag, leaving photo/location empty.
    """
    consent: bool = False
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    accuracy_meters: Optional[float] = None
    photo_b64: Optional[str] = None  # JPEG data URL or raw base64, capped server-side
    photo_skipped: bool = False
    location_skipped: bool = False
    error: Optional[str] = None  # free-text reason if permission was denied


# ---- Off-order (direct) dispatch ----
class OffOrderDispatchItemIn(BaseModel):
    item_id: str
    quantity: int
    # Optional free-text description / note attached to this line. Shows
    # on the printed dispatch slip directly under the item name (e.g.
    # "second quality", "loose pack", "sample"). Persisted verbatim on
    # the dispatch line.
    description: Optional[str] = ""


class OffOrderDispatchIn(BaseModel):
    """Dispatch a list of SKUs to a party that has no pending order.

    Either `customer_id` (existing party) OR a non-empty `customer_name`
    (walk-in / one-off) is required. `transport_name` overrides the
    customer's default; price list is taken from the customer if assigned.
    """
    customer_id: Optional[str] = None
    customer_name: Optional[str] = None
    transport_name: Optional[str] = None
    items: List[OffOrderDispatchItemIn]
    notes: Optional[str] = None
    # Operator-entered number of bags shipped in this dispatch. Optional —
    # editable later from the Daily Dispatch Report.
    bag_count: Optional[int] = None
    # Task 1 — Operator can pick a predefined price list for this customer
    # while dispatching. The choice is persisted on the customer doc so the
    # next dispatch suggests it automatically. `""` clears the assignment;
    # `None` leaves the customer's existing price list untouched.
    price_list_id: Optional[str] = None
    # Optional backdate — see DispatchExecuteIn.dispatched_at for the
    # accepted formats and timezone behaviour.
    dispatched_at: Optional[str] = None
    # Manually-typed Bill Number for parties configured in bill-number mode
    # (used instead of a Private Marka). Stored on the dispatch record.
    bill_number: Optional[str] = None


# ---- Bulk customer admin ----
class CustomerBulkDeleteIn(BaseModel):
    """Admin: delete many customers in one request. Each id is validated
    and the call is rejected (with details) if any of them are referenced
    by an order."""
    ids: List[str]


# ---- Dispatch ledger (GR number edit) ----
class DispatchGrUpdate(BaseModel):
    gr_number: str


# ======================== Helpers ========================
def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()


def verify_password(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode(), hashed.encode())
    except Exception:
        return False


def create_token(user: Dict[str, Any]) -> str:
    payload = {
        "sub": user["id"],
        "email": user["email"],
        "role": user["role"],
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRES_HOURS),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALG)


async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> Dict[str, Any]:
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALG])
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Invalid token")


def require_admin(user: Dict[str, Any] = Depends(get_current_user)) -> Dict[str, Any]:
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return user


def has_action_permission(user: Dict[str, Any], action_key: str) -> bool:
    """True if the user may perform the given edit/delete action.
    Admins always may. Non-admins only if the action key was explicitly
    granted (stored in their `permissions` array)."""
    if not user:
        return False
    if user.get("role") == "admin":
        return True
    perms = user.get("permissions")
    return isinstance(perms, list) and action_key in perms


def require_action(action_key: str):
    """Dependency factory that gates an edit/delete endpoint behind a
    fine-grained action permission. Keeps the injected value name-compatible
    with existing handlers (returns the current user dict)."""
    async def _dep(user: Dict[str, Any] = Depends(get_current_user)) -> Dict[str, Any]:
        if not has_action_permission(user, action_key):
            raise HTTPException(
                status_code=403,
                detail="You don't have permission to edit or delete here. Ask an admin to grant access.",
            )
        return user
    return _dep


# Demo accounts (e.g. JK1) carry the admin role but must NOT be able to view
# or manage the user list. This guard blocks them from the /users endpoints
# while leaving the rest of the admin surface untouched.
DEMO_USERNAMES = {"JK1"}


def require_users_admin(user: Dict[str, Any] = Depends(require_admin)) -> Dict[str, Any]:
    uname = str(user.get("username") or "").strip().upper()
    if uname in {u.upper() for u in DEMO_USERNAMES}:
        raise HTTPException(status_code=403, detail="Not permitted for this account")
    return user


def normalize_product_name(raw: str) -> Optional[str]:
    key = (raw or "").strip().lower()
    if not key:
        return None
    if key in SYNONYM_MAP:
        return SYNONYM_MAP[key]
    # fuzzy against synonyms
    match = rf_process.extractOne(key, list(SYNONYM_MAP.keys()), scorer=fuzz.WRatio)
    if match and match[1] >= 80:
        return SYNONYM_MAP[match[0]]
    # fuzzy against canonical product names
    canonical = list({v for v in SYNONYM_MAP.values()})
    match2 = rf_process.extractOne(key, canonical, scorer=fuzz.WRatio)
    if match2 and match2[1] >= 75:
        return match2[0]
    return None


# ======================== Seeding ========================
async def seed_db():
    # Users
    if await db.users.count_documents({}) == 0:
        admin = {"id": str(uuid.uuid4()), "email": "admin@factory.com", "username": "admin", "name": "Admin",
                 "password": hash_password("admin123"), "role": "admin", "otp_login": False, "created_at": now_iso()}
        user = {"id": str(uuid.uuid4()), "email": "user@factory.com", "username": "user", "name": "Operator",
                "password": hash_password("user123"), "role": "user", "otp_login": False, "created_at": now_iso()}
        await db.users.insert_many([admin, user])
        logger.info("Seeded default users")
    else:
        # Backfill otp_login only where the field is missing (preserve any
        # admin-chosen ON/OFF value). Default OFF — the feature is optional.
        await db.users.update_many({"otp_login": {"$exists": False}}, {"$set": {"otp_login": False}})
        # Backfill username for any pre-existing user (local-part of email, deduped)
        seen = set(u.get("username") for u in await db.users.find({"username": {"$exists": True}}, {"_id": 0, "username": 1}).to_list(1000) if u.get("username"))
        async for u in db.users.find({"username": {"$exists": False}}, {"_id": 0}):
            base = (u.get("email", "") or "").split("@")[0].lower() or u["id"][:8]
            uname = base
            i = 1
            while uname in seen:
                i += 1
                uname = f"{base}{i}"
            seen.add(uname)
            await db.users.update_one({"id": u["id"]}, {"$set": {"username": uname}})
    # Special demo account "JK1" — sees an EMPTY system (blank_view) so the app
    # can be shown/screenshared as if no data was ever entered. Real accounts
    # are unaffected. Idempotent: created once, never overwrites a changed pwd.
    if await db.users.count_documents({"username": "JK1"}) == 0:
        await db.users.insert_one({
            "id": str(uuid.uuid4()),
            "email": "jk1@factory.com",
            "username": "JK1",
            "name": "JK1 (Demo)",
            "password": hash_password("jk1123"),
            "role": "admin",
            "otp_login": False,
            "blank_view": True,
            "created_at": now_iso(),
        })
        logger.info("Seeded JK1 blank-view demo account")
    # One-time backfill (idempotent): existing parties whose private_mark is
    # PURELY NUMERIC are switched to Bill Number mode (mark cleared). Parties
    # whose mark contains letters are left untouched. Runs every startup but
    # only matches purely-numeric marks, so it self-heals across environments.
    _num_migrated = await db.customers.update_many(
        {"private_mark": {"$regex": r"^\s*\d+\s*$"}},
        {"$set": {"bill_number_mode": True, "private_mark": ""}},
    )
    if getattr(_num_migrated, "modified_count", 0):
        logger.info("Bill-number migration: converted %s numeric private marks", _num_migrated.modified_count)
    # Settings singleton — overdue order threshold (admin-configurable)
    if await db.settings.count_documents({"id": "global"}) == 0:
        await db.settings.insert_one({"id": "global", "overdue_days": 15, "edit_window_days": 3, "updated_at": now_iso()})
        logger.info("Seeded default settings (overdue_days=15, edit_window_days=3)")
    else:
        # Back-fill the new edit_window_days field on existing installs
        await db.settings.update_one(
            {"id": "global", "edit_window_days": {"$exists": False}},
            {"$set": {"edit_window_days": 3}},
        )
    # Products — additive: insert any from DEFAULT_PRODUCTS that don't exist yet
    existing_names = {p["name"] for p in await db.products.find({}, {"_id": 0, "name": 1}).to_list(1000)}
    new_products = [p for p in DEFAULT_PRODUCTS if p["name"] not in existing_names]
    if new_products:
        docs = [{"id": str(uuid.uuid4()), **p, "created_at": now_iso()} for p in new_products]
        await db.products.insert_many(docs)
        logger.info("Seeded %d new products", len(new_products))
    # Items — seed from data/items_parsed.json on first run
    if await db.items.count_documents({}) == 0:
        items_path = ROOT_DIR / "data" / "items_parsed.json"
        if items_path.exists():
            import json
            data = json.loads(items_path.read_text())
            # Build product name → id map
            prod_map = {p["name"]: p["id"] for p in await db.products.find({}, {"_id": 0}).to_list(1000)}
            docs = []
            unknown = set()
            for cat, items in data.items():
                product_name = CATEGORY_TO_PRODUCT.get(cat)
                if not product_name:
                    unknown.add(cat)
                    continue
                pid = prod_map.get(product_name)
                if not pid:
                    unknown.add(f"product missing: {product_name}")
                    continue
                for name in items:
                    docs.append({
                        "id": str(uuid.uuid4()),
                        "name": name,
                        "category": cat,
                        "product_id": pid,
                        "product_name": product_name,
                        "created_at": now_iso(),
                    })
            if docs:
                await db.items.insert_many(docs)
                logger.info("Seeded %d item SKUs", len(docs))
            if unknown:
                logger.warning("Item seed: unmapped categories: %s", unknown)


# ======================== Auth Routes ========================
def _user_public(user: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "id": user["id"],
        "email": user["email"],
        "username": user.get("username"),
        "name": user.get("name", ""),
        "role": user["role"],
        "permissions": user.get("permissions"),
    }


def is_blank_view(user: Optional[Dict[str, Any]]) -> bool:
    """True for the special demo account (JK1) that must see an empty system
    — no orders, dispatches or customers — as if nothing was ever entered.
    The real accounts are unaffected and see all live data."""
    return bool(user and user.get("blank_view"))


def _mask_email(addr: str) -> str:
    """Mask an email for display: j***n@gmail.com"""
    try:
        local, domain = (addr or "").split("@", 1)
    except ValueError:
        return addr or ""
    if len(local) <= 2:
        masked = local[0] + "*" if local else "*"
    else:
        masked = local[0] + ("*" * (len(local) - 2)) + local[-1]
    return f"{masked}@{domain}"


OTP_TTL_MINUTES = 10
OTP_MAX_ATTEMPTS = 5


@api_router.post("/auth/login")
async def login(body: UserIn):
    # Allow login by either email OR username (case-insensitive)
    ident = (body.email or "").strip().lower()
    if not ident:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    raw_ident = (body.email or "").strip()
    ci_username = {"username": {"$regex": f"^{re.escape(raw_ident)}$", "$options": "i"}}
    user = await db.users.find_one({"$or": [{"email": ident}, {"username": ident}, ci_username]}, {"_id": 0})
    if not user or not verify_password(body.password, user["password"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")

    # ── Two-step verification (email OTP) — DISABLED ─────────────────────
    # Optional email-OTP second step. Admin can turn this ON/OFF per user
    # from Admin → Users. When ON, the user must complete an email OTP as a
    # second step. The code is emailed to the same address configured for the
    # daily database backup, reusing those Gmail credentials.
    # TEMPORARILY DISABLED: OTP login is turned off. Set to True to re-enable.
    OTP_LOGIN_ENABLED = False
    if OTP_LOGIN_ENABLED and user.get("otp_login"):
        import random
        code = f"{random.randint(0, 999999):06d}"
        challenge_id = str(uuid.uuid4())
        expires_at = datetime.now(timezone.utc) + timedelta(minutes=OTP_TTL_MINUTES)
        await db.admin_otp_challenges.insert_one({
            "id": challenge_id,
            "user_id": user["id"],
            "code_hash": hash_password(code),
            "expires_at": expires_at.isoformat(),
            "attempts": 0,
            "created_at": now_iso(),
        })
        # Send the code by email (reuses the backup Gmail settings).
        sent_to = None
        email_ok = False
        try:
            sent_to = await backup_mod.send_otp_email(db, code)
            email_ok = True
        except Exception as e:
            logger.warning("OTP email send failed: %s", e)
        # Fallback for resilience/testing: always log the code server-side so
        # a misconfigured mailbox can never fully lock the admin out.
        logger.info("Admin OTP for %s (challenge %s): %s", user["email"], challenge_id, code)
        return {
            "otp_required": True,
            "challenge_id": challenge_id,
            "sent_to": _mask_email(sent_to) if sent_to else None,
            "email_sent": email_ok,
        }

    token = create_token(user)
    return {"token": token, "user": _user_public(user)}


@api_router.post("/auth/verify-otp", response_model=TokenOut)
async def verify_otp(body: OtpVerifyIn):
    challenge = await db.admin_otp_challenges.find_one({"id": body.challenge_id}, {"_id": 0})
    if not challenge:
        raise HTTPException(status_code=400, detail="Invalid or expired code. Please sign in again.")
    # Expiry check
    try:
        exp = datetime.fromisoformat(challenge["expires_at"])
    except Exception:
        exp = datetime.now(timezone.utc) - timedelta(seconds=1)
    if datetime.now(timezone.utc) > exp:
        await db.admin_otp_challenges.delete_one({"id": body.challenge_id})
        raise HTTPException(status_code=400, detail="Code expired. Please sign in again.")
    # Attempt limit
    if int(challenge.get("attempts", 0)) >= OTP_MAX_ATTEMPTS:
        await db.admin_otp_challenges.delete_one({"id": body.challenge_id})
        raise HTTPException(status_code=429, detail="Too many attempts. Please sign in again.")
    if not verify_password((body.code or "").strip(), challenge["code_hash"]):
        await db.admin_otp_challenges.update_one({"id": body.challenge_id}, {"$inc": {"attempts": 1}})
        raise HTTPException(status_code=401, detail="Incorrect code. Please try again.")
    # Success — consume the challenge and issue the token.
    await db.admin_otp_challenges.delete_one({"id": body.challenge_id})
    user = await db.users.find_one({"id": challenge["user_id"]}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    token = create_token(user)
    return {"token": token, "user": _user_public(user)}


@api_router.get("/auth/me")
async def me(user=Depends(get_current_user)):
    return user


# ======================== Admin: User Management ========================
@api_router.get("/users")
async def list_users(admin=Depends(require_users_admin)):
    users = await db.users.find({}, {"_id": 0, "password": 0}).sort("created_at", 1).to_list(500)
    return users


@api_router.post("/users")
async def create_user(body: AdminUserCreate, admin=Depends(require_users_admin)):
    if body.role not in ("admin", "user"):
        raise HTTPException(status_code=400, detail="role must be 'admin' or 'user'")
    email = body.email.lower()
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="A user with this email already exists")
    if len(body.password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    username = (body.username or "").strip().lower() or email.split("@")[0]
    # ensure username uniqueness
    base = username
    i = 1
    while await db.users.find_one({"username": username}):
        i += 1
        username = f"{base}{i}"
    doc = {
        "id": str(uuid.uuid4()),
        "email": email,
        "username": username,
        "name": body.name.strip() or email,
        "password": hash_password(body.password),
        "role": body.role,
        "otp_login": bool(body.otp_login),
        "created_at": now_iso(),
    }
    # Optional explicit permission allowlist (validated against the catalog).
    if body.permissions is not None:
        invalid = sorted({p for p in body.permissions if p not in ALL_GRANTABLE_KEYS})
        if invalid:
            raise HTTPException(status_code=400, detail=f"Invalid permission keys: {invalid}")
        doc["permissions"] = [k for k in ALL_GRANTABLE_KEYS if k in set(body.permissions)]
    await db.users.insert_one(doc)
    return {"id": doc["id"], "email": doc["email"], "username": doc["username"], "name": doc["name"], "role": doc["role"], "otp_login": doc["otp_login"], "permissions": doc.get("permissions"), "created_at": doc["created_at"]}


@api_router.patch("/users/{uid}/otp")
async def set_user_otp(uid: str, body: UserOtpUpdate, admin=Depends(require_users_admin)):
    """Admin toggles email-OTP two-step login ON/OFF for any user."""
    res = await db.users.update_one(
        {"id": uid},
        {"$set": {"otp_login": bool(body.otp_login), "updated_at": now_iso()}},
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    user = await db.users.find_one({"id": uid}, {"_id": 0, "password": 0})
    return user


@api_router.delete("/users/{uid}")
async def delete_user(uid: str, admin=Depends(require_users_admin)):
    if uid == admin["id"]:
        raise HTTPException(status_code=400, detail="You cannot delete your own account")
    res = await db.users.delete_one({"id": uid})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"ok": True, "deleted": uid}


@api_router.post("/users/{uid}/reset-password")
async def reset_user_password(uid: str, body: AdminPasswordReset, admin=Depends(require_users_admin)):
    if len(body.password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    res = await db.users.update_one(
        {"id": uid},
        {"$set": {"password": hash_password(body.password), "updated_at": now_iso()}},
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"ok": True}


@api_router.get("/permissions/catalog")
async def permissions_catalog(admin=Depends(require_admin)):
    """All grantable nav keys + the default allowlist for non-admin users.
    The Admin Users page reads this to render the access matrix."""
    return {
        "all": ALL_PERMISSION_KEYS,
        "default_user": DEFAULT_USER_PERMISSIONS,
        "actions": ACTION_PERMISSION_KEYS,
    }


@api_router.patch("/users/{uid}/permissions")
async def set_user_permissions(uid: str, body: UserPermissionsUpdate, admin=Depends(require_users_admin)):
    """Admin sets a user's explicit nav-key allowlist.
    - Passing `permissions: null` clears the override and falls back to role defaults.
    - Passing an empty list `[]` locks the user out of every nav item.
    - Admins themselves are never gated by stored permissions (front-end + back-end
      both bypass the check for `role == admin`), but the field is still saved
      so an admin can be demoted to user later without losing their custom set.

    Every change is logged in the `permission_audit` collection (who / for whom /
    when / before-after / added / removed) so the trail survives across admins.
    """
    if body.permissions is not None:
        invalid = sorted({p for p in body.permissions if p not in ALL_GRANTABLE_KEYS})
        if invalid:
            raise HTTPException(status_code=400, detail=f"Invalid permission keys: {invalid}")
        # Dedupe + preserve catalog order for stable storage
        ordered = [k for k in ALL_GRANTABLE_KEYS if k in set(body.permissions)]
    else:
        ordered = None
    # Snapshot the previous state so we can diff and audit
    prev_user = await db.users.find_one({"id": uid}, {"_id": 0, "password": 0})
    if not prev_user:
        raise HTTPException(status_code=404, detail="User not found")
    prev_perms = prev_user.get("permissions")
    await db.users.update_one(
        {"id": uid},
        {"$set": {"permissions": ordered, "updated_at": now_iso()}},
    )
    # Diff for the audit log. We compare against the *effective* set the user
    # had before — so admins can see what actually changed for the operator,
    # not just the raw stored field.
    prev_effective = set(prev_perms) if isinstance(prev_perms, list) else set(DEFAULT_USER_PERMISSIONS)
    next_effective = set(ordered) if isinstance(ordered, list) else set(DEFAULT_USER_PERMISSIONS)
    audit_doc = {
        "id": str(uuid.uuid4()),
        "when": now_iso(),
        "actor_id": admin.get("id"),
        "actor_username": admin.get("username") or admin.get("email"),
        "actor_name": admin.get("name"),
        "target_id": uid,
        "target_username": prev_user.get("username") or prev_user.get("email"),
        "target_name": prev_user.get("name"),
        "before": prev_perms,
        "after": ordered,
        "added": sorted(next_effective - prev_effective),
        "removed": sorted(prev_effective - next_effective),
        "kind": "clear" if ordered is None else "set",
    }
    await db.permission_audit.insert_one(audit_doc)
    user = await db.users.find_one({"id": uid}, {"_id": 0, "password": 0})
    return user


@api_router.get("/users/{uid}/permissions/audit")
async def user_permission_audit(uid: str, limit: int = 25, admin=Depends(require_admin)):
    """Return the most recent permission changes for the given user (newest first)."""
    if limit <= 0 or limit > 200:
        limit = 25
    rows = await db.permission_audit.find(
        {"target_id": uid}, {"_id": 0}
    ).sort("when", -1).to_list(limit)
    return {"rows": rows, "count": len(rows)}


@api_router.get("/permissions/audit")
async def permission_audit_all(limit: int = 100, admin=Depends(require_admin)):
    """Global feed of all permission changes (newest first)."""
    if limit <= 0 or limit > 500:
        limit = 100
    rows = await db.permission_audit.find({}, {"_id": 0}).sort("when", -1).to_list(limit)
    return {"rows": rows, "count": len(rows)}


# ======================== Admin: Settings (overdue threshold) ========================
async def _get_settings_doc() -> Dict[str, Any]:
    doc = await db.settings.find_one({"id": "global"}, {"_id": 0})
    if not doc:
        doc = {"id": "global", "overdue_days": 15, "edit_window_days": 3}
        await db.settings.insert_one({**doc, "updated_at": now_iso()})
    # Back-fill defaults for older docs.
    if "edit_window_days" not in doc:
        doc["edit_window_days"] = 3
    return doc


@api_router.get("/settings")
async def get_settings(user=Depends(get_current_user)):
    return await _get_settings_doc()


@api_router.patch("/settings")
async def update_settings(body: SettingsUpdate, admin=Depends(require_admin)):
    upd: Dict[str, Any] = {"updated_at": now_iso()}
    if body.overdue_days is not None:
        if body.overdue_days <= 0 or body.overdue_days > 365:
            raise HTTPException(status_code=400, detail="overdue_days must be between 1 and 365")
        upd["overdue_days"] = int(body.overdue_days)
    if body.edit_window_days is not None:
        # 0 = users cannot edit at all (admin-only). Cap at 365 to avoid silliness.
        if body.edit_window_days < 0 or body.edit_window_days > 365:
            raise HTTPException(status_code=400, detail="edit_window_days must be between 0 and 365")
        upd["edit_window_days"] = int(body.edit_window_days)
    if len(upd) == 1:
        # Only updated_at present — nothing to update.
        return await _get_settings_doc()
    await db.settings.update_one(
        {"id": "global"},
        {"$set": upd},
        upsert=True,
    )
    return await _get_settings_doc()


# ======================== Backup & Restore ========================
class BackupSettingsUpdate(BaseModel):
    enabled: Optional[bool] = None
    gmail_user: Optional[str] = None
    gmail_app_password: Optional[str] = None
    send_to: Optional[str] = None
    schedule_hour: Optional[int] = Field(default=None, ge=0, le=23)
    schedule_minute: Optional[int] = Field(default=None, ge=0, le=59)


@api_router.get("/admin/backup/settings")
async def get_backup_settings(admin=Depends(require_admin)):
    doc = await backup_mod.get_settings(db)
    return backup_mod._mask(doc)


@api_router.patch("/admin/backup/settings")
async def patch_backup_settings(body: BackupSettingsUpdate, admin=Depends(require_admin)):
    return await backup_mod.update_settings(db, body.dict(exclude_unset=True))


@api_router.post("/admin/backup/run")
async def run_backup_now(admin=Depends(require_admin)):
    """Trigger an immediate backup-and-email run."""
    return await backup_mod.run_backup_now(db)


@api_router.post("/admin/backup/restore")
async def restore_backup(file: UploadFile = File(...), admin=Depends(require_admin)):
    """Replace the entire database with the contents of the uploaded ZIP.
    Protected collections (e.g. backup settings themselves) are kept intact.
    """
    blob = await file.read()
    if not blob:
        raise HTTPException(status_code=400, detail="Empty upload")
    if len(blob) > 200 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Backup file too large (max 200 MB)")
    try:
        return await backup_mod.restore_from_zip(db, blob)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))



# ======================== Products ========================
@api_router.get("/products")
async def list_products(user=Depends(get_current_user)):
    items = await db.products.find({}, {"_id": 0}).sort("name", 1).to_list(1000)
    return items


@api_router.post("/products")
async def create_product(body: ProductIn, user=Depends(require_admin)):
    if await db.products.find_one({"name": body.name}):
        raise HTTPException(status_code=400, detail="Product already exists")
    doc = {"id": str(uuid.uuid4()), **body.model_dump(), "created_at": now_iso()}
    await db.products.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.patch("/products/{pid}")
async def update_product(pid: str, body: ProductUpdate, user=Depends(require_action("edit:products"))):
    update = {k: v for k, v in body.model_dump().items() if v is not None}
    if not update:
        raise HTTPException(status_code=400, detail="No fields to update")
    # Server-side bag-limit sanity (client also validates)
    mn = update.get("min_per_bag")
    mx = update.get("max_per_bag")
    if mn is not None and mn <= 0:
        raise HTTPException(status_code=400, detail="min_per_bag must be > 0")
    if mx is not None and mx <= 0:
        raise HTTPException(status_code=400, detail="max_per_bag must be > 0")
    if mn is not None and mx is not None and mn > mx:
        raise HTTPException(status_code=400, detail="min_per_bag cannot exceed max_per_bag")
    res = await db.products.update_one({"id": pid}, {"$set": update})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    return await db.products.find_one({"id": pid}, {"_id": 0})


@api_router.delete("/products/{pid}")
async def delete_product(pid: str, user=Depends(require_action("delete:products"))):
    prod = await db.products.find_one({"id": pid})
    if not prod:
        raise HTTPException(status_code=404, detail="Product not found")
    # Block deletion if any item SKU or order still references it
    item_refs = await db.items.count_documents({"product_id": pid})
    if item_refs > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete: {item_refs} item SKU(s) still mapped to this product")
    order_refs = await db.orders.count_documents({"items.product_id": pid})
    if order_refs > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete: {order_refs} order(s) still reference this product")
    await db.products.delete_one({"id": pid})
    return {"ok": True, "deleted": pid}


# ======================== Item SKUs ========================
@api_router.get("/items")
async def list_items(product_id: Optional[str] = None, customer_id: Optional[str] = None,
                     user=Depends(get_current_user)):
    q = {}
    if product_id:
        q["product_id"] = product_id
    # When a customer is in context, hide items that the admin has blocked for
    # that party — those items must never be orderable / dispatchable for the
    # party (product-side or search-side, no matter where the list is used).
    if customer_id:
        cust = await db.customers.find_one({"id": customer_id}, {"_id": 0, "blocked_items": 1})
        blocked = list((cust or {}).get("blocked_items") or [])
        if blocked:
            q["id"] = {"$nin": blocked}
    items = await db.items.find(q, {"_id": 0}).sort("name", 1).to_list(5000)
    return items


@api_router.get("/items/search")
async def search_items(q: str = "", product_id: Optional[str] = None, limit: int = 15,
                       customer_id: Optional[str] = None,
                       user=Depends(get_current_user)):
    q = q.strip()
    filt = {}
    if product_id:
        filt["product_id"] = product_id
    # Same per-party blocklist filter as /items — the SKU search dropdown
    # used across New Order / Dispatch / Estimate / edit dialogs will never
    # surface a blocked SKU for that party.
    if customer_id:
        cust = await db.customers.find_one({"id": customer_id}, {"_id": 0, "blocked_items": 1})
        blocked = list((cust or {}).get("blocked_items") or [])
        if blocked:
            filt["id"] = {"$nin": blocked}
    all_items = await db.items.find(filt, {"_id": 0}).to_list(5000)
    if not q:
        return all_items[:limit]
    names = [it["name"] for it in all_items]
    matches = rf_process.extract(q, names, scorer=fuzz.WRatio,
                                  processor=rf_utils.default_process, limit=limit)
    out = []
    seen = set()
    short_q = len(q) < 5
    for name, score, idx in matches:
        if score < (35 if short_q else 45):
            continue
        it = all_items[idx]
        if it["id"] in seen:
            continue
        seen.add(it["id"])
        out.append({**it, "match_score": score})
    return out


@api_router.get("/items/{iid}")
async def get_item(iid: str, user=Depends(get_current_user)):
    it = await db.items.find_one({"id": iid}, {"_id": 0})
    if not it:
        raise HTTPException(status_code=404, detail="Item not found")
    return it


@api_router.post("/items")
async def create_item(body: ItemCreate, admin=Depends(require_admin)):
    """Admin: create a new SKU under an existing master product."""
    prod = await db.products.find_one({"id": body.product_id}, {"_id": 0})
    if not prod:
        raise HTTPException(status_code=404, detail="Product not found for product_id")
    name = body.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Item name required")
    if await db.items.find_one({"name": name, "product_id": body.product_id}):
        raise HTTPException(status_code=400, detail="An item with this name already exists under this product")
    doc = {
        "id": str(uuid.uuid4()),
        "name": name,
        "category": prod["name"],
        "product_id": body.product_id,
        "product_name": prod["name"],
        "created_at": now_iso(),
    }
    if body.min_per_bag is not None or body.max_per_bag is not None:
        mn = body.min_per_bag if body.min_per_bag is not None else body.max_per_bag
        mx = body.max_per_bag if body.max_per_bag is not None else body.min_per_bag
        if mn is None or mx is None or mn <= 0 or mx <= 0 or mn > mx:
            raise HTTPException(status_code=400, detail="Invalid bag override")
        doc["min_per_bag"] = mn
        doc["max_per_bag"] = mx
    await db.items.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.patch("/items/{iid}")
async def update_item(iid: str, body: ItemEdit, user=Depends(require_action("edit:products"))):
    """Admin: update an item SKU. Accepts any subset of name / product_id /
    bag-override fields. Preserves prior behaviour: passing only min_per_bag +
    max_per_bag still sets the bag override unchanged."""
    existing = await db.items.find_one({"id": iid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Item not found")
    update: Dict[str, Any] = {}
    if body.name is not None:
        nm = body.name.strip()
        if not nm:
            raise HTTPException(status_code=400, detail="Item name cannot be empty")
        # uniqueness within same product
        dup = await db.items.find_one({"name": nm, "product_id": body.product_id or existing["product_id"], "id": {"$ne": iid}})
        if dup:
            raise HTTPException(status_code=400, detail="Another item with this name exists under the same product")
        update["name"] = nm
    if body.product_id is not None and body.product_id != existing.get("product_id"):
        prod = await db.products.find_one({"id": body.product_id}, {"_id": 0})
        if not prod:
            raise HTTPException(status_code=404, detail="Product not found for product_id")
        update["product_id"] = body.product_id
        update["product_name"] = prod["name"]
        update["category"] = prod["name"]
    if body.min_per_bag is not None and body.max_per_bag is not None:
        if body.min_per_bag <= 0 or body.max_per_bag <= 0:
            raise HTTPException(status_code=400, detail="Bag values must be > 0")
        if body.min_per_bag > body.max_per_bag:
            raise HTTPException(status_code=400, detail="min_per_bag cannot exceed max_per_bag")
        update["min_per_bag"] = body.min_per_bag
        update["max_per_bag"] = body.max_per_bag
    if not update:
        raise HTTPException(status_code=400, detail="No fields to update")
    await db.items.update_one({"id": iid}, {"$set": update})
    return {"ok": True}


@api_router.delete("/items/{iid}")
async def delete_item(iid: str, admin=Depends(require_action("delete:products"))):
    """Admin: delete an item SKU. Blocked if any order line references it."""
    existing = await db.items.find_one({"id": iid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Item not found")
    refs = await db.orders.count_documents({"items.item_id": iid})
    if refs > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete: {refs} order(s) reference this SKU")
    await db.items.delete_one({"id": iid})
    return {"ok": True, "deleted": iid}


@api_router.delete("/items/{iid}/bag-override")
async def clear_item_bag_override(iid: str, user=Depends(require_action("edit:products"))):
    """Remove the per-SKU bag override so the item falls back to its master
    product's bag limits."""
    existing = await db.items.find_one({"id": iid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Item not found")
    await db.items.update_one(
        {"id": iid},
        {"$unset": {"min_per_bag": "", "max_per_bag": ""}},
    )
    return {"ok": True}


# ======================== Customers ========================
@api_router.get("/customers")
async def list_customers(user=Depends(get_current_user)):
    if is_blank_view(user):
        return []
    items = await db.customers.find({}, {"_id": 0}).sort("name", 1).to_list(2000)
    return items


@api_router.get("/customers/search")
async def search_customers(q: str = "", user=Depends(get_current_user)):
    if is_blank_view(user):
        return []
    q = q.strip()
    if not q:
        return []
    all_c = await db.customers.find({}, {"_id": 0}).to_list(5000)
    names = [c["name"] for c in all_c]
    matches = rf_process.extract(q, names, scorer=fuzz.WRatio, processor=rf_utils.default_process, limit=8)
    out = []
    seen = set()
    short_q = len(q) < 6
    for name, score, idx in matches:
        if score < (40 if short_q else 50):
            continue
        cust = all_c[idx]
        if cust["id"] in seen:
            continue
        seen.add(cust["id"])
        out.append({**cust, "match_score": score})
    return out


@api_router.post("/customers")
async def create_customer(body: CustomerIn, user=Depends(require_admin)):
    doc = {"id": str(uuid.uuid4()), **body.model_dump(), "created_at": now_iso()}
    await db.customers.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.patch("/customers/{cid}")
async def update_customer(cid: str, body: CustomerUpdate, user=Depends(require_action("edit:customers"))):
    update = {k: v for k, v in body.model_dump().items() if v is not None}
    # Non-admin users can update preferences + the operational labels they
    # actually use on the dispatch floor (private mark, transport). Other
    # fields (name, phone, address, price_list_id, blocked_items, etc.)
    # remain admin-only.
    if user["role"] != "admin":
        allowed = {"preferences", "private_mark", "transport_name"}
        update = {k: v for k, v in update.items() if k in allowed}
    if not update:
        raise HTTPException(status_code=400, detail="No fields to update")
    # Enforce Bill-Number mode: a party configured for Bill Number must NEVER
    # have a private mark auto-saved/reused. Whether the party is already in
    # bill mode or is being switched into it now, any incoming private_mark is
    # forced empty so the software always asks for a fresh Bill Number at
    # dispatch (never silently remembers one).
    # NOTE: use a projection-safe existence check. When the customer doc
    # exists but does NOT contain the `bill_number_mode` field (older /
    # restored records), the projection returns an empty dict `{}` which is
    # falsy — a plain `if not existing` would wrongly 404 a real customer.
    existing = await db.customers.find_one({"id": cid}, {"_id": 0, "bill_number_mode": 1})
    if existing is None:
        raise HTTPException(status_code=404, detail="Customer not found")
    incoming_mode = update.get("bill_number_mode")
    effective_mode = incoming_mode if incoming_mode is not None else bool(existing.get("bill_number_mode"))
    if effective_mode and "private_mark" in update:
        update["private_mark"] = ""
    res = await db.customers.update_one({"id": cid}, {"$set": update})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    return await db.customers.find_one({"id": cid}, {"_id": 0})


class BlockedItemsIn(BaseModel):
    item_ids: List[str] = []


@api_router.get("/customers/{cid}/blocked-items")
async def get_customer_blocked_items(cid: str, user=Depends(get_current_user)):
    """List the item SKUs blocked for this party. Returns the full item docs
    (id, name, product_name, product_id) so the UI can render friendly
    chips without a second round-trip.
    """
    # `blocked_items` may be absent on older / restored customer docs, in
    # which case the projection returns an empty dict `{}` (falsy) even
    # though the customer exists — so check for None explicitly.
    cust = await db.customers.find_one({"id": cid}, {"_id": 0, "blocked_items": 1})
    if cust is None:
        raise HTTPException(status_code=404, detail="Customer not found")
    ids = list(cust.get("blocked_items") or [])
    if not ids:
        return {"item_ids": [], "items": []}
    items = await db.items.find({"id": {"$in": ids}}, {"_id": 0}).to_list(2000)
    # Preserve order given in blocked_items where possible
    by_id = {it["id"]: it for it in items}
    ordered = [by_id[i] for i in ids if i in by_id]
    return {"item_ids": ids, "items": ordered}


@api_router.put("/customers/{cid}/blocked-items")
async def set_customer_blocked_items(cid: str, body: BlockedItemsIn, admin=Depends(require_action("edit:customers"))):
    """Admin: replace the full list of blocked item_ids for this party.
    Any SKUs in this list will never appear in the item search dropdown
    for this customer and cannot be placed on a new order or dispatch.
    """
    cust = await db.customers.find_one({"id": cid}, {"_id": 0})
    if not cust:
        raise HTTPException(status_code=404, detail="Customer not found")
    ids = [str(x) for x in (body.item_ids or []) if x]
    # Dedupe while preserving order
    seen = set()
    clean: List[str] = []
    for x in ids:
        if x in seen:
            continue
        seen.add(x)
        clean.append(x)
    # Validate the ids exist as real items so bad payloads don't silently rot
    if clean:
        existing = await db.items.find({"id": {"$in": clean}}, {"_id": 0, "id": 1}).to_list(5000)
        valid = {it["id"] for it in existing}
        clean = [i for i in clean if i in valid]
    await db.customers.update_one({"id": cid}, {"$set": {"blocked_items": clean}})
    items = []
    if clean:
        docs = await db.items.find({"id": {"$in": clean}}, {"_id": 0}).to_list(2000)
        by_id = {it["id"]: it for it in docs}
        items = [by_id[i] for i in clean if i in by_id]
    return {"item_ids": clean, "items": items}


@api_router.delete("/customers/{cid}")
async def delete_customer(cid: str, admin=Depends(require_action("delete:customers"))):
    """Admin: delete a customer. Blocked if any order references this party."""
    existing = await db.customers.find_one({"id": cid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Customer not found")
    refs = await db.orders.count_documents({"customer_id": cid})
    if refs > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete: {refs} order(s) reference this customer")
    await db.customers.delete_one({"id": cid})
    return {"ok": True, "deleted": cid}


@api_router.post("/customers/bulk-delete")
async def bulk_delete_customers(body: CustomerBulkDeleteIn, admin=Depends(require_action("delete:customers"))):
    """Admin: delete many customers in a single call. The whole call is
    rejected if any of the supplied ids are referenced by an order — the
    response lists the blocking parties so the operator can review."""
    ids = [i for i in (body.ids or []) if i]
    if not ids:
        raise HTTPException(status_code=400, detail="No customer ids supplied")
    # Validate every id exists
    existing = await db.customers.find({"id": {"$in": ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(len(ids))
    existing_ids = {c["id"] for c in existing}
    missing = [i for i in ids if i not in existing_ids]
    if missing:
        raise HTTPException(status_code=404, detail=f"{len(missing)} customer(s) not found")
    # Find any blocking references in orders
    blockers = await db.orders.aggregate([
        {"$match": {"customer_id": {"$in": ids}}},
        {"$group": {"_id": "$customer_id", "count": {"$sum": 1}}},
    ]).to_list(len(ids))
    if blockers:
        name_map = {c["id"]: c["name"] for c in existing}
        details = ", ".join(f"{name_map.get(b['_id'], b['_id'])} ({b['count']})" for b in blockers)
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete: {len(blockers)} customer(s) have orders — {details}",
        )
    res = await db.customers.delete_many({"id": {"$in": ids}})
    return {"ok": True, "deleted": res.deleted_count, "ids": ids}


@api_router.get("/customers/import/template")
async def customer_import_template(admin=Depends(require_admin)):
    """Download a blank Excel template for bulk customer import."""
    from openpyxl import Workbook
    from fastapi.responses import StreamingResponse
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    headers = ["name", "phone", "address", "city", "location", "transport_name", "price_list"]
    ws.append(headers)
    # one example row to make the format obvious (will be skipped if empty)
    ws.append(["EXAMPLE PARTY PVT LTD", "9876543210", "12, Industrial Area", "Indore", "Sapna Sangeeta", "DTDC", ""])
    widths = [40, 16, 36, 18, 22, 22, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="customers_import_template.xlsx"'},
    )


@api_router.post("/customers/import")
async def import_customers(file: UploadFile = File(...), admin=Depends(require_admin)):
    """Admin: bulk import customers from an Excel file.

    Columns recognised (case-insensitive header row required):
      name (or party_name / customer_name), phone (or mobile / contact),
      address, city, location, transport_name (or transport),
      price_list (or pricelist / price_list_name).

    Behaviour (partial-success):
    - Rows with a name duplicating an existing customer or another row in
      the same file are SKIPPED (not failed). The response lists them so
      the operator can review.
    - Rows whose `price_list` doesn't match a known list are SKIPPED.
    - Empty rows are silently skipped.
    - At least one valid row inserts; the response returns
      `{imported, skipped, skipped_reasons}` so the UI can summarise.
    """
    from openpyxl import load_workbook
    blob = await file.read()
    if not blob:
        raise HTTPException(status_code=400, detail="Empty file")
    try:
        wb = load_workbook(io.BytesIO(blob), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel: {e}")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Empty sheet")
    # Parse header — tolerate casing, surrounding spaces, hyphens, and
    # a handful of common column-name synonyms.
    HEADER_ALIASES: Dict[str, str] = {
        "name": "name", "party_name": "name", "party": "name", "customer_name": "name", "customer": "name",
        "phone": "phone", "mobile": "phone", "contact": "phone", "phone_no": "phone", "phone_number": "phone", "mobile_no": "phone",
        "address": "address", "addr": "address",
        "city": "city",
        "location": "location", "area": "location",
        "transport_name": "transport_name", "transport": "transport_name", "transporter": "transport_name",
        "price_list": "price_list", "pricelist": "price_list", "price_list_name": "price_list", "price": "price_list",
    }
    header_row = rows[0]
    header_map: Dict[str, int] = {}
    for idx, cell_v in enumerate(header_row):
        if cell_v is None:
            continue
        raw = str(cell_v).strip().lower().replace("-", "_").replace(" ", "_")
        if not raw:
            continue
        norm = HEADER_ALIASES.get(raw)
        if norm and norm not in header_map:
            header_map[norm] = idx
    if "name" not in header_map:
        raise HTTPException(
            status_code=400,
            detail='Excel must have a "name" column in the first row (also accepted: party_name, customer_name).',
        )

    def cell(r, key: str) -> str:
        idx = header_map.get(key)
        if idx is None or idx >= len(r):
            return ""
        v = r[idx]
        if v is None:
            return ""
        return str(v).strip()

    # Pre-load existing customers and price lists for duplicate detection.
    # A "duplicate" now requires the SAME name AND the SAME city AND the SAME
    # address — same shop name at a different location is treated as a
    # different customer and imported as such. Phone is NOT part of the
    # dedupe key any more either; many shops legitimately share a single
    # owner-mobile across two branches.
    existing = await db.customers.find(
        {}, {"_id": 0, "id": 1, "name": 1, "phone": 1, "city": 1, "address": 1},
    ).to_list(20000)

    def _dkey(name: str, city: str, address: str) -> str:
        """Build the dedupe key from name+city+address, all lower-cased and
        whitespace-collapsed so 'Ram Auto / Delhi / Plot-12' and
        'ram   auto / delhi / plot-12' map to the same key."""
        def n(s: str) -> str:
            return " ".join((s or "").strip().lower().split())
        return f"{n(name)}||{n(city)}||{n(address)}"

    existing_keys: Dict[str, str] = {
        _dkey(c.get("name") or "", c.get("city") or "", c.get("address") or ""): (c.get("name") or "")
        for c in existing
        if c.get("name")
    }
    price_lists = await db.price_lists.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(1000)
    pl_by_name = {pl["name"].strip().lower(): pl["id"] for pl in price_lists}

    parsed_rows: List[Dict[str, Any]] = []
    skipped: List[Dict[str, str]] = []
    seen_key_in_file: Dict[str, int] = {}

    for row_idx, r in enumerate(rows[1:], start=2):  # excel row #
        if not r or all(c is None or (isinstance(c, str) and not c.strip()) for c in r):
            continue
        name = cell(r, "name")
        if not name:
            continue
        phone = cell(r, "phone")
        city = cell(r, "city")
        address = cell(r, "address")
        key = _dkey(name, city, address)
        if key in existing_keys:
            skipped.append({
                "row": str(row_idx),
                "name": name,
                "reason": "same name + city + address already exists in customer list",
            })
            continue
        if key in seen_key_in_file:
            skipped.append({
                "row": str(row_idx),
                "name": name,
                "reason": f"duplicate of row {seen_key_in_file[key]} in this file (same name + city + address)",
            })
            continue

        pl_name = cell(r, "price_list")
        pl_id: Optional[str] = None
        if pl_name:
            pl_id = pl_by_name.get(pl_name.lower())
            if not pl_id:
                skipped.append({"row": str(row_idx), "name": name, "reason": f"unknown price list '{pl_name}'"})
                continue

        seen_key_in_file[key] = row_idx
        parsed_rows.append({
            "id": str(uuid.uuid4()),
            "name": name,
            "phone": phone,
            "address": address,
            "city": city,
            "location": cell(r, "location"),
            "transport_name": cell(r, "transport_name"),
            "price_list_id": pl_id,
            "preferences": {},
            "created_at": now_iso(),
        })

    if not parsed_rows and not skipped:
        raise HTTPException(status_code=400, detail="No customer rows found in the file")
    if not parsed_rows:
        # Everything was skipped — return a 200 with the details so the
        # frontend can show why nothing was imported (instead of a generic
        # 400 that hides the per-row reasons).
        return {"imported": 0, "skipped": len(skipped), "skipped_reasons": skipped[:200]}

    await db.customers.insert_many(parsed_rows)
    return {"imported": len(parsed_rows), "skipped": len(skipped), "skipped_reasons": skipped[:200]}


# ======================== Orders ========================
async def _persist_customer_prefs(customer_id: str, items: List[OrderItemIn]):
    """Memorize variant choices on customer for future orders."""
    products = await db.products.find({}, {"_id": 0}).to_list(1000)
    pmap = {p["name"]: p for p in products}
    prefs_update = {}
    for it in items:
        prod = pmap.get(it.product_name)
        if not prod:
            continue
        vf = prod.get("variant_field")
        if not vf:
            continue
        if it.product_name in ("Center Stand with Kit", "Center Stand without Kit"):
            prefs_update["center_stand_kit"] = "With Kit" if "with Kit" in it.product_name else "Without Kit"
        elif it.variant:
            prefs_update[vf] = it.variant
    if prefs_update:
        cust = await db.customers.find_one({"id": customer_id}, {"_id": 0})
        if cust:
            merged = {**(cust.get("preferences") or {}), **prefs_update}
            await db.customers.update_one({"id": customer_id}, {"$set": {"preferences": merged}})


@api_router.post("/orders")
async def create_order(body: OrderIn, user=Depends(get_current_user)):
    cust = await db.customers.find_one({"id": body.customer_id}, {"_id": 0})
    if not cust:
        raise HTTPException(status_code=404, detail="Customer not found")

    # Strict item-wise validation: every line must have a valid item_id (SKU)
    if not body.items:
        raise HTTPException(status_code=400, detail="Order must contain at least one item")
    item_ids = [it.item_id for it in body.items]
    found = await db.items.find({"id": {"$in": item_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(1000)
    found_by_id = {f["id"]: f for f in found}
    missing = [iid for iid in item_ids if iid not in found_by_id]
    if missing:
        raise HTTPException(status_code=400, detail=f"Unknown item_id(s): {missing}")

    # Per-party block-list enforcement: reject the whole order if any line
    # references a SKU the admin has blocked for this customer. The item
    # search dropdown already hides these; this is the last-line guard so
    # a stale client / bulk import / API caller can't sneak them through.
    blocked_ids = set((cust.get("blocked_items") or []))
    if blocked_ids:
        offending = [iid for iid in item_ids if iid in blocked_ids]
        if offending:
            names = [found_by_id[i].get("name", i) for i in offending]
            raise HTTPException(
                status_code=400,
                detail=f"These items are blocked for {cust.get('name') or 'this party'}: {', '.join(names)}",
            )

    # If clear_previous_pending: mark all this customer's pending orders as Cleared
    if body.clear_previous_pending:
        await db.orders.update_many(
            {"customer_id": body.customer_id, "status": "Pending"},
            {"$set": {"status": "Cleared", "updated_at": now_iso()}},
        )

    # If merge_with_pending: append items to most recent pending order
    if body.merge_with_pending:
        existing = await db.orders.find_one(
            {"customer_id": body.customer_id, "status": "Pending"},
            sort=[("created_at", -1)],
        )
        if existing:
            new_items = existing.get("items", []) + [it.model_dump() for it in body.items]
            await db.orders.update_one(
                {"id": existing["id"]},
                {"$set": {"items": new_items, "updated_at": now_iso(),
                          "delivery_date": body.delivery_date or existing.get("delivery_date"),
                          "notes": (existing.get("notes", "") + " | " + (body.notes or "")).strip(" |")}},
            )
            await _persist_customer_prefs(body.customer_id, body.items)
            return await db.orders.find_one({"id": existing["id"]}, {"_id": 0})

    doc = {
        "id": str(uuid.uuid4()),
        "customer_id": body.customer_id,
        "customer_name": cust["name"],
        "items": [it.model_dump() for it in body.items],
        "order_date": body.order_date or now_iso(),
        "delivery_date": body.delivery_date,
        "status": "Pending",
        "notes": body.notes or "",
        "created_by": user["email"],
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    await db.orders.insert_one(doc)
    await _persist_customer_prefs(body.customer_id, body.items)
    doc.pop("_id", None)
    return doc


@api_router.get("/orders")
async def list_orders(status_filter: Optional[str] = None, user=Depends(get_current_user)):
    if is_blank_view(user):
        return []
    q = {}
    if status_filter == "Dispatched":
        # "Dispatched" view must include PARTIALLY dispatched orders too —
        # those keep status "Pending" (remainder still open) so they would
        # otherwise never appear here, hiding the qty already shipped.
        single_ids = await db.dispatches.distinct("order_id")
        multi_ids = await db.dispatches.distinct("order_ids")
        disp_ids = [i for i in set([*(single_ids or []), *(multi_ids or [])]) if i]
        q = {"$or": [{"status": {"$in": ["Dispatched", "Cleared"]}}, {"id": {"$in": disp_ids}}]}
    elif status_filter:
        q["status"] = status_filter
    items = await db.orders.find(q, {"_id": 0}).sort("created_at", -1).to_list(2000)
    # Annotate overdue flag + days_open for Pending orders, using admin-set threshold
    settings = await _get_settings_doc()
    threshold = int(settings.get("overdue_days", 15))
    now = datetime.now(timezone.utc)

    # Customer location cache — the dispatch UI shows the party's city /
    # location right next to the customer name so operators can group
    # orders by area at a glance. One projection call for all referenced
    # customers keeps this O(1) extra query regardless of order count.
    cust_ids = list({o.get("customer_id") for o in items if o.get("customer_id")})
    cust_loc: Dict[str, Dict[str, str]] = {}
    if cust_ids:
        async for c in db.customers.find(
            {"id": {"$in": cust_ids}},
            {"_id": 0, "id": 1, "city": 1, "location": 1, "address": 1},
        ):
            cust_loc[c["id"]] = {
                "city": c.get("city") or "",
                "location": c.get("location") or "",
                "address": c.get("address") or "",
            }

    # Dispatched-items summary — a fully dispatched order's `items` list is
    # emptied (remaining pending = 0), so the list view would show nothing
    # under a "Dispatched" order. Aggregate what was actually shipped from the
    # dispatches collection so the UI can show "what was dispatched".
    order_ids = [o["id"] for o in items]
    disp_map: Dict[str, Dict[str, Dict[str, Any]]] = {}
    # Per-slip dispatch history per order → powers the "All Status" brief
    # (each shipment shows its date AND slip number). oid → list[{date, slip_no, items}]
    disp_slips: Dict[str, List[Dict[str, Any]]] = {}
    # Track which dispatch ids are already linked to each order so the
    # inference step below doesn't attach a slip that's already shown.
    linked_disp_ids: Dict[str, set] = {}
    if order_ids:
        oid_set = set(order_ids)
        async for d in db.dispatches.find(
            {"$or": [{"order_id": {"$in": order_ids}}, {"order_ids": {"$in": order_ids}}]},
            {"_id": 0, "id": 1, "order_id": 1, "order_ids": 1, "items": 1,
             "dispatched_at": 1, "last_dispatched_at": 1, "slip_no": 1},
        ):
            targets = []
            if d.get("order_id") in oid_set:
                targets = [d["order_id"]]
            else:
                targets = [oid for oid in (d.get("order_ids") or []) if oid in oid_set]
            # Resolve the dispatch day (YYYY-MM-DD) for the brief.
            dts = d.get("dispatched_at") or d.get("last_dispatched_at")
            day = ""
            if dts:
                try:
                    dd = datetime.fromisoformat(str(dts).replace("Z", "+00:00"))
                    day = dd.date().isoformat()
                except Exception:
                    day = str(dts)[:10]
            slip_items = [{
                "item_id": it.get("item_id"),
                "item_name": it.get("item_name"),
                "product_name": it.get("product_name"),
                "variant": it.get("variant"),
                "quantity": int(it.get("quantity") or 0),
            } for it in (d.get("items") or [])]
            for oid in targets:
                bucket = disp_map.setdefault(oid, {})
                for it in slip_items:
                    key = it.get("item_id") or it.get("item_name") or ""
                    row = bucket.setdefault(key, {
                        "item_id": it.get("item_id"),
                        "item_name": it.get("item_name"),
                        "product_name": it.get("product_name"),
                        "variant": it.get("variant"),
                        "quantity": 0,
                    })
                    row["quantity"] += it["quantity"]
                disp_slips.setdefault(oid, []).append({
                    "date": day, "slip_no": d.get("slip_no"), "items": slip_items,
                })
                linked_disp_ids.setdefault(oid, set()).add(d.get("id"))

    # ── Discrepancy detection ────────────────────────────────────────────
    # Catch the case where goods were DISPATCHED before the matching order
    # was actually punched in (a back-dated order entered after the slip).
    # For each still-Pending order we look for a dispatch to the SAME party,
    # sharing at least one SKU, that is NOT linked to this order and whose
    # dispatch date PRE-DATES the order's entry (created_at). We pull every
    # dispatch for the referenced customers once and match in-memory.
    def _parse_iso(s):
        if not s:
            return None
        try:
            d = datetime.fromisoformat(str(s).replace("Z", "+00:00")) if isinstance(s, str) else s
            if d.tzinfo is None:
                d = d.replace(tzinfo=timezone.utc)
            return d
        except Exception:
            return None

    cust_disp: Dict[str, List[Dict[str, Any]]] = {}
    if cust_ids:
        async for d in db.dispatches.find(
            {"customer_id": {"$in": cust_ids}},
            {"_id": 0, "id": 1, "slip_no": 1, "order_id": 1, "order_ids": 1,
             "customer_id": 1, "items": 1, "dispatched_at": 1},
        ):
            cust_disp.setdefault(d.get("customer_id") or "", []).append(d)

    for o in items:
        days_open = None
        ref = o.get("order_date") or o.get("created_at")
        if ref:
            try:
                dt = datetime.fromisoformat(ref.replace("Z", "+00:00")) if isinstance(ref, str) else ref
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=timezone.utc)
                days_open = (now - dt).days
            except Exception:
                days_open = None
        o["days_open"] = days_open
        o["is_overdue"] = bool(o.get("status") == "Pending" and days_open is not None and days_open >= threshold)
        loc = cust_loc.get(o.get("customer_id") or "", {})
        o["customer_city"] = loc.get("city", "")
        o["customer_location"] = loc.get("location", "")
        o["customer_address"] = loc.get("address", "")
        o["dispatched_items"] = list(disp_map.get(o["id"], {}).values())
        # Per-slip brief: [{date, slip_no, items:[...]}, ...] sorted oldest→newest.
        slips = sorted(disp_slips.get(o["id"], []), key=lambda s: (s.get("date") or ""))
        o["dispatch_summary"] = slips
        o["dispatch_inferred"] = False

        # ── Repair display for "orphan" dispatched orders ─────────────────
        # Some orders were marked Dispatched but their slip isn't linked via
        # order_id/order_ids (e.g. dispatched off-order or on a merged slip),
        # so the brief was empty → "no slip on record". Infer the correct
        # slip(s) from this customer's dispatches by matching the order's SKUs
        # (uses original_items when items were emptied on dispatch).
        if not slips and o.get("status") in ("Dispatched", "Cleared"):
            match_items = o.get("items") or o.get("original_items") or []
            want_ids = {it.get("item_id") for it in match_items if it.get("item_id")}
            want_names = {(it.get("item_name") or "").strip().lower()
                          for it in match_items if it.get("item_name")}
            already = linked_disp_ids.get(o["id"], set())
            inferred = []
            for d in cust_disp.get(o.get("customer_id") or "", []):
                if d.get("id") in already:
                    continue
                d_items = [it for it in (d.get("items") or [])
                           if (it.get("item_id") and it.get("item_id") in want_ids)
                           or ((it.get("item_name") or "").strip().lower() in want_names)]
                if not d_items:
                    continue
                dts = d.get("dispatched_at") or d.get("last_dispatched_at")
                day = ""
                if dts:
                    try:
                        day = datetime.fromisoformat(str(dts).replace("Z", "+00:00")).date().isoformat()
                    except Exception:
                        day = str(dts)[:10]
                inferred.append({
                    "date": day,
                    "slip_no": d.get("slip_no"),
                    "items": [{
                        "item_id": it.get("item_id"),
                        "item_name": it.get("item_name"),
                        "product_name": it.get("product_name"),
                        "variant": it.get("variant"),
                        "quantity": int(it.get("quantity") or 0),
                    } for it in d_items],
                })
            if inferred:
                inferred.sort(key=lambda s: (s.get("date") or ""))
                o["dispatch_summary"] = inferred
                o["dispatch_inferred"] = True

        # Attach a discrepancy suggestion if one is found (and not dismissed).
        o["discrepancy"] = None
        if o.get("status") == "Pending" and not o.get("discrepancy_dismissed"):
            entered = _parse_iso(o.get("created_at"))
            # Set of this order's SKUs by id and by normalised name.
            oid_ids = {it.get("item_id") for it in (o.get("items") or []) if it.get("item_id")}
            oid_names = {(it.get("item_name") or "").strip().lower()
                         for it in (o.get("items") or []) if it.get("item_name")}
            best = None
            best_dt = None
            for d in cust_disp.get(o.get("customer_id") or "", []):
                # Skip dispatches already linked to THIS order.
                if d.get("order_id") == o["id"] or o["id"] in (d.get("order_ids") or []):
                    continue
                disp_dt = _parse_iso(d.get("dispatched_at"))
                if not disp_dt or not entered:
                    continue
                # Core signal: goods shipped BEFORE this order was entered.
                if not (disp_dt < entered):
                    continue
                # Require at least one shared SKU.
                matched = []
                for it in (d.get("items") or []):
                    iid = it.get("item_id")
                    inm = (it.get("item_name") or "").strip().lower()
                    if (iid and iid in oid_ids) or (inm and inm in oid_names):
                        matched.append({
                            "item_name": it.get("item_name"),
                            "product_name": it.get("product_name"),
                            "variant": it.get("variant"),
                            "quantity": int(it.get("quantity") or 0),
                        })
                if not matched:
                    continue
                # Prefer the most recent qualifying dispatch.
                if best_dt is None or disp_dt > best_dt:
                    best_dt = disp_dt
                    best = {
                        "dispatch_id": d.get("id"),
                        "slip_no": d.get("slip_no"),
                        "dispatched_at": d.get("dispatched_at"),
                        "order_date": o.get("order_date"),
                        "entered_at": o.get("created_at"),
                        "items": matched,
                    }
            o["discrepancy"] = best
    return items


@api_router.patch("/orders/{oid}/status")
async def update_order_status(oid: str, body: OrderStatusUpdate, user=Depends(require_action("edit:orders"))):
    if body.status not in ("Pending", "Dispatched", "Cleared"):
        raise HTTPException(status_code=400, detail="Invalid status")
    res = await db.orders.update_one({"id": oid}, {"$set": {"status": body.status, "updated_at": now_iso()}})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Order not found")
    return await db.orders.find_one({"id": oid}, {"_id": 0})


@api_router.patch("/orders/{oid}")
async def admin_update_order(oid: str, body: OrderUpdate, admin=Depends(require_action("edit:orders"))):
    """Admin-only full edit of an order: customer, items, dates, notes, status."""
    existing = await db.orders.find_one({"id": oid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Order not found")
    update: Dict[str, Any] = {}
    if body.customer_id is not None and body.customer_id != existing.get("customer_id"):
        cust = await db.customers.find_one({"id": body.customer_id}, {"_id": 0})
        if not cust:
            raise HTTPException(status_code=404, detail="Customer not found")
        update["customer_id"] = body.customer_id
        update["customer_name"] = cust["name"]
    if body.items is not None:
        if not body.items:
            raise HTTPException(status_code=400, detail="Order must contain at least one item")
        item_ids = [it.item_id for it in body.items]
        found = await db.items.find({"id": {"$in": item_ids}}, {"_id": 0, "id": 1}).to_list(1000)
        found_ids = {f["id"] for f in found}
        missing = [iid for iid in item_ids if iid not in found_ids]
        if missing:
            raise HTTPException(status_code=400, detail=f"Unknown item_id(s): {missing}")
        update["items"] = [it.model_dump() for it in body.items]
    if body.order_date is not None:
        update["order_date"] = body.order_date
    if body.delivery_date is not None:
        update["delivery_date"] = body.delivery_date
    if body.notes is not None:
        update["notes"] = body.notes
    if body.status is not None:
        if body.status not in ("Pending", "Dispatched", "Cleared"):
            raise HTTPException(status_code=400, detail="Invalid status")
        update["status"] = body.status
    if not update:
        raise HTTPException(status_code=400, detail="No fields to update")
    update["updated_at"] = now_iso()
    await db.orders.update_one({"id": oid}, {"$set": update})
    return await db.orders.find_one({"id": oid}, {"_id": 0})


@api_router.delete("/orders/{oid}")
async def delete_order(oid: str, user=Depends(require_action("delete:orders"))):
    res = await db.orders.delete_one({"id": oid})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Order not found")
    return {"ok": True}


class DiscrepancyResolveIn(BaseModel):
    # One of: "update_date" | "clear" | "delete" | "keep"
    action: str
    dispatch_id: Optional[str] = None


@api_router.post("/orders/{oid}/resolve-discrepancy")
async def resolve_discrepancy(oid: str, body: DiscrepancyResolveIn,
                              user=Depends(get_current_user)):
    """Resolve a dispatch-before-order discrepancy on a Pending order.

    Actions:
      • update_date → set the order's date to the dispatch date, then dismiss.
      • clear       → reconcile: link the dispatch to this order and mark the
                      order Dispatched (its items were already shipped).
      • delete      → remove this (duplicate) order entry.
      • keep        → keep it Pending and stop flagging (dismiss the prompt).
    """
    action = (body.action or "").strip()
    if action not in ("update_date", "clear", "delete", "keep"):
        raise HTTPException(status_code=400, detail="Invalid action")

    order = await db.orders.find_one({"id": oid}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    # Permission: deleting needs delete rights; everything else needs edit.
    need = "delete:orders" if action == "delete" else "edit:orders"
    if not has_action_permission(user, need):
        raise HTTPException(status_code=403, detail="You don't have permission for this action.")

    if action == "delete":
        await db.orders.delete_one({"id": oid})
        return {"ok": True, "action": action, "deleted": True}

    if action == "keep":
        await db.orders.update_one(
            {"id": oid},
            {"$set": {"discrepancy_dismissed": True, "updated_at": now_iso()}},
        )
        return {"ok": True, "action": action, "order": await db.orders.find_one({"id": oid}, {"_id": 0})}

    # The remaining actions reference the matched dispatch.
    disp = None
    if body.dispatch_id:
        disp = await db.dispatches.find_one({"id": body.dispatch_id}, {"_id": 0})
    if not disp:
        raise HTTPException(status_code=404, detail="Linked dispatch not found")

    if action == "update_date":
        # Align the order's date with the day the goods actually went out.
        await db.orders.update_one(
            {"id": oid},
            {"$set": {"order_date": disp.get("dispatched_at") or order.get("order_date"),
                      "discrepancy_dismissed": True, "updated_at": now_iso()}},
        )
        return {"ok": True, "action": action, "order": await db.orders.find_one({"id": oid}, {"_id": 0})}

    # action == "clear" → reconcile the order against the existing dispatch.
    # IMPORTANT: only the items ACTUALLY on the dispatch slip are treated as
    # shipped. A discrepancy dispatch frequently covers just SOME of the
    # order's SKUs, so clearing the WHOLE order (as this used to do) wrongly
    # showed every line as dispatched. Items that are NOT on the slip stay
    # Pending so a later slip can ship them. The order becomes fully
    # Dispatched only when the slip covers every remaining line.
    slip_ids = {it.get("item_id") for it in (disp.get("items") or []) if it.get("item_id")}
    slip_names = {(it.get("item_name") or "").strip().lower()
                  for it in (disp.get("items") or []) if it.get("item_name")}

    def _on_slip(it: Dict[str, Any]) -> bool:
        iid = it.get("item_id")
        inm = (it.get("item_name") or "").strip().lower()
        return bool((iid and iid in slip_ids) or (inm and inm in slip_names))

    order_items = order.get("items") or []
    remaining_items = [it for it in order_items if not _on_slip(it)]
    fully_dispatched = len(remaining_items) == 0

    # Link this order onto the dispatch for traceability …
    order_ids = list(disp.get("order_ids") or [])
    if oid not in order_ids:
        order_ids.append(oid)
    await db.dispatches.update_one(
        {"id": disp["id"]},
        {"$set": {"order_ids": order_ids,
                  "order_fully_dispatched": fully_dispatched,
                  "updated_at": now_iso()}},
    )
    # Snapshot the pre-clear item list once so restores/reports keep a stable
    # baseline of what the order originally contained.
    if "original_items" not in order:
        await db.orders.update_one(
            {"id": oid}, {"$set": {"original_items": order_items}}
        )
    # Remove ONLY the shipped (on-slip) lines. Keep the rest Pending.
    await db.orders.update_one(
        {"id": oid},
        {"$set": {"items": remaining_items,
                  "status": "Dispatched" if fully_dispatched else "Pending",
                  "discrepancy_dismissed": True, "updated_at": now_iso()}},
    )
    return {"ok": True, "action": action,
            "fully_dispatched": fully_dispatched,
            "order": await db.orders.find_one({"id": oid}, {"_id": 0})}


# ======================== Dashboard Summary ========================
@api_router.get("/dashboard/summary")
async def dashboard_summary(user=Depends(get_current_user)):
    if is_blank_view(user):
        return {
            "stats": {
                "total_orders": 0,
                "pending_orders": 0,
                "dispatched_orders": 0,
                "cleared_orders": 0,
                "customers": 0,
                "products": 0,
            },
            "item_totals": [],
            "product_totals": [],
            "party_breakdown": [],
            "overdue_customers": [],
            "overdue_threshold_days": 15,
        }
    pending = await db.orders.find({"status": "Pending"}, {"_id": 0}).to_list(5000)

    # Read overdue threshold (default 15) so the dashboard can rank overdue customers.
    settings_doc = await _get_settings_doc()
    overdue_days_threshold = int(settings_doc.get("overdue_days", 15))
    now = datetime.now(timezone.utc)

    # Strict item-wise aggregation. Key = item_id (fallback to item_name for
    # legacy rows that may exist before the strict requirement landed).
    item_totals: Dict[str, Dict[str, Any]] = {}
    party_breakdown: Dict[str, Dict[str, Dict[str, Any]]] = {}
    # customer_id -> aggregated overdue stats
    overdue_customers: Dict[str, Dict[str, Any]] = {}
    for o in pending:
        cust_name = o["customer_name"]
        cust_id = o.get("customer_id")
        order_id = o.get("id")
        order_date = o.get("order_date") or o.get("created_at")
        party_breakdown.setdefault(cust_name, {})

        # days_open from order_date
        days_open = 0
        odt = o.get("order_date") or o.get("created_at")
        if isinstance(odt, str):
            try:
                dt = datetime.fromisoformat(odt.replace("Z", "+00:00"))
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=timezone.utc)
                days_open = max(0, (now - dt).days)
            except Exception:
                days_open = 0
        elif isinstance(odt, datetime):
            dt = odt if odt.tzinfo else odt.replace(tzinfo=timezone.utc)
            days_open = max(0, (now - dt).days)

        line_total = 0
        for it in o.get("items", []):
            # Robust qty parsing — strip stray commas/spaces from string-typed
            # legacy values so the grand total is always an integer sum.
            raw_qty = it.get("quantity") or 0
            if isinstance(raw_qty, str):
                raw_qty = raw_qty.replace(",", "").replace(" ", "").strip() or 0
            try:
                qty = int(float(raw_qty))
            except (TypeError, ValueError):
                qty = 0
            line_total += qty
            iid = it.get("item_id") or f"legacy:{it.get('item_name') or it.get('product_name')}"
            iname = it.get("item_name") or it.get("product_name") or "Unknown"
            pname = it.get("product_name") or ""

            row = item_totals.setdefault(iid, {
                "item_id": iid, "item_name": iname,
                "product_name": pname, "quantity": 0,
                "order_count": 0, "breakdown": [],
            })
            row["quantity"] += qty
            row["order_count"] += 1
            row["breakdown"].append({
                "order_id": order_id,
                "customer_name": cust_name,
                "quantity": qty,
                "order_date": order_date,
            })

            prow = party_breakdown[cust_name].setdefault(iid, {
                "item_id": iid, "item_name": iname,
                "product_name": pname, "quantity": 0,
            })
            prow["quantity"] += qty

        # Aggregate overdue stats per customer (only if past threshold).
        if days_open >= overdue_days_threshold:
            key = cust_id or f"name:{cust_name}"
            entry = overdue_customers.setdefault(key, {
                "customer_id": cust_id,
                "customer_name": cust_name,
                "oldest_days": 0,
                "pending_count": 0,
                "total_pcs": 0,
            })
            entry["oldest_days"] = max(entry["oldest_days"], days_open)
            entry["pending_count"] += 1
            entry["total_pcs"] += line_total

    total_orders = await db.orders.count_documents({})
    pending_count = len(pending)
    dispatched_count = await db.orders.count_documents({"status": "Dispatched"})
    cleared_count = await db.orders.count_documents({"status": "Cleared"})
    customers_count = await db.customers.count_documents({})
    products_count = await db.products.count_documents({})

    # Sort breakdown by quantity desc within each SKU row.
    for row in item_totals.values():
        row["breakdown"].sort(key=lambda x: -x["quantity"])

    item_totals_list = sorted(item_totals.values(), key=lambda x: -x["quantity"])
    party_list = [
        {"customer_name": c, "items": sorted(items.values(), key=lambda x: -x["quantity"])}
        for c, items in party_breakdown.items()
    ]
    overdue_customers_list = sorted(
        overdue_customers.values(),
        key=lambda x: (-x["oldest_days"], -x["pending_count"]),
    )

    return {
        "stats": {
            "total_orders": total_orders,
            "pending_orders": pending_count,
            "dispatched_orders": dispatched_count,
            "cleared_orders": cleared_count,
            "customers": customers_count,
            "products": products_count,
        },
        # Item-wise pending totals (strict requirement)
        "item_totals": item_totals_list,
        # Kept for backward compat — same item-wise rows
        "product_totals": item_totals_list,
        "party_breakdown": party_list,
        "overdue_customers": overdue_customers_list,
        "overdue_threshold_days": overdue_days_threshold,
    }


# ======================== Dispatch Matching ========================
@api_router.post("/dispatch/match")
async def dispatch_match(body: DispatchStockIn, user=Depends(get_current_user)):
    """Strict item-wise dispatch matching — show ALL candidates.

    Input: `items` dict where keys are **item_id** (SKU) and values are qty
    available. Returns every pending order that has demand for any of the
    input SKUs, sorted by oldest order first, so the operator can decide
    who to dispatch to (instead of greedily allocating to the first order
    and starving the rest).

    Per-order `allocated` = min(need, INPUT_stock_for_that_sku) — i.e. how
    much that order COULD receive if chosen — NOT how much remains after
    earlier orders. The actual deduction happens in `/dispatch/execute`."""
    products = await db.products.find({}, {"_id": 0}).to_list(1000)
    pmap = {p["name"]: p for p in products}

    # Look up item metadata for everything in the stock input so we can
    # render the response with proper item_name/product_name.
    stock_item_ids = [k for k, v in body.items.items() if int(v) > 0]
    item_docs = await db.items.find({"id": {"$in": stock_item_ids}}, {"_id": 0}).to_list(1000)
    imap = {it["id"]: it for it in item_docs}

    # Sort by the BUSINESS order date (when the customer actually placed
    # the order) ascending — i.e. oldest order first. Use `created_at` as
    # a tiebreaker for orders placed the same day.
    pending = (
        await db.orders.find({"status": "Pending"}, {"_id": 0})
        .sort([("order_date", 1), ("created_at", 1)])
        .to_list(5000)
    )

    # Reference stock map — NEVER decremented during candidate generation.
    # Allocation for each order is computed against this full input so that
    # every order that has demand is surfaced to the operator.
    input_stock: Dict[str, int] = {k: int(v) for k, v in body.items.items() if int(v) > 0}

    # Customer price-list cache (Task 1 — surface the saved price list for
    # each pending order so the dispatch UI can pre-select it). Also caches
    # city/location so the dispatch UI can show WHERE the party is — a
    # frequent operator ask so orders can be routed by area at a glance.
    cust_ids = list({o.get("customer_id") for o in pending if o.get("customer_id")})
    cust_pl: Dict[str, Optional[str]] = {}
    cust_loc: Dict[str, Dict[str, str]] = {}
    if cust_ids:
        async for c in db.customers.find(
            {"id": {"$in": cust_ids}},
            {"_id": 0, "id": 1, "price_list_id": 1, "city": 1, "location": 1},
        ):
            cust_pl[c["id"]] = c.get("price_list_id")
            cust_loc[c["id"]] = {
                "city": c.get("city") or "",
                "location": c.get("location") or "",
            }

    suggestions = []  # per order — all candidates
    # Bag calculation hint based on the INPUT stock (independent of which
    # order the user ends up selecting). Useful as a packing reference.
    per_item_allocated: Dict[str, int] = dict(input_stock)
    per_product_allocated: Dict[str, int] = {}
    for iid, qty in input_stock.items():
        pn = (imap.get(iid) or {}).get("product_name") or ""
        if pn:
            per_product_allocated[pn] = per_product_allocated.get(pn, 0) + qty

    for o in pending:
        order_alloc = []
        any_demand = False  # this order needs at least one of the input SKUs
        for it in o.get("items", []):
            iid = it.get("item_id")
            iname = it.get("item_name") or it.get("product_name") or "Unknown"
            pn = it.get("product_name") or ""
            need = int(it.get("quantity") or 0)
            stock_for_sku = input_stock.get(iid, 0) if iid else 0
            # `give` is how much this order WOULD receive if dispatched — uses
            # the FULL input stock, not a decremented residual.
            give = min(need, stock_for_sku) if (iid and need > 0) else 0
            if give > 0:
                any_demand = True
            order_alloc.append({
                "item_id": iid,
                "item_name": iname,
                "product_name": pn,
                "needed": need,
                "allocated": give,
                "shortfall": need - give,
                "variant": it.get("variant"),
                "fully_fulfilled": give == need and need > 0,
            })
        if any_demand:
            loc = cust_loc.get(o.get("customer_id") or "", {})
            suggestions.append({
                "order_id": o["id"],
                "customer_id": o["customer_id"],
                "customer_name": o["customer_name"],
                "customer_city": loc.get("city", ""),
                "customer_location": loc.get("location", ""),
                "order_date": o.get("order_date"),
                "delivery_date": o.get("delivery_date"),
                "allocations": order_alloc,
                "fully_fulfilled": all(a["fully_fulfilled"] or a["needed"] == 0 for a in order_alloc),
                # Task 1 — surfaces the customer's saved price list so the
                # operator can pre-select it (or override) before dispatch.
                "price_list_id": cust_pl.get(o["customer_id"]),
            })

    # Bag calculation: per-SKU when the item has its own override
    # (item.min_per_bag / item.max_per_bag), else grouped per master product.
    # Bags never mix products. SKUs with custom bag size pack separately.
    bag_calc = []
    product_residual: Dict[str, int] = {}  # qty per master product (no item override)
    for iid, qty in per_item_allocated.items():
        meta = imap.get(iid, {})
        i_min = meta.get("min_per_bag")
        i_max = meta.get("max_per_bag")
        if i_min and i_max and i_min > 0 and i_max > 0:
            min_bags = -(-qty // i_max)
            max_bags = -(-qty // i_min)
            bag_calc.append({
                "scope": "item",
                "item_id": iid,
                "item_name": meta.get("name", "Unknown"),
                "product_name": meta.get("product_name", ""),
                "allocated_qty": qty,
                "min_per_bag": i_min,
                "max_per_bag": i_max,
                "min_bags": min_bags,
                "max_bags": max_bags,
                "bag_range_label": f"{min_bags} bag{'s' if min_bags != 1 else ''}" if min_bags == max_bags else f"{min_bags}–{max_bags} bags",
            })
        else:
            pn = meta.get("product_name", "")
            product_residual[pn] = product_residual.get(pn, 0) + qty

    for pn, qty in product_residual.items():
        prod = pmap.get(pn)
        if prod:
            min_b = prod.get("min_per_bag") or 1
            max_b = prod.get("max_per_bag") or min_b
            min_bags = -(-qty // max_b) if max_b > 0 else 0  # ceil
            max_bags = -(-qty // min_b) if min_b > 0 else 0
            bag_calc.append({
                "scope": "product",
                "product_name": pn,
                "allocated_qty": qty,
                "min_per_bag": min_b,
                "max_per_bag": max_b,
                "min_bags": min_bags,
                "max_bags": max_bags,
                "bag_range_label": f"{min_bags} bag{'s' if min_bags != 1 else ''}" if min_bags == max_bags else f"{min_bags}–{max_bags} bags",
            })
        else:
            bag_calc.append({"scope": "product", "product_name": pn, "allocated_qty": qty, "min_bags": 0, "max_bags": 0, "bag_range_label": "N/A"})

    # `leftover_stock` no longer makes sense — we show ALL candidates and
    # the operator picks one. The actual leftover is computed in the UI
    # after they confirm a dispatch (the stock row is decremented locally).
    leftover = []

    return {
        "suggestions": suggestions,
        "bag_calculation": bag_calc,
        "leftover_stock": leftover,
        "per_item_allocated": [
            {
                "item_id": iid,
                "item_name": imap.get(iid, {}).get("name", "Unknown"),
                "product_name": imap.get(iid, {}).get("product_name", ""),
                "allocated_qty": qty,
            }
            for iid, qty in per_item_allocated.items()
        ],
        "input_stock": body.items,
    }


@api_router.post("/dispatch/execute")
async def dispatch_execute(body: DispatchExecuteIn, user=Depends(get_current_user)):
    """Partially fulfill a pending order. Subtracts the given quantities from
    each item line. Items hitting 0 are removed. If the order has no items
    left, it is marked Dispatched; otherwise it stays Pending so the
    remaining lot can be shipped later. A dispatch history record is saved."""
    order = await db.orders.find_one({"id": body.order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.get("status") != "Pending":
        raise HTTPException(status_code=400, detail=f"Order is {order.get('status')}, not Pending")
    if not body.allocations:
        raise HTTPException(status_code=400, detail="No allocations supplied")

    # Index order items by item_id for quick lookup
    order_items = order.get("items", [])
    by_iid: Dict[str, Dict[str, Any]] = {}
    for it in order_items:
        iid = it.get("item_id")
        if iid:
            by_iid[iid] = it

    # Validate allocations + price lookup against customer's assigned price list
    cust = await db.customers.find_one({"id": order["customer_id"]}, {"_id": 0}) or {}

    # Task 1 — if the operator explicitly chose a price list during this
    # dispatch, persist it as the customer's saved list AND use it for this
    # slip's line pricing. `None` ⇒ keep what's already on the customer.
    if body.price_list_id is not None:
        new_pl = body.price_list_id or None  # "" clears
        if new_pl and new_pl != cust.get("price_list_id"):
            pl_exists = await db.price_lists.count_documents({"id": new_pl})
            if pl_exists == 0:
                raise HTTPException(status_code=404, detail="Selected price list not found")
        await db.customers.update_one(
            {"id": order["customer_id"]},
            {"$set": {"price_list_id": new_pl}},
        )
        cust["price_list_id"] = new_pl
    cust_price_list = cust.get("price_list_id")
    cust_transport = cust.get("transport_name") or ""

    dispatched_lines: List[Dict[str, Any]] = []
    for alloc in body.allocations:
        if alloc.quantity <= 0:
            continue
        line = by_iid.get(alloc.item_id)
        if not line:
            raise HTTPException(status_code=400, detail=f"Item {alloc.item_id} not in this order")
        remaining = int(line.get("quantity") or 0)
        if alloc.quantity > remaining:
            raise HTTPException(
                status_code=400,
                detail=f"Cannot dispatch {alloc.quantity} of '{line.get('item_name')}' — only {remaining} remaining in order",
            )
        pricing = await compute_line_pricing(
            cust_price_list,
            alloc.item_id,
            line.get("product_name") or "",
        )
        dispatched_lines.append({
            "item_id": alloc.item_id,
            "item_name": line.get("item_name") or line.get("product_name"),
            "product_name": line.get("product_name"),
            "variant": line.get("variant"),
            "quantity": alloc.quantity,
            "description": (alloc.description or "").strip(),
            **pricing,
        })

    if not dispatched_lines:
        raise HTTPException(status_code=400, detail="All allocation quantities are zero")

    # Apply subtraction + 85% line-clear rule. `original_items` (snapshot
    # at the very first partial dispatch) drives the threshold so the rule
    # is stable across many partial dispatches on the same order.
    original_items_for_rule = order.get("original_items") or order_items
    dispatched_map = {d["item_id"]: d["quantity"] for d in dispatched_lines}
    new_items = _apply_line_clear_threshold(order_items, dispatched_map, original_items_for_rule)

    new_status = "Dispatched" if not new_items else "Pending"

    # NOTE: Raw-material stock is **clamped at zero** on consumption (see
    # `_consume_bom_for_lines` below) — the dispatch is ALWAYS allowed even
    # when the BOM would otherwise drive stock negative. This is intentional:
    # operators may not have keyed in every historical purchase yet, so the
    # system soft-tracks consumption and gradually aligns to physical stock
    # as more purchases get recorded.

    update_doc = {
        "items": new_items,
        "status": new_status,
        "updated_at": now_iso(),
    }
    # Preserve original item list on the very first dispatch (regardless of
    # whether the resulting status is Pending or Dispatched) so the 85%
    # auto-clear rule and any subsequent edit-unwind have a stable baseline
    # to refer back to. Without this snapshot, restoring a previously
    # auto-cleared SKU on a dispatch edit/delete can only return the delta
    # instead of the true residual.
    if "original_items" not in order:
        update_doc["original_items"] = order_items

    await db.orders.update_one({"id": body.order_id}, {"$set": update_doc})

    # Save dispatch history record — MERGE into today's existing slip for
    # this customer if one exists, so multiple same-day dispatches to one
    # party produce ONE consolidated slip instead of many.
    # NOTE: total_value (the printed Bill Amount) is NEVER auto-computed
    # from item pricing — it must be entered manually by the operator in
    # Daily Report / Dispatch Ledger edit. New dispatches start at 0.
    dispatch_ts = _resolve_dispatch_ts(body.dispatched_at)
    existing = await _find_open_dispatch_today(
        order["customer_id"], order["customer_name"], for_iso_ts=dispatch_ts,
    )
    if existing:
        merged_items = _merge_dispatch_lines(existing.get("items") or [], dispatched_lines)
        merged_total_pcs = sum(int(it.get("quantity") or 0) for it in merged_items)
        # Track all parent orders contributing to this slip (for traceability)
        order_ids = list(existing.get("order_ids") or [])
        if existing.get("order_id") and existing["order_id"] not in order_ids:
            order_ids.append(existing["order_id"])
        if body.order_id and body.order_id not in order_ids:
            order_ids.append(body.order_id)
        merged_notes = (existing.get("notes") or "").strip()
        if body.notes:
            merged_notes = (merged_notes + " | " + body.notes.strip()).strip(" |") if merged_notes else body.notes.strip()
        update_set: Dict[str, Any] = {
            "items": merged_items,
            "total_pcs": merged_total_pcs,
            # Task 2 — Bill Amount is NEVER auto-populated. It stays at
            # whatever value an admin has manually entered for the slip
            # (or 0 if untouched). Adding new items to a same-day slip
            # MUST NOT silently overwrite that manual entry.
            "order_ids": order_ids,
            "notes": merged_notes,
            "last_dispatched_at": now_iso(),
            "last_dispatched_by": user["email"],
        }
        # Once any line fully clears its order, propagate the "fully" flag up
        if new_status == "Dispatched":
            update_set["order_fully_dispatched"] = True
        await db.dispatches.update_one({"id": existing["id"]}, {"$set": update_set})
        dispatch_doc = await db.dispatches.find_one({"id": existing["id"]}, {"_id": 0})
    else:
        dispatch_doc = {
            "id": str(uuid.uuid4()),
            "slip_no": await next_slip_no(),
            "order_id": body.order_id,
            "order_ids": [body.order_id],
            "customer_id": order["customer_id"],
            "customer_name": order["customer_name"],
            "transport_name": cust_transport,
            "price_list_id": cust_price_list,
            "items": dispatched_lines,
            "total_pcs": sum(d["quantity"] for d in dispatched_lines),
            # Task 2 — Bill Amount starts EMPTY (0). The operator must
            # enter it manually from Daily Report / Dispatch Ledger.
            "total_value": 0.0,
            "notes": body.notes or "",
            "dispatched_by": user["email"],
            "dispatched_at": dispatch_ts,
            "order_fully_dispatched": new_status == "Dispatched",
        }
        await db.dispatches.insert_one(dispatch_doc)
        dispatch_doc.pop("_id", None)

    # Auto-consume raw materials from stock based on each dispatched line's
    # product BOM. Operates only on `dispatched_lines` (the NEW quantities
    # just sent out), so merging into an existing slip doesn't double-deduct.
    await _consume_bom_for_lines(dispatched_lines, dispatch_doc["id"], user.get("email") or "")

    updated_order = await db.orders.find_one({"id": body.order_id}, {"_id": 0})
    return {
        "dispatch": dispatch_doc,
        "order": updated_order,
        "fully_dispatched": new_status == "Dispatched",
    }


async def next_slip_no() -> int:
    """Atomically increment and return the next sequential dispatch slip number.

    The counter document is defensively bumped past any pre-existing slip
    number in ``db.dispatches`` before returning. This protects against the
    (rare but crashing) case where the ``db.counters`` document was restored
    from backup / seeded / hand-created after some dispatches were already
    persisted, leaving the counter behind the actual max ``slip_no`` and
    triggering ``DuplicateKeyError: slip_no_unique`` on the very next
    dispatch. Retries up to a few times to ride out concurrent bumps.
    """
    for _ in range(5):
        doc = await db.counters.find_one_and_update(
            {"_id": "dispatch_slip"},
            {"$inc": {"seq": 1}},
            upsert=True,
            return_document=True,  # ReturnDocument.AFTER
        )
        candidate = int(doc["seq"]) if doc else 1

        # Guard rail: if a dispatch with this slip number already exists,
        # advance the counter past the actual DB max and try again. This
        # can only happen once (or a very small number of times) — the
        # $max update below permanently corrects the drift.
        clash = await db.dispatches.count_documents({"slip_no": candidate}, limit=1)
        if not clash:
            return candidate

        highest = await db.dispatches.find_one(
            {}, {"_id": 0, "slip_no": 1}, sort=[("slip_no", -1)]
        )
        max_used = int((highest or {}).get("slip_no") or 0)
        await db.counters.update_one(
            {"_id": "dispatch_slip"},
            {"$max": {"seq": max_used}},
            upsert=True,
        )

    # Extremely unlikely fallback — surface a clean 500 instead of a raw
    # DuplicateKeyError bubbling out of ``insert_one``.
    raise HTTPException(
        status_code=500,
        detail="Could not allocate a unique dispatch slip number; please retry.",
    )


async def _find_open_dispatch_today(
    customer_id: Optional[str], customer_name: Optional[str],
    for_iso_ts: Optional[str] = None,
) -> Optional[Dict[str, Any]]:
    """Find an existing dispatch slip created on the SAME business day (IST)
    for the same customer. Used to merge multiple same-day dispatches into a
    single slip. Matches by customer_id when present, else by customer_name
    (walk-in). India Standard Time (UTC+5:30) is used so the "day" boundary
    matches when the factory actually operates.

    `for_iso_ts` lets the caller bucket against an explicit timestamp (e.g.
    a backdated dispatch) instead of "now"."""
    IST = timezone(timedelta(hours=5, minutes=30))
    if for_iso_ts:
        try:
            anchor = datetime.fromisoformat(str(for_iso_ts).replace("Z", "+00:00"))
            if anchor.tzinfo is None:
                anchor = anchor.replace(tzinfo=timezone.utc)
            day = anchor.astimezone(IST).date()
        except Exception:
            day = datetime.now(IST).date()
    else:
        day = datetime.now(IST).date()
    # Window covering all of `day` (IST), expressed in UTC for the ISO-string
    # comparison against `dispatched_at` (which is stored in UTC).
    start = datetime.combine(day, datetime.min.time(), tzinfo=IST).astimezone(timezone.utc).isoformat()
    end = datetime.combine(day, datetime.max.time(), tzinfo=IST).astimezone(timezone.utc).isoformat()
    q: Dict[str, Any] = {"dispatched_at": {"$gte": start, "$lte": end}}
    if customer_id:
        q["customer_id"] = customer_id
    else:
        q["customer_id"] = None
        q["customer_name"] = (customer_name or "").strip()
    return await db.dispatches.find_one(q, {"_id": 0}, sort=[("dispatched_at", 1)])


def _resolve_dispatch_ts(value: Optional[str]) -> str:
    """Normalise a user-supplied backdate into a UTC ISO timestamp.

    Accepts:
      * `None` / empty  → returns `now_iso()` (current UTC).
      * `"YYYY-MM-DD"`  → noon IST on that date (so the slip lands cleanly
        in that day's IST bucket on every report).
      * full ISO datetime → returned as UTC ISO string.

    Falls back to "now" on any parse failure so a bad client value never
    blocks a dispatch from being recorded."""
    if not value:
        return now_iso()
    s = str(value).strip()
    if not s:
        return now_iso()
    IST = timezone(timedelta(hours=5, minutes=30))
    try:
        if len(s) == 10 and s.count("-") == 2:
            y, m, d = (int(x) for x in s.split("-"))
            ist_noon = datetime(y, m, d, 12, 0, 0, tzinfo=IST)
            return ist_noon.astimezone(timezone.utc).isoformat()
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=IST)
        return dt.astimezone(timezone.utc).isoformat()
    except Exception:
        return now_iso()


def _merge_dispatch_lines(
    existing: List[Dict[str, Any]], incoming: List[Dict[str, Any]]
) -> List[Dict[str, Any]]:
    """Combine new dispatch lines into the existing list. Lines with the same
    item_id, same net_unit_price AND same description are merged (quantities
    summed); otherwise the new line is appended as-is so different pricing
    or descriptions stay auditable."""
    out: List[Dict[str, Any]] = [dict(x) for x in existing]
    for new in incoming:
        new_iid = new.get("item_id")
        new_net = float(new.get("net_unit_price") or 0)
        new_desc = (new.get("description") or "").strip()
        match_idx = None
        for i, ex in enumerate(out):
            if (
                ex.get("item_id") == new_iid
                and float(ex.get("net_unit_price") or 0) == new_net
                and (ex.get("description") or "").strip() == new_desc
            ):
                match_idx = i
                break
        if match_idx is not None:
            merged = dict(out[match_idx])
            merged["quantity"] = int(merged.get("quantity") or 0) + int(new.get("quantity") or 0)
            net = float(merged.get("net_unit_price") or 0)
            merged["line_value"] = round(net * merged["quantity"], 2)
            out[match_idx] = merged
        else:
            line = dict(new)
            if "line_value" not in line:
                line["line_value"] = round(
                    float(line.get("net_unit_price") or 0) * int(line.get("quantity") or 0), 2
                )
            out.append(line)
    return out


def _sum_line_values(items: List[Dict[str, Any]]) -> float:
    """Sum of `line_value` across dispatch line items. Falls back to
    `net_unit_price * quantity` if line_value is missing."""
    total = 0.0
    for it in items or []:
        lv = it.get("line_value")
        if lv is None:
            lv = float(it.get("net_unit_price") or 0) * int(it.get("quantity") or 0)
        total += float(lv or 0)
    return round(total, 2)


async def next_receipt_no() -> int:
    """Atomically increment and return the next sequential payment receipt number."""
    doc = await db.counters.find_one_and_update(
        {"_id": "payment_receipt"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True,
    )
    return int(doc["seq"]) if doc else 1


@api_router.get("/dispatches")
async def list_dispatches(order_id: Optional[str] = None, customer_id: Optional[str] = None,
                          user=Depends(get_current_user)):
    if is_blank_view(user):
        return []
    q = {}
    if order_id:
        q["order_id"] = order_id
    if customer_id:
        q["customer_id"] = customer_id
    items = await db.dispatches.find(q, {"_id": 0}).sort("dispatched_at", -1).to_list(2000)
    return items


# ======================== Admin Dispatch Ledger (per-dispatch GR + slip) ========================
@api_router.get("/admin/dispatch-ledger")
async def admin_dispatch_ledger(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    customer_id: Optional[str] = None,
    limit: int = 200,
    skip: int = 0,
    user=Depends(get_current_user),
):
    """Dispatch ledger — every dispatch (regular + off-order) with its GR
    number. Accessible to any authenticated user so it can live in the main
    interface (not buried under Admin Settings). Newest first."""
    if is_blank_view(user):
        return {"total": 0, "items": [], "grand_total_value": 0, "grand_total_pcs": 0}
    q: Dict[str, Any] = {}
    if start_date or end_date:
        IST = timezone(timedelta(hours=5, minutes=30))
        rng: Dict[str, str] = {}
        if start_date:
            try:
                d = datetime.strptime(start_date, "%Y-%m-%d").date()
            except ValueError:
                raise HTTPException(status_code=400, detail="start_date must be YYYY-MM-DD")
            rng["$gte"] = datetime.combine(d, datetime.min.time(), tzinfo=IST).astimezone(timezone.utc).isoformat()
        if end_date:
            try:
                d = datetime.strptime(end_date, "%Y-%m-%d").date()
            except ValueError:
                raise HTTPException(status_code=400, detail="end_date must be YYYY-MM-DD")
            rng["$lte"] = datetime.combine(d, datetime.max.time(), tzinfo=IST).astimezone(timezone.utc).isoformat()
        q["dispatched_at"] = rng
    if customer_id:
        q["customer_id"] = customer_id
    cursor = (
        db.dispatches.find(q, {"_id": 0})
        .sort("dispatched_at", -1)
        .skip(max(0, int(skip)))
        .limit(max(1, min(500, int(limit))))
    )
    rows = await cursor.to_list(length=500)
    total = await db.dispatches.count_documents(q)
    grand_value = round(sum(float(r.get("total_value") or 0) for r in rows), 2)
    grand_pcs = sum(int(r.get("total_pcs") or 0) for r in rows)
    return {"total": total, "items": rows, "grand_total_value": grand_value, "grand_total_pcs": grand_pcs}


@api_router.patch("/dispatches/{did}/gr")
async def update_dispatch_gr(did: str, body: DispatchGrUpdate, user=Depends(require_action("edit:dispatch"))):
    """Set / update the GR (Goods Receipt) number for a dispatch.
    Any authenticated user can edit this field (per product spec)."""
    gr = (body.gr_number or "").strip()
    res = await db.dispatches.update_one(
        {"id": did},
        {"$set": {
            "gr_number": gr,
            "gr_updated_at": now_iso(),
            "gr_updated_by": user["email"],
        }},
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Dispatch not found")
    return await db.dispatches.find_one({"id": did}, {"_id": 0})


# ---- Dispatch row edit / delete (used by Single Party Ledger row actions) ----


class DispatchEditItem(BaseModel):
    item_id: Optional[str] = None
    item_name: str
    product_name: Optional[str] = ""
    variant: Optional[str] = ""
    quantity: float
    unit_price: Optional[float] = 0.0
    net_unit_price: Optional[float] = None
    discount_value: Optional[float] = 0.0
    discount_type: Optional[str] = ""
    # Free-text description / note printed under the item name on the slip.
    description: Optional[str] = ""


class DispatchEdit(BaseModel):
    gr_number: Optional[str] = None
    gr_date: Optional[str] = None  # YYYY-MM-DD (or empty string to clear)
    transport_name: Optional[str] = None
    notes: Optional[str] = None
    total_value: Optional[float] = None  # allow overriding bill amount (debit)
    bag_count: Optional[int] = None  # operator-entered number of bags shipped
    items: Optional[List[DispatchEditItem]] = None  # if provided, replaces line items
    # Reassign this slip to a different customer. When set, the slip's
    # customer_id + customer_name are updated; because the customer ledger /
    # dispatch report are derived by customer_id, the slip moves cleanly to
    # the new party's ledger.
    customer_id: Optional[str] = None
    # Manually-typed Bill Number for bill-number-mode parties (per dispatch).
    bill_number: Optional[str] = None
    # Optional backdate / date correction. Accepts "YYYY-MM-DD" or full ISO
    # datetime; resolved to UTC at noon IST when only a date is given so
    # the slip lands cleanly in that day's bucket on the dispatch report.
    dispatched_at: Optional[str] = None


@api_router.patch("/dispatches/{did}")
async def update_dispatch(did: str, body: DispatchEdit, user=Depends(require_action("edit:dispatch"))):
    """Edit a dispatch's bookkeeping fields (GR, transport, notes, bill
    amount, bag count) and optionally its line items (name / qty / price).

    Behaviour (Jan 2026): when `items` is provided, the per-SKU delta
    between the existing and new lines is also reflected on the parent
    order(s):
      • Increased qty / new lines → deducted from the customer's pending
        orders FIFO (same logic as Off-Order Dispatch).
      • Decreased qty / removed lines → the difference is restored back
        into the dispatch's parent order(s); if the parent order had been
        marked Dispatched it is reopened to Pending.
    Raw-material stock is intentionally NOT rebalanced on PATCH; use
    DELETE (which reverses BOM consumption) to fully roll back a slip.

    Non-admin users may only edit a dispatch within `edit_window_days`
    (admin-configurable in /api/settings, default 3 days) of when it was
    punched. Admins can edit at any time.
    """
    existing = await db.dispatches.find_one({"id": did}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Dispatch not found")
    # Enforce the user-edit window. Admins are always allowed.
    if user.get("role") != "admin":
        settings = await _get_settings_doc()
        window_days = int(settings.get("edit_window_days", 3) or 0)
        ts = existing.get("dispatched_at") or existing.get("created_at")
        if ts:
            try:
                punched = datetime.fromisoformat(str(ts).replace("Z", "+00:00"))
                if punched.tzinfo is None:
                    punched = punched.replace(tzinfo=timezone.utc)
                age_days = (datetime.now(timezone.utc) - punched).total_seconds() / 86400.0
            except Exception:
                age_days = 0.0
        else:
            age_days = 0.0
        if window_days <= 0 or age_days > window_days:
            raise HTTPException(
                status_code=403,
                detail=f"Editing is locked for users after {window_days} day(s). Ask an admin to make this change.",
            )
    upd: Dict[str, Any] = {"updated_at": now_iso(), "updated_by": user["email"]}
    # Reassign the slip to a different customer (admin can correct a slip
    # punched under the wrong party). The ledger/report follow customer_id.
    if body.customer_id is not None and body.customer_id and body.customer_id != existing.get("customer_id"):
        cust = await db.customers.find_one({"id": body.customer_id}, {"_id": 0})
        if not cust:
            raise HTTPException(status_code=404, detail="Customer not found")
        upd["customer_id"] = cust["id"]
        upd["customer_name"] = cust.get("name") or existing.get("customer_name")
    if body.gr_number is not None:
        upd["gr_number"] = body.gr_number.strip()
    if body.gr_date is not None:
        # Accept "YYYY-MM-DD" (or empty string to clear). Also tolerate a
        # full ISO datetime and take the date portion so operators can
        # paste any of the app's date formats.
        raw = (body.gr_date or "").strip()
        if not raw:
            upd["gr_date"] = ""
        else:
            m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", raw)
            if not m:
                raise HTTPException(status_code=400, detail="gr_date must be in YYYY-MM-DD format")
            try:
                y, mo, dd = int(m.group(1)), int(m.group(2)), int(m.group(3))
                datetime(y, mo, dd)  # validate calendar date
            except Exception:
                raise HTTPException(status_code=400, detail="gr_date is not a valid calendar date")
            upd["gr_date"] = f"{y:04d}-{mo:02d}-{dd:02d}"
    if body.transport_name is not None:
        upd["transport_name"] = body.transport_name.strip()
        # Persist this transport as the default for the dispatch's customer
        # (Jan 2026) so subsequent dispatches default to the same value.
        await _persist_customer_transport(existing.get("customer_id"), body.transport_name)
    if body.notes is not None:
        upd["notes"] = body.notes.strip()
    if body.bill_number is not None:
        upd["bill_number"] = body.bill_number.strip()
    if body.bag_count is not None:
        if int(body.bag_count) < 0:
            raise HTTPException(status_code=400, detail="bag_count cannot be negative")
        upd["bag_count"] = int(body.bag_count)
    if body.dispatched_at is not None:
        # Allow operators / admins to correct the slip date (e.g. when a
        # slip was punched late and needs to be moved to its actual day).
        # Empty string clears any override → falls back to current value.
        new_ts = (body.dispatched_at or "").strip()
        if new_ts:
            upd["dispatched_at"] = _resolve_dispatch_ts(new_ts)
    restore_lines: List[Dict[str, Any]] = []
    if body.items is not None:
        new_items: List[Dict[str, Any]] = []
        for it in body.items:
            qty = int(round(float(it.quantity or 0)))
            if qty <= 0 or not (it.item_name or "").strip():
                # skip empty / invalid rows silently
                continue
            unit = float(it.unit_price or 0)
            net = float(it.net_unit_price) if it.net_unit_price is not None else unit
            new_items.append({
                "item_id": it.item_id or str(uuid.uuid4()),
                "item_name": it.item_name.strip(),
                "product_name": (it.product_name or "").strip(),
                "variant": (it.variant or "").strip(),
                "quantity": qty,
                "unit_price": round(unit, 2),
                "discount_value": round(float(it.discount_value or 0), 2),
                "discount_type": (it.discount_type or ""),
                "net_unit_price": round(net, 2),
                "line_value": round(net * qty, 2),
                "description": (it.description or "").strip(),
                # Manual edit via the Edit-Dispatch dialog → treat this
                # item's price as an operator override so downstream views
                # (Dispatch Report, Slip preview, Share PDF) show the
                # exact number the operator typed instead of silently
                # recomputing from the current price list. Only rows with
                # a positive stored net_unit_price get flagged.
                "price_override": bool(round(net, 2) > 0),
            })
        if not new_items:
            raise HTTPException(status_code=400, detail="At least one item with name and quantity > 0 is required")
        upd["items"] = new_items
        upd["total_pcs"] = sum(it["quantity"] for it in new_items)
        # Task 2 — Bill Amount is NEVER auto-recomputed when items change.
        # The operator's manually-entered total_value is preserved; if they
        # want to refresh it after editing items, they can re-enter it from
        # Daily Report / Dispatch Ledger.

        # Rebalance the parent order(s) by the per-item diff so the pending
        # order quantities stay in sync with the edited slip (Jan 2026):
        #   delta > 0 (more dispatched) → deduct extra from pending orders
        #   delta < 0 (less dispatched) → restore the difference back
        old_by_iid: Dict[str, Dict[str, Any]] = {
            it.get("item_id"): it for it in (existing.get("items") or []) if it.get("item_id")
        }
        new_by_iid: Dict[str, Dict[str, Any]] = {
            it["item_id"]: it for it in new_items if it.get("item_id")
        }
        add_lines: List[Dict[str, Any]] = []
        for iid, new_line in new_by_iid.items():
            old_qty = int((old_by_iid.get(iid) or {}).get("quantity") or 0)
            new_qty = int(new_line.get("quantity") or 0)
            delta = new_qty - old_qty
            if delta > 0:
                add_lines.append({**new_line, "quantity": delta})
            elif delta < 0:
                restore_lines.append({**new_line, "quantity": -delta})
        # Items removed entirely → restore full old qty.
        for iid, old_line in old_by_iid.items():
            if iid not in new_by_iid:
                restore_lines.append({**old_line, "quantity": int(old_line.get("quantity") or 0)})

        if add_lines:
            extra_touched = await _deduct_off_order_from_pending_orders(
                existing.get("customer_id"), add_lines
            )
            if extra_touched:
                merged_oids = list(existing.get("order_ids") or [])
                for oid in extra_touched:
                    if oid not in merged_oids:
                        merged_oids.append(oid)
                upd["order_ids"] = merged_oids
    if body.total_value is not None:
        if float(body.total_value) < 0:
            raise HTTPException(status_code=400, detail="total_value cannot be negative")
        upd["total_value"] = round(float(body.total_value), 2)
    # Persist the dispatch update FIRST so that the subsequent restore pass
    # sees the new line quantities when it sums dispatched totals across all
    # live dispatches for a given order. This is required for the 85%
    # auto-clear unwind: if an item was previously dropped from the order
    # because cumulative dispatch crossed the threshold and the user now
    # reduces the dispatch below it, we must recompute pending from the
    # original_items snapshot, not from the stale per-edit delta.
    res = await db.dispatches.update_one({"id": did}, {"$set": upd})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Dispatch not found")
    if body.items is not None and restore_lines:
        await _restore_dispatch_qty_to_orders(existing, restore_lines)
    return await db.dispatches.find_one({"id": did}, {"_id": 0})


@api_router.delete("/dispatches/{did}")
async def delete_dispatch(did: str, admin=Depends(require_action("delete:dispatch"))):
    """Delete a dispatch record (admin only) and rebalance the customer's
    parent order(s) + raw-material stock so the system stays consistent.

    Behaviour (Jan 2026):
    - Every item on the dispatch is added back to its parent order. The
      order's status flips from `Dispatched` → `Pending` if it gets items
      again. Lines that were auto-cleared by the 85% rule are recreated
      from the order's `original_items` snapshot (or from the dispatch
      line itself as a last resort).
    - BOM raw-material consumption is reversed: every `dispatch` movement
      written by this dispatch is matched with a paired `dispatch_revert`
      movement that re-credits stock.
    - The dispatch document is then removed from `dispatches`.
    """
    existing = await db.dispatches.find_one({"id": did}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Dispatch not found")

    restore_items = list(existing.get("items") or [])
    # Delete the dispatch first so subsequent restoration logic sums only
    # the remaining live dispatches when reconstructing pending quantities
    # for lines that were previously auto-cleared by the 85% rule.
    await db.dispatches.delete_one({"id": did})
    touched_orders = await _restore_dispatch_qty_to_orders(existing, restore_items)
    restored_rm = await _revert_bom_consumption_for_dispatch(did, admin.get("email") or "")

    # Tombstone — preserve the slip number so a future lookup on this
    # number returns a clear "this slip was deleted" message instead of a
    # generic 404. The global slip counter (`next_slip_no`) is INCREMENT-
    # ONLY, so this slip_no will never be re-issued to a new dispatch.
    if existing.get("slip_no") is not None:
        await db.deleted_slips.update_one(
            {"slip_no": int(existing["slip_no"])},
            {"$set": {
                "slip_no": int(existing["slip_no"]),
                "dispatch_id": did,
                "customer_id": existing.get("customer_id"),
                "customer_name": existing.get("customer_name") or "",
                "original_dispatched_at": existing.get("dispatched_at") or existing.get("created_at"),
                "total_value": float(existing.get("total_value") or 0),
                "total_pcs": int(existing.get("total_pcs") or 0),
                "deleted_at": now_iso(),
                "deleted_by": admin.get("email") or admin.get("username") or "admin",
            }},
            upsert=True,
        )

    return {
        "ok": True,
        "deleted": did,
        "order_ids_restored": touched_orders,
        "raw_materials_restored": restored_rm,
        "slip_no": existing.get("slip_no"),
    }


@api_router.get("/admin/dispatch-ledger/{did}/slip")
async def admin_dispatch_slip(did: str, user=Depends(get_current_user)):
    """Full slip payload (customer details + lines + totals) used by the
    frontend to render a printable dispatch slip. Accessible to any
    authenticated user."""
    d = await db.dispatches.find_one({"id": did}, {"_id": 0})
    if not d:
        raise HTTPException(status_code=404, detail="Dispatch not found")
    cust: Dict[str, Any] = {}
    if d.get("customer_id"):
        cust = await db.customers.find_one({"id": d["customer_id"]}, {"_id": 0}) or {}
    return {
        "dispatch": d,
        "customer": {
            "id": cust.get("id"),
            "name": d.get("customer_name") or cust.get("name") or "—",
            "phone": cust.get("phone") or "",
            "address": cust.get("address") or "",
            "city": cust.get("city") or "",
            "location": cust.get("location") or "",
            "transport_name": d.get("transport_name") or cust.get("transport_name") or "",
            "private_mark": cust.get("private_mark") or "",
        },
    }


@api_router.get("/dispatches/by-slip/{slip_no}")
async def get_dispatch_by_slip(slip_no: int, user=Depends(get_current_user)):
    """Global slip-number lookup.

    Each dispatch is assigned a globally-unique sequential `slip_no` (see
    `next_slip_no`). This endpoint mirrors the invoice-search behaviour of
    accounting software: type any slip number and we return the dispatch
    plus its customer context so the frontend can jump straight to it on
    the Customer Ledger.

    If the slip was deleted, returns HTTP 410 Gone with a `deleted` body
    so the UI can clearly say "Slip #N was deleted on <date> by <user>".
    The slip number itself is NEVER reused — the global counter only
    increments — so a 410 here is a permanent state, not a transient one.
    """
    try:
        n = int(slip_no)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid slip number")
    d = await db.dispatches.find_one({"slip_no": n}, {"_id": 0})
    if not d:
        # Check tombstone collection so users get a clear message instead
        # of a generic 404 for previously-issued but now-deleted slips.
        tomb = await db.deleted_slips.find_one({"slip_no": n}, {"_id": 0})
        if tomb:
            raise HTTPException(
                status_code=410,
                detail={
                    "message": f"Slip #{n} was deleted",
                    "deleted": True,
                    "slip_no": n,
                    "deleted_at": tomb.get("deleted_at"),
                    "deleted_by": tomb.get("deleted_by"),
                    "customer_name": tomb.get("customer_name") or "",
                    "original_dispatched_at": tomb.get("original_dispatched_at"),
                    "total_value": tomb.get("total_value") or 0,
                    "total_pcs": tomb.get("total_pcs") or 0,
                },
            )
        raise HTTPException(status_code=404, detail=f"No slip with number {n}")
    cust: Dict[str, Any] = {}
    if d.get("customer_id"):
        cust = await db.customers.find_one({"id": d["customer_id"]}, {"_id": 0}) or {}
    return {
        "dispatch": d,
        "customer": {
            "id": cust.get("id") or d.get("customer_id"),
            "name": d.get("customer_name") or cust.get("name") or "—",
            "phone": cust.get("phone") or "",
            "address": cust.get("address") or "",
            "city": cust.get("city") or "",
            "location": cust.get("location") or "",
            "transport_name": d.get("transport_name") or cust.get("transport_name") or "",
            "private_mark": cust.get("private_mark") or "",
        },
    }



# ======================== Price Lists ========================
async def compute_line_pricing(price_list_id: Optional[str], item_id: str, product_name: str) -> Dict[str, Any]:
    """Resolve unit_price, discount_value, discount_type and net_unit_price
    for one line, given an optional price-list assignment.
    If no price list / no entry exists, returns 0 values gracefully."""
    out = {
        "unit_price": 0.0,
        "discount_value": 0.0,
        "discount_type": "",
        "net_unit_price": 0.0,
    }
    if not price_list_id:
        return out
    pli = await db.price_list_items.find_one(
        {"price_list_id": price_list_id, "item_id": item_id}, {"_id": 0}
    )
    if pli:
        out["unit_price"] = float(pli.get("price") or 0)
    disc = await db.price_list_category_discounts.find_one(
        {"price_list_id": price_list_id, "product_name": product_name}, {"_id": 0}
    )
    if disc:
        out["discount_value"] = float(disc.get("discount_value") or 0)
        out["discount_type"] = disc.get("discount_type") or ""
    # Defensive: if a discount value is set but type is missing (legacy /
    # half-saved data), treat it as flat rupees off so the discount
    # actually applies instead of being silently ignored.
    if out["discount_value"] > 0 and out["discount_type"] not in ("₹", "%"):
        out["discount_type"] = "₹"
    # Net unit price after discount
    net = out["unit_price"]
    if out["discount_type"] == "%":
        net = max(0.0, net * (1 - out["discount_value"] / 100.0))
    elif out["discount_type"] == "₹":
        net = max(0.0, net - out["discount_value"])
    out["net_unit_price"] = round(net, 2)
    return out


# ─────────────────────────────────────────────────────────────────────────────
# ESTIMATES — quick quotation for any customer without creating an order.
# Given a customer + item list (item_id or name + quantity), the server pulls
# the customer's assigned price list, resolves per-line pricing/discounts and
# returns a full breakdown ready for on-screen display or PDF/share.
# The optional `bill_amount` is what the party pays "on the bill" and the
# remainder becomes the cash portion (same formula the dispatch slip uses).
# ─────────────────────────────────────────────────────────────────────────────

class EstimateItemIn(BaseModel):
    item_id: Optional[str] = None
    item_name: Optional[str] = None      # fallback if item_id missing
    quantity: float

class EstimateIn(BaseModel):
    customer_id: str
    items: List[EstimateItemIn]
    bill_amount: Optional[float] = 0     # what the customer will pay "on bill"
    # Optional — override the customer's assigned price list for this
    # estimate only. When empty/None, the customer's default is used.
    price_list_id_override: Optional[str] = None

async def _compute_estimate(body: EstimateIn) -> Dict[str, Any]:
    """Core estimate builder — resolves the customer's price list, prices
    every line and returns the full breakdown. Shared by the live compute
    endpoint (no writes) and the save endpoint (persists a record)."""
    cust = await db.customers.find_one({"id": body.customer_id}, {"_id": 0})
    if not cust:
        raise HTTPException(status_code=404, detail="Customer not found")
    # If the request explicitly picks a price list, that OVERRIDES the customer's
    # default assignment. Empty-string / None ⇒ fall back to the customer's own.
    override_pl = getattr(body, "price_list_id_override", None)
    price_list_id = (override_pl or cust.get("price_list_id") or None)
    price_list_name = None
    if price_list_id:
        pl = await db.price_lists.find_one({"id": price_list_id}, {"_id": 0, "name": 1})
        price_list_name = pl.get("name") if pl else None

    lines: List[Dict[str, Any]] = []
    line_total = 0.0
    for row in body.items:
        qty = float(row.quantity or 0)
        if qty <= 0:
            continue
        item = None
        if row.item_id:
            item = await db.items.find_one({"id": row.item_id}, {"_id": 0})
        if not item and row.item_name:
            # Case-insensitive exact/substring match on name.
            item = await db.items.find_one(
                {"name": {"$regex": f"^{row.item_name.strip()}$", "$options": "i"}}, {"_id": 0}
            )
        if not item:
            # Unknown SKU — surface it so the UI can flag it.
            lines.append({
                "item_id": row.item_id or "",
                "item_name": row.item_name or "(unknown)",
                "product_name": "",
                "quantity": qty,
                "unit_price": 0.0,
                "discount_value": 0.0,
                "discount_type": "",
                "net_unit_price": 0.0,
                "line_value": 0.0,
                "found": False,
            })
            continue
        pricing = await compute_line_pricing(price_list_id, item["id"], item.get("product_name") or "")
        line_value = round(pricing["net_unit_price"] * qty, 2)
        line_total += line_value
        lines.append({
            "item_id": item["id"],
            "item_name": item["name"],
            "product_name": item.get("product_name") or "",
            "quantity": qty,
            "unit_price": pricing["unit_price"],
            "discount_value": pricing["discount_value"],
            "discount_type": pricing["discount_type"],
            "net_unit_price": pricing["net_unit_price"],
            "line_value": line_value,
            "found": True,
        })

    # Cash / bill split — same math the dispatch slip uses.
    #   gst      = round(bill_amount * 18 / 118)
    #   grand    = round(line_total + gst)
    #   cash     = max(0, grand - bill_amount)
    bill_amount = float(body.bill_amount or 0)
    gst = round(bill_amount * 18 / 118)
    grand_total = round(line_total + gst)
    cash_amount = max(0, grand_total - bill_amount)

    return {
        "customer": {
            "id": cust["id"],
            "name": cust.get("name"),
            "phone": cust.get("phone"),
            "address": cust.get("address"),
            "city": cust.get("city"),
            "location": cust.get("location"),
            "private_mark": cust.get("private_mark") or "",
            "transport_name": cust.get("transport_name") or "",
        },
        "price_list_id": price_list_id,
        "price_list_name": price_list_name,
        "lines": lines,
        "totals": {
            "line_total": round(line_total, 2),
            "bill_amount": round(bill_amount, 2),
            "gst": gst,
            "grand_total": grand_total,
            "cash_amount": cash_amount,
        },
        "generated_at": now_iso(),
    }


@api_router.post("/estimates/compute")
async def compute_estimate(body: EstimateIn, user=Depends(get_current_user)):
    """Return a fully-priced estimate for a customer. No DB writes."""
    return await _compute_estimate(body)


async def next_estimate_no() -> int:
    """Atomically increment and return the next sequential estimate number.

    Mirrors ``next_slip_no`` — the counter is defensively bumped past any
    pre-existing ``estimate_no`` in ``db.estimates`` so a restored counter
    can never re-issue a number that is already on a saved estimate.
    """
    for _ in range(5):
        doc = await db.counters.find_one_and_update(
            {"_id": "estimate_no"},
            {"$inc": {"seq": 1}},
            upsert=True,
            return_document=True,
        )
        candidate = int(doc["seq"]) if doc else 1
        clash = await db.estimates.count_documents({"estimate_no": candidate}, limit=1)
        if not clash:
            return candidate
        highest = await db.estimates.find_one(
            {}, {"_id": 0, "estimate_no": 1}, sort=[("estimate_no", -1)]
        )
        max_used = int((highest or {}).get("estimate_no") or 0)
        await db.counters.update_one(
            {"_id": "estimate_no"}, {"$max": {"seq": max_used}}, upsert=True,
        )
    raise HTTPException(status_code=500, detail="Could not allocate an estimate number")


@api_router.post("/estimates")
async def save_estimate(body: EstimateIn, user=Depends(get_current_user)):
    """Compute AND persist an estimate, assigning a globally-unique
    sequential ``estimate_no``. Returns the saved record so the UI can show
    the number on the slip and in the Records list."""
    payload = await _compute_estimate(body)
    est_no = await next_estimate_no()
    doc = {
        "id": str(uuid.uuid4()),
        "estimate_no": est_no,
        "customer_id": body.customer_id,
        "customer_name": payload["customer"].get("name") or "",
        "customer": payload["customer"],
        "price_list_id": payload.get("price_list_id"),
        "price_list_name": payload.get("price_list_name"),
        "lines": payload["lines"],
        "totals": payload["totals"],
        "bill_amount": payload["totals"].get("bill_amount", 0),
        "created_at": now_iso(),
        "created_by": user.get("email"),
    }
    await db.estimates.insert_one(dict(doc))
    doc.pop("_id", None)
    return doc


@api_router.get("/estimates")
async def list_estimates(
    q: Optional[str] = None,
    limit: int = 500,
    user=Depends(get_current_user),
):
    """List all saved estimates, newest first. Optional `q` filters by
    customer name (case-insensitive) or by estimate number."""
    filt: Dict[str, Any] = {}
    if q and q.strip():
        term = q.strip()
        ors: List[Dict[str, Any]] = [
            {"customer_name": {"$regex": re.escape(term), "$options": "i"}}
        ]
        if term.isdigit():
            ors.append({"estimate_no": int(term)})
        filt = {"$or": ors}
    docs = await db.estimates.find(filt, {"_id": 0}).sort("estimate_no", -1).to_list(int(limit or 500))
    # Lightweight summary rows for the table (full doc still available on GET/{id}).
    out = []
    for d in docs:
        totals = d.get("totals") or {}
        lines = d.get("lines") or []
        out.append({
            "id": d.get("id"),
            "estimate_no": d.get("estimate_no"),
            "customer_id": d.get("customer_id"),
            "customer_name": d.get("customer_name") or (d.get("customer") or {}).get("name") or "",
            "item_count": len(lines),
            "total_pcs": sum(float(l.get("quantity") or 0) for l in lines),
            "grand_total": totals.get("grand_total", 0),
            "bill_amount": totals.get("bill_amount", 0),
            "cash_amount": totals.get("cash_amount", 0),
            "price_list_name": d.get("price_list_name"),
            "created_at": d.get("created_at"),
            "created_by": d.get("created_by"),
        })
    return {"estimates": out, "count": len(out)}


@api_router.get("/estimates/{eid}")
async def get_estimate(eid: str, user=Depends(get_current_user)):
    """Fetch a single saved estimate by its id (full breakdown for
    re-printing / viewing)."""
    doc = await db.estimates.find_one({"id": eid}, {"_id": 0})
    if doc is None:
        raise HTTPException(status_code=404, detail="Estimate not found")
    # Normalise into the same shape the compute endpoint returns so the UI
    # slip renderer can be reused verbatim.
    doc["generated_at"] = doc.get("created_at")
    return doc


@api_router.delete("/estimates/{eid}")
async def delete_estimate(eid: str, admin=Depends(require_admin)):
    """Admin: delete a saved estimate. The sequential number is NOT
    re-issued (the counter is increment-only)."""
    existing = await db.estimates.find_one({"id": eid}, {"_id": 0, "id": 1})
    if existing is None:
        raise HTTPException(status_code=404, detail="Estimate not found")
    await db.estimates.delete_one({"id": eid})
    return {"ok": True, "deleted": eid}




@api_router.get("/price-lists")
async def list_price_lists(user=Depends(get_current_user)):
    lists = await db.price_lists.find({}, {"_id": 0}).sort("name", 1).to_list(500)
    # attach counts
    out = []
    for pl in lists:
        cnt = await db.price_list_items.count_documents({"price_list_id": pl["id"]})
        dcnt = await db.price_list_category_discounts.count_documents({"price_list_id": pl["id"]})
        ccnt = await db.customers.count_documents({"price_list_id": pl["id"]})
        out.append({**pl, "items_count": cnt, "discounts_count": dcnt, "customers_count": ccnt})
    return out


@api_router.post("/price-lists")
async def create_price_list(body: PriceListIn, admin=Depends(require_admin)):
    name = body.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Name required")
    if await db.price_lists.find_one({"name": name}):
        raise HTTPException(status_code=400, detail="A price list with this name already exists")
    doc = {
        "id": str(uuid.uuid4()),
        "name": name,
        "description": (body.description or "").strip(),
        "bill_amount_required": bool(body.bill_amount_required),
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    await db.price_lists.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.patch("/price-lists/{plid}")
async def update_price_list(plid: str, body: PriceListUpdate, admin=Depends(require_action("edit:priceLists"))):
    update = {k: v for k, v in body.model_dump().items() if v is not None}
    if not update:
        raise HTTPException(status_code=400, detail="No fields to update")
    if "name" in update:
        update["name"] = update["name"].strip()
    update["updated_at"] = now_iso()
    res = await db.price_lists.update_one({"id": plid}, {"$set": update})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Price list not found")
    return await db.price_lists.find_one({"id": plid}, {"_id": 0})


@api_router.delete("/price-lists/{plid}")
async def delete_price_list(plid: str, admin=Depends(require_action("delete:priceLists"))):
    existing = await db.price_lists.find_one({"id": plid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Price list not found")
    refs = await db.customers.count_documents({"price_list_id": plid})
    if refs > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete: {refs} customer(s) are assigned to this price list")
    await db.price_lists.delete_one({"id": plid})
    await db.price_list_items.delete_many({"price_list_id": plid})
    await db.price_list_category_discounts.delete_many({"price_list_id": plid})
    return {"ok": True, "deleted": plid}


@api_router.post("/price-lists/{plid}/clone")
async def clone_price_list(plid: str, body: PriceListCloneIn, admin=Depends(require_admin)):
    """Duplicate an existing price list under a new name, copying every
    per-item price and per-category discount as-is. Customer linkages are
    NOT copied — the new list starts unassigned."""
    src = await db.price_lists.find_one({"id": plid}, {"_id": 0})
    if not src:
        raise HTTPException(status_code=404, detail="Price list not found")
    new_name = (body.name or "").strip()
    if not new_name:
        raise HTTPException(status_code=400, detail="Name required")
    if await db.price_lists.find_one({"name": new_name}):
        raise HTTPException(status_code=400, detail="A price list with this name already exists")

    new_id = str(uuid.uuid4())
    now = now_iso()
    new_desc = body.description
    if new_desc is None:
        base = (src.get("description") or "").strip()
        new_desc = (f"{base} (Copy)" if base else f"Cloned from {src.get('name', '')}").strip()
    new_doc = {
        "id": new_id,
        "name": new_name,
        "description": new_desc,
        "created_at": now,
        "updated_at": now,
    }
    await db.price_lists.insert_one(new_doc)

    # Copy per-item prices
    src_items = await db.price_list_items.find(
        {"price_list_id": plid}, {"_id": 0}
    ).to_list(10000)
    if src_items:
        new_items = []
        for it in src_items:
            it2 = {k: v for k, v in it.items() if k != "_id"}
            it2["id"] = str(uuid.uuid4())
            it2["price_list_id"] = new_id
            it2["updated_at"] = now
            new_items.append(it2)
        await db.price_list_items.insert_many(new_items)

    # Copy per-category discounts
    src_discs = await db.price_list_category_discounts.find(
        {"price_list_id": plid}, {"_id": 0}
    ).to_list(10000)
    if src_discs:
        new_discs = []
        for d in src_discs:
            d2 = {k: v for k, v in d.items() if k != "_id"}
            d2["id"] = str(uuid.uuid4())
            d2["price_list_id"] = new_id
            d2["updated_at"] = now
            new_discs.append(d2)
        await db.price_list_category_discounts.insert_many(new_discs)

    new_doc.pop("_id", None)
    return {
        "ok": True,
        "cloned_from": plid,
        "items_copied": len(src_items),
        "discounts_copied": len(src_discs),
        **new_doc,
    }


@api_router.post("/price-lists/{plid}/delink-customers")
async def delink_price_list_from_customers(plid: str, admin=Depends(require_admin)):
    """Task 3 — Bulk-detach a price list from every customer currently
    assigned to it. The list itself remains intact (items + discounts
    preserved); only the per-customer `price_list_id` references are
    cleared. Returns the number of customers updated."""
    pl = await db.price_lists.find_one({"id": plid}, {"_id": 0, "id": 1, "name": 1})
    if not pl:
        raise HTTPException(status_code=404, detail="Price list not found")
    res = await db.customers.update_many(
        {"price_list_id": plid},
        {"$set": {"price_list_id": None}},
    )
    return {
        "ok": True,
        "list_id": plid,
        "list_name": pl.get("name"),
        "delinked_customers": int(getattr(res, "modified_count", 0) or 0),
    }



@api_router.get("/price-lists/{plid}")
async def get_price_list_detail(plid: str, user=Depends(get_current_user)):
    pl = await db.price_lists.find_one({"id": plid}, {"_id": 0})
    if not pl:
        raise HTTPException(status_code=404, detail="Price list not found")
    items = await db.items.find({}, {"_id": 0}).sort("name", 1).to_list(5000)
    pli_list = await db.price_list_items.find({"price_list_id": plid}, {"_id": 0}).to_list(10000)
    price_map = {p["item_id"]: float(p.get("price") or 0) for p in pli_list}
    rows = []
    for it in items:
        rows.append({
            "item_id": it["id"],
            "item_name": it["name"],
            "product_name": it.get("product_name") or "",
            "price": price_map.get(it["id"], 0.0),
        })
    discounts = await db.price_list_category_discounts.find({"price_list_id": plid}, {"_id": 0}).to_list(500)
    customers_count = await db.customers.count_documents({"price_list_id": plid})
    return {"price_list": pl, "items": rows, "discounts": discounts, "customers_count": customers_count}


@api_router.post("/price-lists/{plid}/items")
async def set_price_list_item(plid: str, body: PriceListItemIn, admin=Depends(require_admin)):
    pl = await db.price_lists.find_one({"id": plid}, {"_id": 0})
    if not pl:
        raise HTTPException(status_code=404, detail="Price list not found")
    item = await db.items.find_one({"id": body.item_id}, {"_id": 0})
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    doc = {
        "price_list_id": plid,
        "item_id": body.item_id,
        "item_name": item["name"],
        "product_name": item.get("product_name") or "",
        "price": float(body.price or 0),
        "updated_at": now_iso(),
    }
    await db.price_list_items.update_one(
        {"price_list_id": plid, "item_id": body.item_id},
        {"$set": doc},
        upsert=True,
    )
    return doc


@api_router.post("/price-lists/{plid}/discounts")
async def set_price_list_category_discount(plid: str, body: CategoryDiscountIn, admin=Depends(require_admin)):
    pl = await db.price_lists.find_one({"id": plid}, {"_id": 0})
    if not pl:
        raise HTTPException(status_code=404, detail="Price list not found")
    if body.discount_type not in ("₹", "%", ""):
        raise HTTPException(status_code=400, detail="discount_type must be '₹', '%' or ''")
    doc = {
        "price_list_id": plid,
        "product_name": body.product_name,
        "discount_value": float(body.discount_value or 0),
        "discount_type": body.discount_type,
        "updated_at": now_iso(),
    }
    await db.price_list_category_discounts.update_one(
        {"price_list_id": plid, "product_name": body.product_name},
        {"$set": doc},
        upsert=True,
    )
    return doc


@api_router.get("/price-lists/{plid}/export")
async def export_price_list(plid: str, user=Depends(get_current_user)):
    """Download Excel with two columns: Item Name | Price (₹)."""
    from openpyxl import Workbook
    from fastapi.responses import StreamingResponse
    pl = await db.price_lists.find_one({"id": plid}, {"_id": 0})
    if not pl:
        raise HTTPException(status_code=404, detail="Price list not found")
    items = await db.items.find({}, {"_id": 0}).sort("name", 1).to_list(5000)
    pli_list = await db.price_list_items.find({"price_list_id": plid}, {"_id": 0}).to_list(10000)
    price_map = {p["item_id"]: float(p.get("price") or 0) for p in pli_list}
    wb = Workbook()
    ws = wb.active
    ws.title = "Prices"
    ws.append(["Item Name", "Price (Rs)"])
    for it in items:
        ws.append([it["name"], price_map.get(it["id"], 0.0)])
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 15
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = "".join(c if c.isalnum() else "_" for c in pl["name"])
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="price_list_{safe_name}.xlsx"'},
    )


@api_router.post("/price-lists/{plid}/import")
async def import_price_list(plid: str, file: UploadFile = File(...), admin=Depends(require_admin)):
    """Upload Excel with rows: Item Name | Price. Matches each row to an
    existing item (case-insensitive, fuzzy fallback ≥85). Unknown rows are
    returned for review. Existing prices are overwritten."""
    from openpyxl import load_workbook
    pl = await db.price_lists.find_one({"id": plid}, {"_id": 0})
    if not pl:
        raise HTTPException(status_code=404, detail="Price list not found")
    blob = await file.read()
    if not blob:
        raise HTTPException(status_code=400, detail="Empty file")
    try:
        wb = load_workbook(io.BytesIO(blob), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel: {e}")
    ws = wb.active
    items = await db.items.find({}, {"_id": 0}).to_list(5000)
    by_lower = {it["name"].strip().lower(): it for it in items}
    item_names = [it["name"] for it in items]
    updated = 0
    unknown: List[Dict[str, Any]] = []
    rows_iter = ws.iter_rows(values_only=True)
    first_row = next(rows_iter, None)
    # Detect header — if first cell is non-numeric and looks like "Item Name"-ish, skip it
    def is_header(r):
        if not r or len(r) < 2:
            return False
        b = r[1]
        if isinstance(b, (int, float)):
            return False
        return True
    if first_row and not is_header(first_row):
        # Treat first row as data
        rows_to_process = [first_row]
    else:
        rows_to_process = []
    rows_to_process.extend(rows_iter)
    for r in rows_to_process:
        if not r or r[0] is None:
            continue
        name = str(r[0]).strip()
        if not name:
            continue
        price_raw = r[1] if len(r) > 1 else None
        try:
            price = float(price_raw) if price_raw is not None and str(price_raw).strip() != "" else 0.0
        except Exception:
            unknown.append({"item_name": name, "reason": f"invalid price '{price_raw}'"})
            continue
        # Match item by exact lowercase first
        match = by_lower.get(name.lower())
        if not match:
            # fuzzy fallback
            best = rf_process.extractOne(name, item_names, scorer=fuzz.WRatio, processor=rf_utils.default_process)
            if best and best[1] >= 85:
                match = next((it for it in items if it["name"] == best[0]), None)
        if not match:
            unknown.append({"item_name": name, "reason": "no matching SKU"})
            continue
        await db.price_list_items.update_one(
            {"price_list_id": plid, "item_id": match["id"]},
            {"$set": {
                "price_list_id": plid,
                "item_id": match["id"],
                "item_name": match["name"],
                "product_name": match.get("product_name") or "",
                "price": price,
                "updated_at": now_iso(),
            }},
            upsert=True,
        )
        updated += 1
    await db.price_lists.update_one({"id": plid}, {"$set": {"updated_at": now_iso()}})
    return {"updated": updated, "unknown_count": len(unknown), "unknown": unknown[:50]}


# ======================== Daily Dispatch Report ========================
@api_router.get("/reports/daily-dispatch")
async def daily_dispatch_report(date: Optional[str] = None, user=Depends(get_current_user)):
    """Consolidated end-of-day report grouped by party (customer).
    `date` is YYYY-MM-DD; defaults to today's IST date.

    Day boundaries use India Standard Time (UTC+5:30) so the report's
    grouping matches the factory's actual working day — same convention
    used by the dispatch auto-merge logic and the single-party ledger
    (which renders `dispatched_at` in the browser's local timezone). This
    avoids the off-by-one mismatch that happened for dispatches recorded
    between 00:00 and 05:30 IST (which fall on the previous UTC date)."""
    IST = timezone(timedelta(hours=5, minutes=30))
    today_ist = datetime.now(IST).date()
    if date:
        try:
            target = datetime.strptime(date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD")
    else:
        target = today_ist
    # Edit-window metadata so the UI can lock/unlock the edit controls.
    settings = await _get_settings_doc()
    edit_window_days = int(settings.get("edit_window_days", 3) or 0)
    is_admin = user.get("role") == "admin"
    if is_blank_view(user):
        return {
            "date": target.isoformat(),
            "groups": [],
            "grand_total_pcs": 0,
            "grand_total_value": 0,
            "dispatch_count": 0,
            "edit_window_days": edit_window_days,
            "is_admin": is_admin,
        }
    now_utc = datetime.now(timezone.utc)
    def _can_edit(d: Dict[str, Any]) -> bool:
        if is_admin:
            return True
        if edit_window_days <= 0:
            return False
        ts = d.get("dispatched_at") or d.get("created_at")
        if not ts:
            return False
        try:
            punched = datetime.fromisoformat(str(ts).replace("Z", "+00:00"))
            if punched.tzinfo is None:
                punched = punched.replace(tzinfo=timezone.utc)
            age_days = (now_utc - punched).total_seconds() / 86400.0
            return age_days <= edit_window_days
        except Exception:
            return False
    # IST day window expressed in UTC for the ISO-string compare against
    # `dispatched_at` (which is stored in UTC).
    start = datetime.combine(target, datetime.min.time(), tzinfo=IST).astimezone(timezone.utc).isoformat()
    end = datetime.combine(target, datetime.max.time(), tzinfo=IST).astimezone(timezone.utc).isoformat()
    dispatches = await db.dispatches.find(
        {"dispatched_at": {"$gte": start, "$lte": end}},
        {"_id": 0},
    ).sort("dispatched_at", 1).to_list(5000)
    # Group by customer
    groups: Dict[str, Dict[str, Any]] = {}
    customers_cache: Dict[str, Dict[str, Any]] = {}
    price_list_cache: Dict[str, Dict[str, Any]] = {}

    async def _price_list_meta(pl_id: Optional[str]) -> Dict[str, Any]:
        """Cache both the display name AND the bill_amount_required toggle
        for each price list so the Dispatch Report can render/skip the
        Bill Amount field per-party without repeated DB round-trips."""
        if not pl_id:
            return {"name": "", "bill_amount_required": True}
        if pl_id in price_list_cache:
            return price_list_cache[pl_id]
        pl = await db.price_lists.find_one(
            {"id": pl_id},
            {"_id": 0, "name": 1, "bill_amount_required": 1},
        )
        meta = {
            "name": (pl or {}).get("name") or "",
            # Default to True so pre-existing lists (without the field)
            # keep the current "bill required" behaviour.
            "bill_amount_required": (pl or {}).get("bill_amount_required", True) if pl else True,
        }
        price_list_cache[pl_id] = meta
        return meta

    async def _price_list_name(pl_id: Optional[str]) -> str:
        return (await _price_list_meta(pl_id))["name"]

    grand_pcs = 0
    grand_value = 0.0
    for d in dispatches:
        cid = d.get("customer_id") or "unknown"
        cust = customers_cache.get(cid)
        if cust is None and cid != "unknown":
            cust = await db.customers.find_one({"id": cid}, {"_id": 0}) or {}
            customers_cache[cid] = cust
        if cust is None:
            cust = {}
        g = groups.get(cid)
        if not g:
            pl_id = cust.get("price_list_id")
            pl_meta = await _price_list_meta(pl_id)
            g = {
                "customer_id": cid,
                "customer_name": d.get("customer_name") or cust.get("name") or "—",
                "transport_name": cust.get("transport_name") or "",
                "phone": cust.get("phone") or "",
                "address": cust.get("address") or "",
                "city": cust.get("city") or "",
                "location": cust.get("location") or "",
                "private_mark": cust.get("private_mark") or "",
                "bill_number_mode": bool(cust.get("bill_number_mode")),
                "price_list_id": pl_id or "",
                "price_list_name": pl_meta["name"],
                # Per-party flag driving the Dispatch Report Bill Amount
                # prompt. When False, the frontend hides the input and
                # skips it in the "Missing / Complete" check.
                "bill_amount_required": bool(pl_meta["bill_amount_required"]),
                "lines": [],
                "dispatches": [],
                "total_pcs": 0,
                "total_value": 0.0,
                "dispatch_count": 0,
            }
            groups[cid] = g
        g["dispatch_count"] += 1
        # Enrich the raw items with live pricing so the client-side Edit
        # dialog opens with the SAME net/unit price the report row shows
        # (older dispatches often have net_unit_price=0 stored — recompute
        # against the customer's current price list so the operator can
        # actually see and edit the correct number). Items with a saved
        # `price_override` are left verbatim so manual edits stick.
        enriched_items: List[Dict[str, Any]] = []
        cust_pl_for_enrich = (cust or {}).get("price_list_id")
        for _raw in (d.get("items") or []):
            _stored_unit = float(_raw.get("unit_price") or 0)
            _stored_net = float(_raw.get("net_unit_price") or 0)
            if _raw.get("price_override") and _stored_net > 0:
                # Operator manually set a price via Edit dialog — respect it.
                eff_unit = _stored_unit if _stored_unit > 0 else _stored_net
                eff_net = _stored_net
            else:
                try:
                    _live = await compute_line_pricing(
                        cust_pl_for_enrich,
                        _raw.get("item_id"),
                        _raw.get("product_name") or "",
                    )
                    _live_unit = float(_live.get("unit_price") or 0)
                    _live_net = float(_live.get("net_unit_price") or _live_unit)
                except Exception:
                    _live_unit = 0.0
                    _live_net = 0.0
                eff_unit = _live_unit if _live_unit > 0 else _stored_unit
                eff_net = _live_net if _live_net > 0 else (_stored_net if _stored_net > 0 else eff_unit)
            enriched_items.append({
                **_raw,
                "unit_price": eff_unit,
                "net_unit_price": eff_net,
            })
        g["dispatches"].append({
            "id": d.get("id"),
            "slip_no": d.get("slip_no"),
            "gr_number": d.get("gr_number") or "",
            "bill_number": d.get("bill_number") or "",
            "gr_date": d.get("gr_date") or "",
            "total_value": float(d.get("total_value") or 0),
            "total_pcs": int(d.get("total_pcs") or 0),
            "dispatched_at": d.get("dispatched_at"),
            "bag_count": int(d.get("bag_count") or 0),
            "items": enriched_items,
            "can_edit": _can_edit(d),
        })
        for line in d.get("items", []):
            qty = int(line.get("quantity") or 0)
            # Manual edit via Edit-Dispatch dialog wins — respect the
            # operator's saved net/unit price verbatim. Otherwise recompute
            # LIVE against the customer's currently-assigned price list +
            # discounts, so the report reflects any list changes.
            if line.get("price_override") and float(line.get("net_unit_price") or 0) > 0:
                unit_price = float(line.get("unit_price") or 0)
                discount_value = float(line.get("discount_value") or 0)
                discount_type = line.get("discount_type") or ""
                net_unit_price = float(line.get("net_unit_price") or unit_price)
            else:
                cust_pl = (cust or {}).get("price_list_id")
                live = await compute_line_pricing(
                    cust_pl,
                    line.get("item_id"),
                    line.get("product_name") or "",
                )
                unit_price = float(live["unit_price"] or line.get("unit_price") or 0)
                discount_value = float(live["discount_value"] or 0)
                discount_type = live["discount_type"] or ""
                net_unit_price = float(live["net_unit_price"] or unit_price)
                # If the customer has no price list assigned now but the
                # dispatch snapshot has values, fall back to the snapshot
                # so historical rows aren't blanked out.
                if not cust_pl:
                    unit_price = float(line.get("unit_price") or 0)
                    discount_value = float(line.get("discount_value") or 0)
                    discount_type = line.get("discount_type") or ""
                    net_unit_price = float(line.get("net_unit_price") or unit_price)
            value = round(net_unit_price * qty, 2)
            g["lines"].append({
                "item_id": line.get("item_id"),
                "item_name": line.get("item_name"),
                "product_name": line.get("product_name"),
                "variant": line.get("variant"),
                "quantity": qty,
                "unit_price": unit_price,
                "discount_value": discount_value,
                "discount_type": discount_type,
                "net_unit_price": net_unit_price,
                "line_value": value,
                "dispatched_at": d.get("dispatched_at"),
                "dispatch_id": d.get("id"),
            })
            g["total_pcs"] += qty
            g["total_value"] += value
            grand_pcs += qty
            grand_value += value
        g["total_value"] = round(g["total_value"], 2)
    out_groups = sorted(groups.values(), key=lambda g: g["customer_name"].lower())
    return {
        "date": target.isoformat(),
        "groups": out_groups,
        "grand_total_pcs": grand_pcs,
        "grand_total_value": round(grand_value, 2),
        "dispatch_count": len(dispatches),
        "edit_window_days": edit_window_days,
        "is_admin": is_admin,
    }


# ======================== Voice Parsing (text-only) ========================
class VoiceParseIn(BaseModel):
    text: str
    customer_hint: Optional[str] = None


@api_router.post("/voice/parse")
async def voice_parse(body: VoiceParseIn, user=Depends(get_current_user)):
    """Parse a free-form Hinglish/English order transcript into structured
    items + a fuzzy-matched customer, WITHOUT calling Whisper. Used when
    the operator types/edits the transcript manually or when an external
    STT pipeline supplies the text. Returns the same shape as
    `/voice/transcribe` minus the `text` field already passed in."""
    text = (body.text or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="Empty text")
    parsed = await parse_voice_order_with_items(text)
    # If the caller passes a customer hint, prefer matching against it so a
    # known party name spoken loosely still resolves correctly. Otherwise
    # fall back to matching against the entire text.
    customer_match = await match_customer_from_voice(body.customer_hint or text)
    return {"text": text, "parsed_items": parsed, "parsed_customer": customer_match}


# ======================== Voice Transcription ========================
@api_router.post("/voice/transcribe")
async def voice_transcribe(file: UploadFile = File(...), user=Depends(get_current_user)):
    # Validate input BEFORE attempting external service call — empty audio must
    # return 400, not 500, even when the LLM key is missing.
    audio_bytes = await file.read()
    if not audio_bytes:
        raise HTTPException(status_code=400, detail="Empty audio file")
    try:
        from emergentintegrations.llm.openai import OpenAISpeechToText
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Speech library not available: {e}")
    api_key = os.environ.get("EMERGENT_LLM_KEY")
    if not api_key:
        raise HTTPException(status_code=503, detail="EMERGENT_LLM_KEY not configured")

    stt = OpenAISpeechToText(api_key=api_key)
    file_like = io.BytesIO(audio_bytes)
    file_like.name = file.filename or "audio.webm"
    # Build a biasing prompt that includes a sample of real party names so
    # Whisper transcribes them spelled correctly instead of phonetic noise.
    try:
        party_sample = await db.customers.find({}, {"_id": 0, "name": 1}).limit(60).to_list(60)
        party_names = ", ".join([c["name"] for c in party_sample if c.get("name")])
    except Exception:
        party_names = ""
    bias_prompt = (
        "Factory order for two-wheeler spare parts dictated in Hinglish / English / Hindi. "
        "Products: side stand, center stand with kit, center stand without kit, center stand pin, "
        "footrest rod, seat kunda, side seat handle, lady footrest, side footrest, handlebar, "
        "number plate, engine plate, v-bracket, luggage rod, side mirror clump, rear seat handle. "
        "Quantities in pieces, e.g. do sau, teen sau, char sau, paanch sau, ek hazaar."
    )
    if party_names:
        bias_prompt += f" Customer parties may include: {party_names}."
    try:
        resp = await stt.transcribe(
            file=file_like,
            model="whisper-1",
            response_format="json",
            prompt=bias_prompt,
            temperature=0.0,
        )
        text = getattr(resp, "text", None) or (resp.get("text") if isinstance(resp, dict) else None) or str(resp)
    except Exception as e:
        logger.exception("Transcription failed")
        raise HTTPException(status_code=500, detail=f"Transcription error: {e}")

    # Parse text into order items
    parsed = await parse_voice_order_with_items(text)
    customer_match = await match_customer_from_voice(text)
    return {"text": text, "parsed_items": parsed, "parsed_customer": customer_match}


# ======================== Voice AI Agent ========================
# A general-purpose multilingual voice command agent. Captures audio,
# transcribes it via whisper-1, then asks an LLM to map the transcript
# to a structured intent the frontend can dispatch.
VOICE_AGENT_INTENT_SCHEMA = """
You are the voice command interpreter for "JK Products", a factory order
management web app for two-wheeler spare parts. Operators speak commands
in English, Hindi or Hinglish (Hindi written in Roman/English script).
Convert the transcript into ONE structured JSON action. Respond with
JSON ONLY (no markdown, no commentary).

Schema:
{
  "intent": "<one of the intents below>",
  "params": { ... intent-specific parameters ... },
  "spoken_reply": "<one short sentence to acknowledge the user, in the language they used>",
  "confidence": <0..1 float>
}

# Navigation / read-only intents
- "navigate"                → params.page is one of:
   "dashboard", "orders", "new_order", "dispatch", "dispatch_ledger",
   "customers", "products", "raw_materials", "vendor_ledger",
   "vendor_price_lists", "price_lists", "daily_report", "admin_users",
   "admin_settings", "login_attestations", "purchase_center", "suppliers"
- "search_customer"          → params.query (string)
- "search_product"           → params.query (string)
- "show_customer_ledger"     → params.customer (customer name as spoken)
- "show_vendor_ledger"       → params.vendor (vendor name as spoken)
- "filter_orders"            → params.status one of "Pending","Dispatched","Cleared","All"
- "help"                     → params {}
- "unknown"                  → params {}

# Read-aloud (Q&A) intents — backend will fetch data and speak the answer
- "query_closing_balance"    → params.customer (customer name)
- "query_stock"              → params.item (product / item name)
- "query_daily_summary"      → params {} (today's totals)
- "query_pending_count"      → params {} (number of pending orders)
- "query_vendor_balance"     → params.vendor (vendor name)

# Mutation intents — backend resolves entities; frontend asks for
# confirmation before calling the real API
- "record_customer_payment"  → params.customer, params.amount (number, rupees),
                                params.source ("cash"|"upi"|"bank_transfer"|"neft"|"rtgs"|"cheque"|"card"|"adjustment"|"other"),
                                params.reference (optional)
- "record_supplier_purchase" → params.vendor, params.amount, params.bill_number (opt),
                                params.material (opt, free text), params.notes (opt)
- "record_supplier_payment"  → params.vendor, params.amount, params.source, params.reference (opt)
- "update_order_status"      → params.order_ref (slip/order id or last n digits),
                                params.new_status ("Pending"|"Dispatched"|"Cleared")
- "set_private_mark"         → params.customer, params.private_mark
- "prefill_stock_match"      → params.items: [{"name":<item>, "quantity":<int>}] (opens dispatch w/ pre-fill)
- "prefill_new_order"        → params.customer, params.items: [{"name","quantity"}]
- "add_customer"             → params.name, params.phone (opt), params.city (opt), params.address (opt)
- "add_supplier"             → params.name, params.phone (opt), params.city (opt),
                                params.material_category (opt)
- "delete_dispatch"          → params.dispatch_ref
- "delete_payment"           → params.payment_ref
- "update_price"             → params.item, params.new_price (number)

Number parsing (CRITICAL):
- "do hazaar" or "two thousand" → 2000
- "das hazaar" or "ten thousand" → 10000
- "paanch sau" or "five hundred" → 500
- "ek lakh" or "one lakh" → 100000
- "saade teen sau" → 350
- "char hazaar paanch sau" or "forty five hundred" → 4500
- Always emit `amount` and `quantity` as numbers (no commas, no units).

Examples:
- "Open dispatch"                              → {"intent":"navigate","params":{"page":"dispatch"},"spoken_reply":"Opening dispatch.","confidence":0.95}
- "Sharma Auto se das hazaar cash mila"        → {"intent":"record_customer_payment","params":{"customer":"Sharma Auto","amount":10000,"source":"cash"},"spoken_reply":"Sharma Auto se 10000 rupee cash record karne ke liye confirm karein.","confidence":0.9}
- "Record payment of 2500 UPI from A M Auto"   → {"intent":"record_customer_payment","params":{"customer":"A M Auto","amount":2500,"source":"upi"},"spoken_reply":"A M Auto ka 2500 rupee UPI payment confirm karein.","confidence":0.9}
- "Steel Traders ko paanch hazaar UPI diya"    → {"intent":"record_supplier_payment","params":{"vendor":"Steel Traders","amount":5000,"source":"upi"},"spoken_reply":"Steel Traders ko 5000 UPI payment confirm karein.","confidence":0.9}
- "Naya purchase Steel Traders se 12000 ka, bill 234, MS rod"
                                               → {"intent":"record_supplier_purchase","params":{"vendor":"Steel Traders","amount":12000,"bill_number":"234","material":"MS rod"},"spoken_reply":"Steel Traders ka 12000 rupee purchase confirm karein.","confidence":0.88}
- "Side stand char sau pieces dispatch karo"   → {"intent":"prefill_stock_match","params":{"items":[{"name":"side stand","quantity":400}]},"spoken_reply":"Dispatch form me 400 side stand bhar raha hoon. Match dabaayein.","confidence":0.9}
- "A M Auto ke liye center stand do sau aur side stand teen sau"
                                               → {"intent":"prefill_new_order","params":{"customer":"A M Auto","items":[{"name":"center stand","quantity":200},{"name":"side stand","quantity":300}]},"spoken_reply":"A M Auto ka naya order bhar raha hoon. Confirm karein.","confidence":0.9}
- "Order ABC123 ko dispatched mark karo"       → {"intent":"update_order_status","params":{"order_ref":"ABC123","new_status":"Dispatched"},"spoken_reply":"Order ABC123 ko Dispatched mark karne ke liye confirm karein.","confidence":0.88}
- "Sharma Auto ko private mark RG laga do"     → {"intent":"set_private_mark","params":{"customer":"Sharma Auto","private_mark":"RG"},"spoken_reply":"Sharma Auto ka private mark RG set karne ke liye confirm karein.","confidence":0.85}
- "Add new customer Ramesh Auto, Ludhiana, phone 9812345678"
                                               → {"intent":"add_customer","params":{"name":"Ramesh Auto","city":"Ludhiana","phone":"9812345678"},"spoken_reply":"Naya customer Ramesh Auto add karne ke liye confirm karein.","confidence":0.85}
- "Side stand without kit ka rate 640 kar do"  → {"intent":"update_price","params":{"item":"side stand without kit","new_price":640},"spoken_reply":"Side stand without kit ka rate 640 karne ke liye confirm karein.","confidence":0.82}
- "Sharma Auto ka closing balance kya hai"     → {"intent":"query_closing_balance","params":{"customer":"Sharma Auto"},"spoken_reply":"Calculating.","confidence":0.92}
- "Side stand kitna stock hai"                 → {"intent":"query_stock","params":{"item":"side stand"},"spoken_reply":"Checking stock.","confidence":0.9}
- "Aaj ka summary sunao"                       → {"intent":"query_daily_summary","params":{},"spoken_reply":"Aaj ka summary nikal raha hoon.","confidence":0.9}
- "Kitne pending orders hain"                  → {"intent":"query_pending_count","params":{},"spoken_reply":"Counting pending orders.","confidence":0.9}
- "Delete last dispatch"                       → {"intent":"delete_dispatch","params":{"dispatch_ref":"last"},"spoken_reply":"Last dispatch ko delete karne ke liye confirm karein.","confidence":0.8}

Rules:
1. Respond with valid JSON only, nothing else.
2. Pick the BEST single intent. If unsure, use "unknown".
3. Always include "spoken_reply" in the user's spoken language (English / Hindi / Hinglish).
4. Never invent customer or vendor names — copy the user's words.
5. For payments and purchases ALWAYS emit `amount` as a plain JSON number.
6. For mutations, the spoken_reply MUST ask the user to confirm (kyunki yeh data change karega).
"""


class VoiceAgentTextIn(BaseModel):
    """Allows the agent to be invoked with already-transcribed text
    (useful for typed fallback / accessibility)."""
    text: str
    language_hint: Optional[str] = None


async def _classify_voice_command(transcript: str) -> Dict[str, Any]:
    """Send the transcript to the LLM and parse a structured intent."""
    import json as _json
    from emergentintegrations.llm.chat import LlmChat, UserMessage

    api_key = os.environ.get("EMERGENT_LLM_KEY")
    if not api_key:
        raise HTTPException(status_code=503, detail="EMERGENT_LLM_KEY not configured")

    chat = (
        LlmChat(
            api_key=api_key,
            session_id=f"voice-agent-{uuid.uuid4()}",
            system_message=VOICE_AGENT_INTENT_SCHEMA,
        )
        # Claude understands Hindi / English / Hinglish far better than the
        # previous model, which is critical for accurate voice-command intent.
        .with_model("anthropic", "claude-sonnet-4-6")
    )

    try:
        reply = await chat.send_message(
            UserMessage(text=f"Transcript: {transcript}\n\nReturn JSON only.")
        )
    except Exception as e:
        logger.exception("Voice agent classification failed")
        raise HTTPException(status_code=502, detail=f"LLM error: {e}")

    raw = reply if isinstance(reply, str) else getattr(reply, "content", str(reply))
    # Strip code fences if any
    cleaned = raw.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned)
        cleaned = re.sub(r"\s*```$", "", cleaned)
    try:
        parsed = _json.loads(cleaned)
    except Exception:
        logger.warning("Voice agent returned non-JSON: %s", raw[:300])
        parsed = {"intent": "unknown", "params": {}, "spoken_reply": "Sorry, I did not understand. Please try again.", "confidence": 0.0}

    intent = parsed.get("intent") or "unknown"
    params = parsed.get("params") or {}
    spoken_reply = parsed.get("spoken_reply") or ""
    confidence = float(parsed.get("confidence") or 0.0)

    # Resolve customer / vendor names against the DB so the frontend can
    # navigate straight to the matched record.
    resolved: Dict[str, Any] = {}

    customer_intents = {
        "show_customer_ledger", "search_customer", "create_order",
        "record_customer_payment", "set_private_mark", "prefill_new_order",
        "query_closing_balance", "add_customer",
    }
    vendor_intents = {
        "show_vendor_ledger", "record_supplier_payment", "record_supplier_purchase",
        "query_vendor_balance", "add_supplier",
    }

    async def _resolve_vendor(name: str) -> Optional[Dict[str, Any]]:
        if not name:
            return None
        vendors = await db.suppliers.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(2000)
        if not vendors:
            return None
        names = [v["name"] for v in vendors]
        lower = name.lower()
        best = None
        for i, n in enumerate(names):
            s = max(
                fuzz.token_set_ratio(lower, n.lower()),
                fuzz.partial_ratio(lower, n.lower()) if len(n) >= 4 else 0,
            )
            if best is None or s > best[1]:
                best = (i, s)
        if best and best[1] >= 65:
            idx, score = best
            return {"id": vendors[idx]["id"], "name": vendors[idx]["name"], "score": int(score)}
        return None

    if intent in customer_intents:
        name = params.get("customer") or params.get("query")
        # add_customer creates a NEW one — don't try to match the typed name
        if intent != "add_customer" and name:
            match = await match_customer_from_voice(name)
            if match:
                resolved["customer_id"] = match["id"]
                resolved["customer_name"] = match["name"]
                resolved["customer_score"] = match["score"]
    if intent in vendor_intents:
        name = params.get("vendor") or params.get("query")
        if intent != "add_supplier" and name:
            match = await _resolve_vendor(name)
            if match:
                resolved["vendor_id"] = match["id"]
                resolved["vendor_name"] = match["name"]
                resolved["vendor_score"] = match["score"]

    # Item / product name fuzzy match for prefill / stock / price intents
    async def _resolve_item(name: str) -> Optional[Dict[str, Any]]:
        if not name:
            return None
        # First try the SYNONYM_MAP for a canonical product name
        ln = (name or "").lower().strip()
        canonical = SYNONYM_MAP.get(ln)
        # Search items collection by item_name / product_name
        all_items = await db.items.find({}, {"_id": 0, "id": 1, "item_name": 1, "product_name": 1}).to_list(5000)
        if not all_items:
            return None
        # Match against item_name first, then product_name
        candidates_name = [i.get("item_name", "") for i in all_items]
        candidates_prod = [i.get("product_name", "") for i in all_items]
        target = canonical or name
        best = None
        for i, (iname, pname) in enumerate(zip(candidates_name, candidates_prod)):
            s = max(
                fuzz.token_set_ratio(target.lower(), iname.lower()),
                fuzz.token_set_ratio(target.lower(), pname.lower()),
            )
            if best is None or s > best[1]:
                best = (i, s)
        if best and best[1] >= 65:
            idx, score = best
            it = all_items[idx]
            return {
                "id": it["id"],
                "item_name": it.get("item_name") or "",
                "product_name": it.get("product_name") or "",
                "score": int(score),
            }
        return None

    if intent in ("query_stock", "update_price"):
        m = await _resolve_item(params.get("item") or "")
        if m:
            resolved["item_id"] = m["id"]
            resolved["item_name"] = m["item_name"]
            resolved["product_name"] = m["product_name"]
            resolved["item_score"] = m["score"]

    if intent in ("prefill_stock_match", "prefill_new_order"):
        in_items = params.get("items") or []
        out_items = []
        for it in in_items:
            nm = (it or {}).get("name") or ""
            qty = (it or {}).get("quantity")
            try:
                qty = int(float(qty)) if qty is not None else 0
            except Exception:
                qty = 0
            m = await _resolve_item(nm)
            out_items.append({
                "name": nm,
                "quantity": qty,
                "item_id": m["id"] if m else None,
                "item_name": m["item_name"] if m else nm,
                "product_name": m["product_name"] if m else "",
                "score": m["score"] if m else 0,
            })
        resolved["items"] = out_items

    # Order-ref → real order id (fuzzy match by slip_no or id prefix)
    if intent in ("update_order_status", "delete_dispatch", "delete_payment"):
        ref = (params.get("order_ref") or params.get("dispatch_ref") or params.get("payment_ref") or "").strip()
        if ref and ref.lower() != "last":
            # Try to find by order id prefix (first 8 chars) or slip_no
            ref_norm = ref.upper().replace("-", "").replace("#", "").strip()
            order = await db.orders.find_one(
                {"$or": [
                    {"id": {"$regex": f"^{re.escape(ref.lower())}", "$options": "i"}},
                ]},
                {"_id": 0, "id": 1, "customer_name": 1, "status": 1, "total_value": 1},
            )
            if order:
                resolved["order_id"] = order["id"]
                resolved["order_customer_name"] = order.get("customer_name")
                resolved["order_status"] = order.get("status")
            # Try dispatches by slip_no
            if intent == "delete_dispatch":
                try:
                    seq = int(ref_norm)
                    disp = await db.dispatches.find_one({"slip_no": seq}, {"_id": 0, "id": 1, "customer_name": 1, "total_value": 1})
                    if disp:
                        resolved["dispatch_id"] = disp["id"]
                        resolved["dispatch_customer_name"] = disp.get("customer_name")
                except (ValueError, TypeError):
                    pass
            if intent == "delete_payment":
                try:
                    seq = int(ref_norm)
                    pay = await db.payments.find_one({"receipt_no": seq}, {"_id": 0, "id": 1, "customer_name": 1, "amount": 1})
                    if pay:
                        resolved["payment_id"] = pay["id"]
                        resolved["payment_customer_name"] = pay.get("customer_name")
                except (ValueError, TypeError):
                    pass
        elif ref.lower() == "last":
            # "delete last dispatch" → most recent dispatch by created_at
            if intent == "delete_dispatch":
                last_disp = await db.dispatches.find_one(
                    {}, {"_id": 0, "id": 1, "customer_name": 1, "slip_no": 1, "total_value": 1, "dispatched_at": 1},
                    sort=[("dispatched_at", -1)],
                )
                if last_disp:
                    resolved["dispatch_id"] = last_disp["id"]
                    resolved["dispatch_customer_name"] = last_disp.get("customer_name")
                    resolved["dispatch_slip_no"] = last_disp.get("slip_no")
            if intent == "delete_payment":
                last_pay = await db.payments.find_one(
                    {}, {"_id": 0, "id": 1, "customer_name": 1, "amount": 1, "paid_at": 1},
                    sort=[("paid_at", -1)],
                )
                if last_pay:
                    resolved["payment_id"] = last_pay["id"]
                    resolved["payment_customer_name"] = last_pay.get("customer_name")

    # ─── Read-aloud (Q&A) intents — fetch the answer and embed in spoken_reply ───
    if intent == "query_closing_balance" and resolved.get("customer_id"):
        cid = resolved["customer_id"]
        dispatches = await db.dispatches.find(
            {"customer_id": cid}, {"_id": 0, "total_value": 1}
        ).to_list(20000)
        payments = await db.payments.find(
            {"customer_id": cid}, {"_id": 0, "amount": 1}
        ).to_list(20000)
        sale_returns = await db.sale_returns.find(
            {"customer_id": cid}, {"_id": 0, "amount": 1}
        ).to_list(20000)
        debit = sum(float(d.get("total_value") or 0) for d in dispatches)
        credit = (
            sum(float(p.get("amount") or 0) for p in payments)
            + sum(float(r.get("amount") or 0) for r in sale_returns)
        )
        outstanding = round(debit - credit, 2)
        resolved["closing_balance"] = outstanding
        resolved["total_debit"] = round(debit, 2)
        resolved["total_credit"] = round(credit, 2)
        if outstanding >= 0:
            spoken_reply = f"{resolved['customer_name']} ka closing balance hai ₹{outstanding:,.0f} due."
        else:
            spoken_reply = f"{resolved['customer_name']} ka closing balance hai ₹{abs(outstanding):,.0f} advance."

    if intent == "query_vendor_balance" and resolved.get("vendor_id"):
        sid = resolved["vendor_id"]
        purchases = await db.supplier_purchases.find(
            {"supplier_id": sid}, {"_id": 0, "amount": 1}
        ).to_list(20000)
        sup_payments = await db.supplier_payments.find(
            {"supplier_id": sid}, {"_id": 0, "amount": 1}
        ).to_list(20000)
        purchase_returns = await db.purchase_returns.find(
            {"supplier_id": sid}, {"_id": 0, "amount": 1}
        ).to_list(20000)
        debit = sum(float(p.get("amount") or 0) for p in purchases)
        credit = (
            sum(float(p.get("amount") or 0) for p in sup_payments)
            + sum(float(r.get("amount") or 0) for r in purchase_returns)
        )
        outstanding = round(debit - credit, 2)
        resolved["closing_balance"] = outstanding
        if outstanding >= 0:
            spoken_reply = f"{resolved['vendor_name']} ko ₹{outstanding:,.0f} dene hain."
        else:
            spoken_reply = f"{resolved['vendor_name']} se ₹{abs(outstanding):,.0f} advance hai."

    if intent == "query_stock" and resolved.get("item_id"):
        # Sum allocations vs. order items? "Stock" is loosely defined here —
        # we compute remaining quantity across pending orders (i.e. how much
        # demand is open) since this app doesn't track raw inventory at item
        # level. Use the items collection's reported stock if present.
        it = await db.items.find_one({"id": resolved["item_id"]}, {"_id": 0})
        stock = 0
        if it:
            stock = int(it.get("stock_qty") or 0)
        # Also count open pending demand for this SKU
        pending = await db.orders.find({"status": "Pending"}, {"_id": 0, "items": 1}).to_list(20000)
        demand = 0
        for o in pending:
            for line in o.get("items") or []:
                if line.get("item_id") == resolved["item_id"]:
                    demand += int(line.get("quantity") or 0)
        resolved["stock_qty"] = stock
        resolved["open_demand"] = demand
        display_name = resolved.get("item_name") or resolved.get("product_name") or "Item"
        spoken_reply = f"{display_name} ka pending demand hai {demand} pieces."
        if stock:
            spoken_reply = f"{display_name} ka stock hai {stock}, pending demand {demand}."

    if intent == "query_daily_summary":
        today_iso = datetime.now(timezone.utc).date().isoformat()
        d_today = await db.dispatches.find(
            {"dispatched_at": {"$gte": today_iso}}, {"_id": 0, "total_value": 1, "total_pcs": 1}
        ).to_list(20000)
        p_today = await db.payments.find(
            {"paid_at": {"$gte": today_iso}}, {"_id": 0, "amount": 1}
        ).to_list(20000)
        d_count = len(d_today)
        d_value = sum(float(d.get("total_value") or 0) for d in d_today)
        d_pcs = sum(int(d.get("total_pcs") or 0) for d in d_today)
        p_count = len(p_today)
        p_amount = sum(float(p.get("amount") or 0) for p in p_today)
        resolved["dispatch_count"] = d_count
        resolved["dispatch_value"] = round(d_value, 2)
        resolved["dispatch_pcs"] = d_pcs
        resolved["payment_count"] = p_count
        resolved["payment_amount"] = round(p_amount, 2)
        spoken_reply = (
            f"Aaj ke {d_count} dispatches, total {d_pcs} pieces, value ₹{d_value:,.0f}. "
            f"Payments {p_count}, total ₹{p_amount:,.0f}."
        )

    if intent == "query_pending_count":
        pending = await db.orders.count_documents({"status": "Pending"})
        resolved["pending_count"] = pending
        spoken_reply = f"Abhi {pending} pending orders hain."

    return {
        "intent": intent,
        "params": params,
        "resolved": resolved,
        "spoken_reply": spoken_reply,
        "confidence": confidence,
    }


@api_router.post("/voice/agent")
async def voice_agent(file: UploadFile = File(...), user=Depends(get_current_user)):
    """Transcribe a voice command and return a structured intent the
    frontend can dispatch (navigate, filter, search, etc.)."""
    audio_bytes = await file.read()
    if not audio_bytes:
        raise HTTPException(status_code=400, detail="Empty audio file")
    try:
        from emergentintegrations.llm.openai import OpenAISpeechToText
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Speech library not available: {e}")
    api_key = os.environ.get("EMERGENT_LLM_KEY")
    if not api_key:
        raise HTTPException(status_code=503, detail="EMERGENT_LLM_KEY not configured")

    stt = OpenAISpeechToText(api_key=api_key)
    file_like = io.BytesIO(audio_bytes)
    file_like.name = file.filename or "voice-agent.webm"
    bias_prompt = (
        "Factory order management voice command in English, Hindi or Hinglish. "
        "Common intents: open dispatch, open orders, naya order, customer ledger, "
        "vendor ledger, find customer, search product. Customer names may sound like "
        "M Auto, A M Auto, Sharma Auto, Ramesh Sons, etc."
    )
    try:
        resp = await stt.transcribe(
            file=file_like,
            model="whisper-1",
            response_format="json",
            prompt=bias_prompt,
            temperature=0.0,
        )
        text = getattr(resp, "text", None) or (resp.get("text") if isinstance(resp, dict) else None) or str(resp)
    except Exception as e:
        logger.exception("Voice agent transcription failed")
        raise HTTPException(status_code=500, detail=f"Transcription error: {e}")

    if not text or not text.strip():
        return {
            "transcript": "",
            "intent": "unknown",
            "params": {},
            "resolved": {},
            "spoken_reply": "I did not catch that. Please try again.",
            "confidence": 0.0,
        }

    out = await _classify_voice_command(text)
    return {"transcript": text, **out}


@api_router.post("/voice/agent/text")
async def voice_agent_text(body: VoiceAgentTextIn, user=Depends(get_current_user)):
    """Typed-text fallback for the voice agent (no audio capture)."""
    text = (body.text or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="Empty command")
    out = await _classify_voice_command(text)
    return {"transcript": text, **out}


async def match_customer_from_voice(text: str) -> Optional[Dict[str, Any]]:
    """Fuzzy-match a customer name out of a free-form voice transcript.

    The transcript usually starts with something like "<party name> ke liye ..."
    or "<party name> ko ..." or just dictates a list of items. We compare the
    *whole* transcript against every saved customer name using rapidfuzz's
    `token_set_ratio` (handles word reordering, partial mentions, transliteration
    noise from Whisper). A score ≥ 65 is treated as a confident hit so floor
    operators can correct the rest by ear.
    """
    if not text or not text.strip():
        return None
    customers = await db.customers.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(5000)
    if not customers:
        return None
    names = [c["name"] for c in customers]
    lower_text = text.lower()
    lower_names = [n.lower() for n in names]
    # Strip product/SKU vocabulary so it doesn't drown the customer signal.
    stripped = lower_text
    for key in SYNONYM_MAP.keys():
        stripped = stripped.replace(key, " ")
    # token_set_ratio is robust to extra/unordered words. For longer party
    # names we also consider partial_ratio (handles "ramesh auto sons" being
    # mentioned as just "ramesh sons"), but only when the saved name is long
    # enough that partial_ratio won't trivially match random short substrings.
    best = None
    for i, n in enumerate(lower_names):
        s1 = fuzz.token_set_ratio(stripped, n)
        s2 = fuzz.partial_ratio(stripped, n) if len(n) >= 6 else 0
        score = max(s1, s2)
        if best is None or score > best[1]:
            best = (i, score)
    # 70 is the sweet spot: low enough to catch heavily-mangled Whisper output,
    # high enough that "xyzxyz random" doesn't false-positive on a 3-letter
    # party like "CAD".
    if not best or best[1] < 70:
        return None
    idx, score = best
    return {"id": customers[idx]["id"], "name": names[idx], "score": int(score)}


# Hinglish + English word-number lookups for voice-order quantity parsing.
_WORD_NUMS = {
    # English ones
    "zero": 0, "one": 1, "two": 2, "three": 3, "four": 4, "five": 5,
    "six": 6, "seven": 7, "eight": 8, "nine": 9, "ten": 10,
    "eleven": 11, "twelve": 12, "fifteen": 15, "twenty": 20, "thirty": 30,
    "forty": 40, "fifty": 50, "sixty": 60, "seventy": 70, "eighty": 80, "ninety": 90,
    "hundred": 100, "thousand": 1000,
    # Hinglish ones / tens / hundreds
    "ek": 1, "do": 2, "teen": 3, "char": 4, "chaar": 4, "paanch": 5, "panch": 5,
    "che": 6, "chhe": 6, "saat": 7, "aath": 7, "nau": 9, "das": 10, "dus": 10,
    "bees": 20, "tees": 30, "chalis": 40, "pachas": 50,
    "saath": 60, "sattar": 70, "assi": 80, "nabbe": 90,
    "sau": 100, "hazar": 1000, "hazaar": 1000,
    # Common Whisper misspellings of Hinglish numbers
    "dosa": 200,        # "do sau" merged
    "doso": 200,
    "tinso": 300, "tinsau": 300, "teensau": 300,  # "teen sau"
    "charsau": 400, "chaarsau": 400, "charso": 400,
    "panchsau": 500, "paanchsau": 500, "paanchso": 500, "panchso": 500,
    "chesau": 600, "chhesau": 600, "cheso": 600, "chheso": 600,
    "saatsau": 700, "saatso": 700,
    "aathsau": 800, "aathso": 800,
    "nausau": 900, "nauso": 900,
    # Devanagari numerals (Whisper sometimes returns these for pure-Hindi input)
    "०": 0, "१": 1, "२": 2, "३": 3, "४": 4, "५": 5,
    "६": 6, "७": 7, "८": 8, "९": 9,
    # Hindi number words (Devanagari)
    "एक": 1, "दो": 2, "तीन": 3, "चार": 4, "पाँच": 5, "पांच": 5,
    "छह": 6, "छः": 6, "सात": 7, "आठ": 8, "नौ": 9, "दस": 10,
    "बीस": 20, "तीस": 30, "चालीस": 40, "पचास": 50,
    "साठ": 60, "सत्तर": 70, "अस्सी": 80, "नब्बे": 90,
    "सौ": 100, "हज़ार": 1000, "हजार": 1000,
}


def _parse_word_number(phrase: str) -> int:
    """Parse a small word-number phrase like 'do sau' (200), 'two hundred' (200),
    'teen sau pachas' (350), or pure Hindi 'दो सौ' (200). Falls back to 0 if
    nothing matches."""
    import re
    # Keep Latin letters AND Devanagari (\u0900-\u097F) so Hindi tokens survive
    # the cleanup. Hyphens become spaces so "two-hundred" → ["two", "hundred"].
    cleaned = re.sub(r"[^a-z\u0900-\u097F\s]", "", phrase.lower().replace("-", " "))
    tokens = [t for t in cleaned.split() if t]
    total = 0
    current = 0
    matched = False
    for t in tokens:
        if t not in _WORD_NUMS:
            continue
        matched = True
        v = _WORD_NUMS[t]
        if v >= 100:
            current = (current or 1) * v
            total += current
            current = 0
        else:
            current += v
    total += current
    return total if matched else 0


def parse_voice_order(text: str) -> List[Dict[str, Any]]:
    """Parse Hinglish voice transcript into [{product_name, quantity, matched_phrase}].

    Splits text by clause boundaries (, . ; ! ?) so a number can't bleed across
    clauses. Within each clause:
      - find non-overlapping product hits (longest synonym wins)
      - collect digit + Hinglish/English word-number candidates
      - assign each candidate to its NEAREST product hit; product takes its
        closest assigned candidate.
    """
    import re
    out: List[Dict[str, Any]] = []
    if not text:
        return out

    sorted_keys = sorted(SYNONYM_MAP.keys(), key=lambda x: -len(x))

    # Normalise whitespace and strip stray punctuation that breaks word
    # boundaries (Whisper sometimes injects double-quotes / Hindi danda etc).
    text = re.sub(r"[\"'`]+", " ", text)
    text = text.replace("।", ".")  # Hindi danda → period

    # Split into clauses
    clauses = re.split(r"[,.;!?\n]+", text.lower())

    for clause in clauses:
        if not clause.strip():
            continue
        # 1. find product hits in clause
        hits = []
        used_spans = []
        for key in sorted_keys:
            for m in re.finditer(re.escape(key), clause):
                s, e = m.span()
                if any(us <= s < ue or us < e <= ue for (us, ue) in used_spans):
                    continue
                used_spans.append((s, e))
                hits.append((s, e, key))
        if not hits:
            continue
        hits.sort(key=lambda x: x[0])

        # 2. collect quantity candidates within the clause
        def _in_product(pos: int) -> bool:
            return any(s <= pos < e for (s, e, _k) in hits)

        candidates: List[tuple] = []
        for m in re.finditer(r"\d{1,5}", clause):
            if not _in_product(m.start()):
                candidates.append((m.start(), int(m.group(0))))
        token_iter = list(re.finditer(r"[a-z\u0900-\u097F]+", clause))
        i_tok = 0
        while i_tok < len(token_iter):
            tok = token_iter[i_tok].group(0)
            if tok in _WORD_NUMS and not _in_product(token_iter[i_tok].start()):
                run_start = token_iter[i_tok].start()
                run_end = token_iter[i_tok].end()
                j = i_tok + 1
                while (
                    j < len(token_iter)
                    and token_iter[j].group(0) in _WORD_NUMS
                    and not _in_product(token_iter[j].start())
                ):
                    run_end = token_iter[j].end()
                    j += 1
                qty = _parse_word_number(clause[run_start:run_end])
                if qty > 0:
                    candidates.append((run_start, qty))
                i_tok = j
            else:
                i_tok += 1

        # 3. assign each candidate to nearest product hit using greedy 1-to-1
        # matching. Plain "nearest product" was incorrect: a long sentence like
        # "I need 500 side stands and three hundred footrest rods" causes BOTH
        # 500 and 300 to be distance-closer to side_stand than to footrest_rod,
        # leaving footrest_rod with qty 0. Greedy 1-to-1 fixes this: pair the
        # globally-shortest (candidate, product) edge first, then remove both
        # from the pool, then the next shortest, etc.
        def _dist(pos: int, s: int, e: int) -> int:
            if pos < s:
                return s - pos
            if pos >= e:
                return pos - e + 1
            return 0

        edges = []  # (distance, cand_index, hit_index, qty)
        for ci, (pos, qty) in enumerate(candidates):
            for hi, (s, e, _k) in enumerate(hits):
                edges.append((_dist(pos, s, e), ci, hi, qty))
        edges.sort(key=lambda x: x[0])

        product_qty: Dict[int, int] = {}
        used_cands: set = set()
        used_hits: set = set()
        for _d, ci, hi, qty in edges:
            if ci in used_cands or hi in used_hits:
                continue
            product_qty[hi] = qty
            used_cands.add(ci)
            used_hits.add(hi)
            if len(used_hits) == len(hits) or len(used_cands) == len(candidates):
                break

        for i, (s, e, key) in enumerate(hits):
            out.append({"product_name": SYNONYM_MAP[key], "quantity": product_qty.get(i, 0), "matched_phrase": key})

    return out


async def parse_voice_order_with_items(text: str) -> List[Dict[str, Any]]:
    """Wrap parse_voice_order() and try to match a specific item SKU within
    the surrounding text for each detected product hit. If no item match is
    strong enough, returns the row with item_id/item_name as None — the UI
    will prompt the user to pick an item before submit.
    """
    import re
    base = parse_voice_order(text)
    if not base:
        return base
    # Pull all items from DB once
    all_items = await db.items.find({}, {"_id": 0}).to_list(5000)
    by_product: Dict[str, List[Dict[str, Any]]] = {}
    for it in all_items:
        by_product.setdefault(it["product_name"], []).append(it)

    lower_text = (text or "").lower()
    for row in base:
        pname = row.get("product_name")
        candidates = by_product.get(pname) or []
        if not candidates:
            row["item_id"] = None
            row["item_name"] = None
            continue
        names = [c["name"] for c in candidates]
        # Use full text as query; rapidfuzz token_set/WRatio handles short queries vs long candidates
        match = rf_process.extractOne(lower_text, [n.lower() for n in names], scorer=fuzz.token_set_ratio)
        if match and match[1] >= 70:
            idx = match[2]
            chosen = candidates[idx]
            row["item_id"] = chosen["id"]
            row["item_name"] = chosen["name"]
            row["item_match_score"] = match[1]
        else:
            row["item_id"] = None
            row["item_name"] = None
    return base


# ======================== Login Attestation (security audit) ========================
# Stores an admin-only audit record per login containing the user's location
# and a webcam photo. Capture is consent-based: the frontend shows a clear
# notice and standard browser permission prompts. When the user declines,
# we still log the event with `consent=False` so admins see WHO declined.

MAX_PHOTO_BYTES = 600 * 1024  # cap stored base64 size at ~600 KB to avoid bloating the DB


def _client_ip(request: Request) -> str:
    fwd = request.headers.get("x-forwarded-for") or ""
    if fwd:
        return fwd.split(",")[0].strip()
    return request.client.host if request.client else ""


_MOBILE_UA_RE = re.compile(
    r"Android|webOS|iPhone|iPad|iPod|BlackBerry|IEMobile|Opera Mini|Mobi|Tablet",
    re.IGNORECASE,
)


def _is_mobile_ua(ua: str) -> bool:
    """Best-effort mobile/tablet detection from the User-Agent header."""
    if not ua:
        return False
    return bool(_MOBILE_UA_RE.search(ua))


@api_router.post("/auth/attestation")
async def create_login_attestation(
    body: LoginAttestationIn,
    request: Request,
    user=Depends(get_current_user),
):
    """Record a consent-based login security capture (photo + location)."""
    photo_b64 = body.photo_b64 or ""
    # Strip "data:image/jpeg;base64," prefix if present
    if photo_b64.startswith("data:") and "," in photo_b64:
        photo_b64 = photo_b64.split(",", 1)[1]
    if len(photo_b64) > MAX_PHOTO_BYTES * 4 // 3:  # base64 is ~33% larger than raw
        # Reject oversized payloads rather than silently truncating
        raise HTTPException(status_code=413, detail="Photo too large (max ~600 KB)")

    doc = {
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        "user_email": user["email"],
        "username": user.get("username") or user["email"],
        "user_name": user.get("name") or user.get("username") or user["email"],
        "role": user.get("role") or "user",
        "ip": _client_ip(request),
        "user_agent": request.headers.get("user-agent") or "",
        "is_mobile": _is_mobile_ua(request.headers.get("user-agent") or ""),
        "consent": bool(body.consent),
        "latitude": body.latitude,
        "longitude": body.longitude,
        "accuracy_meters": body.accuracy_meters,
        "has_photo": bool(photo_b64),
        "photo_b64": photo_b64 or None,
        "photo_skipped": bool(body.photo_skipped),
        "location_skipped": bool(body.location_skipped),
        "error": (body.error or "")[:500],
        "created_at": now_iso(),
    }
    await db.login_attestations.insert_one(doc)
    return {"id": doc["id"], "stored": True, "has_photo": doc["has_photo"]}


@api_router.get("/admin/login-attestations")
async def list_login_attestations(
    limit: int = 50,
    skip: int = 0,
    user_id: Optional[str] = None,
    consent: Optional[bool] = None,
    admin=Depends(require_admin),
):
    """Admin-only audit log of all login captures (most recent first)."""
    q: Dict[str, Any] = {}
    if user_id:
        q["user_id"] = user_id
    if consent is not None:
        q["consent"] = consent
    cursor = (
        db.login_attestations.find(q, {"_id": 0, "photo_b64": 0})
        .sort("created_at", -1)
        .skip(max(0, int(skip)))
        .limit(max(1, min(200, int(limit))))
    )
    rows = await cursor.to_list(length=200)
    total = await db.login_attestations.count_documents(q)
    return {"total": total, "items": rows}


@api_router.get("/admin/login-attestations/{att_id}/photo")
async def get_login_attestation_photo(att_id: str, admin=Depends(require_admin)):
    """Admin-only: stream the JPEG photo for a single attestation record."""
    rec = await db.login_attestations.find_one({"id": att_id}, {"_id": 0, "photo_b64": 1})
    if not rec or not rec.get("photo_b64"):
        raise HTTPException(status_code=404, detail="Photo not available")
    try:
        raw = base64.b64decode(rec["photo_b64"])
    except Exception:
        raise HTTPException(status_code=500, detail="Stored photo is not valid base64")
    return Response(content=raw, media_type="image/jpeg")


async def _persist_customer_transport(customer_id: Optional[str], transport_name: Optional[str]) -> None:
    """Persist a freshly-typed transport name back onto the customer record so
    it becomes the default for that party next time. No-op for walk-ins
    (no customer_id) and when the value is empty or unchanged.
    """
    if not customer_id:
        return
    tn = (transport_name or "").strip()
    if not tn:
        return
    cur = await db.customers.find_one(
        {"id": customer_id}, {"_id": 0, "transport_name": 1}
    )
    if not cur:
        return
    if (cur.get("transport_name") or "").strip() == tn:
        return
    await db.customers.update_one(
        {"id": customer_id},
        {"$set": {"transport_name": tn, "updated_at": now_iso()}},
    )


# ======================== Off-Order (Direct) Dispatch ========================
async def _restore_dispatch_qty_to_orders(
    dispatch: Dict[str, Any],
    restore_items: List[Dict[str, Any]],
) -> List[str]:
    """Add the given (item_id, quantity) pairs back into the dispatch's
    parent order(s). Used by DELETE /dispatches/{did} and by PATCH when the
    edited slip dispatches LESS of an item than before.

    Resolution strategy for picking which order to credit:
      1. `dispatch.order_id` if it still exists in the DB.
      2. Any id in `dispatch.order_ids` that still exists, oldest first.
      3. Otherwise the customer's most-recent existing order (Pending or
         Dispatched) that already has the SKU in `items` or `original_items`.
      4. If still nothing, the SKU is silently dropped from the restore
         (true walk-in with no parent order to credit).

    When an order's items list goes from empty → non-empty, its status is
    flipped back from "Dispatched" to "Pending". Returns the list of order
    ids touched.
    """
    if not restore_items:
        return []

    # Build the candidate order list (preserving order_ids order).
    candidate_ids: List[str] = []
    if dispatch.get("order_id"):
        candidate_ids.append(dispatch["order_id"])
    for oid in (dispatch.get("order_ids") or []):
        if oid and oid not in candidate_ids:
            candidate_ids.append(oid)
    candidate_orders: List[Dict[str, Any]] = []
    if candidate_ids:
        found = await db.orders.find({"id": {"$in": candidate_ids}}, {"_id": 0}).to_list(2000)
        # Preserve the candidate_ids order
        by_id = {o["id"]: o for o in found}
        for oid in candidate_ids:
            if oid in by_id:
                candidate_orders.append(by_id[oid])

    customer_id = dispatch.get("customer_id")
    touched: Dict[str, Dict[str, Any]] = {}  # id -> live order doc (mutated)

    for r in restore_items:
        iid = r.get("item_id")
        qty = int(r.get("quantity") or 0)
        if not iid or qty <= 0:
            continue

        # Pick a target order: first the dispatch's tracked candidates, then
        # any other order of this customer that already knows this SKU.
        target = None
        for o in candidate_orders:
            target = touched.get(o["id"], o)
            break
        if target is None and customer_id:
            other = await db.orders.find(
                {"customer_id": customer_id}, {"_id": 0}
            ).sort("created_at", -1).to_list(2000)
            for o in other:
                items_list = o.get("items") or []
                orig_list = o.get("original_items") or []
                if any(i.get("item_id") == iid for i in items_list) or any(
                    i.get("item_id") == iid for i in orig_list
                ):
                    target = touched.get(o["id"], o)
                    break
        if target is None:
            continue  # nothing sensible to restore into

        items_list = list(target.get("items") or [])
        # If the line is still there, just bump its qty.
        bumped = False
        for it in items_list:
            if it.get("item_id") == iid:
                it["quantity"] = int(it.get("quantity") or 0) + qty
                bumped = True
                break
        if not bumped:
            # Line not present — either (a) fully dispatched and removed, or
            # (b) auto-cleared by the 85% rule. The correct pending qty to
            # restore is NOT just the restore delta `qty`, but the residual
            # implied by the order's original_items snapshot minus whatever
            # is currently dispatched across all live dispatches for this
            # (order, item). This makes the auto-clear rule reversible:
            # reducing a dispatch below the 85% threshold reopens the full
            # remaining residual, not just the amount that was removed.
            template = None
            original_qty = 0
            for oi in (target.get("original_items") or []):
                if oi.get("item_id") == iid:
                    template = {**oi}
                    original_qty = int(oi.get("quantity") or 0)
                    break

            new_qty_to_add = qty  # fall-back when there is no snapshot
            if original_qty > 0:
                live_disp = await db.dispatches.find(
                    {
                        "$or": [
                            {"order_id": target["id"]},
                            {"order_ids": target["id"]},
                        ]
                    },
                    {"_id": 0, "items": 1},
                ).to_list(2000)
                total_dispatched = 0
                for ld in live_disp:
                    for li in (ld.get("items") or []):
                        if li.get("item_id") == iid:
                            total_dispatched += int(li.get("quantity") or 0)
                residual = original_qty - total_dispatched
                if residual <= 0:
                    # Still fully dispatched after the edit — nothing to add.
                    continue
                # Reapply the 85% auto-clear rule with the new totals: if the
                # SKU is still at/above the threshold, leave it cleared.
                if (original_qty - residual) / original_qty >= LINE_CLEAR_THRESHOLD:
                    continue
                new_qty_to_add = residual

            if template is None:
                template = {
                    "item_id": iid,
                    "item_name": r.get("item_name"),
                    "product_name": r.get("product_name"),
                    "variant": r.get("variant"),
                }
            template["quantity"] = new_qty_to_add
            items_list.append(template)

        target["items"] = items_list
        # If the order was Dispatched and now has items again, reopen it.
        if items_list and target.get("status") != "Pending":
            target["status"] = "Pending"
        touched[target["id"]] = target

    # Persist each touched order.
    for oid, o in touched.items():
        await db.orders.update_one(
            {"id": oid},
            {"$set": {
                "items": o.get("items") or [],
                "status": o.get("status") or "Pending",
                "updated_at": now_iso(),
            }},
        )
    return list(touched.keys())


async def _revert_bom_consumption_for_dispatch(
    dispatch_id: str,
    actor: str,
) -> Dict[str, float]:
    """For every `raw_material_movements` row written by this dispatch
    (`kind='dispatch'`, `reference_id=dispatch_id`), add the consumed qty
    back to the raw material's stock and log a paired `dispatch_revert`
    movement. Returns {rm_id: qty_restored}.
    """
    rows = await db.raw_material_movements.find(
        {"reference_id": dispatch_id, "kind": "dispatch"}, {"_id": 0}
    ).to_list(2000)
    restored: Dict[str, float] = {}
    for row in rows:
        rid = row.get("raw_material_id")
        delta = float(row.get("delta") or 0)
        # delta on dispatch is negative; we want to add |delta| back.
        give_back = -delta
        if not rid or give_back <= 0:
            continue
        await _apply_rm_movement(
            raw_material_id=rid,
            delta=give_back,
            kind="dispatch_revert",
            reference_id=dispatch_id,
            actor=actor or "",
            notes=f"Reversed by dispatch {dispatch_id} delete/edit",
        )
        restored[rid] = restored.get(rid, 0.0) + give_back
    return restored


async def _deduct_off_order_from_pending_orders(
    customer_id: Optional[str],
    dispatched_lines: List[Dict[str, Any]],
) -> List[str]:
    """If the customer has pending orders that include any of the dispatched
    SKUs, deduct the dispatched quantities from those orders FIFO (oldest
    Pending order first). Mirrors the consumption logic used by the regular
    `/dispatch/execute` flow so an Off-Order Dispatch also shrinks the
    matching parent order(s).

    - Walk-in dispatches (no customer_id) are no-ops.
    - Quantity dispatched in excess of total pending qty for that SKU is
      kept as off-order (no order is touched for the excess) — Off-Order
      Dispatch must never fail because of a missing/short pending order.
    - When an order's items list becomes empty, the order is marked
      Dispatched. Otherwise it stays Pending with remaining items.

    Returns the list of order ids that were modified (for traceability on
    the resulting dispatch document).
    """
    if not customer_id or not dispatched_lines:
        return []

    # FIFO across the customer's pending orders (oldest first).
    pending_orders = await db.orders.find(
        {"customer_id": customer_id, "status": "Pending"}, {"_id": 0}
    ).sort("created_at", 1).to_list(2000)
    if not pending_orders:
        return []

    # Remaining qty to apply for each dispatched item.
    remaining_by_iid: Dict[str, int] = {
        d["item_id"]: int(d["quantity"]) for d in dispatched_lines
    }

    touched: List[str] = []
    for order in pending_orders:
        if all(v <= 0 for v in remaining_by_iid.values()):
            break
        items = order.get("items") or []
        # Plan how much we will deduct from each line of this order.
        per_line_take: Dict[str, int] = {}
        any_change = False
        for it in items:
            iid = it.get("item_id")
            take = remaining_by_iid.get(iid, 0)
            cur_qty = int(it.get("quantity") or 0)
            if take > 0 and cur_qty > 0:
                applied = min(take, cur_qty)
                per_line_take[iid] = applied
                remaining_by_iid[iid] = take - applied
                any_change = True
        if not any_change:
            continue

        original_items_for_rule = order.get("original_items") or items
        new_items = _apply_line_clear_threshold(items, per_line_take, original_items_for_rule)

        update_doc: Dict[str, Any] = {
            "items": new_items,
            "updated_at": now_iso(),
        }
        if not new_items:
            update_doc["status"] = "Dispatched"
        # Preserve original item list on first dispatch — snapshot on any
        # dispatch (Pending or Dispatched outcome) so the 85% auto-clear
        # rule has a stable baseline for later edit/delete unwind.
        if "original_items" not in order:
            update_doc["original_items"] = items

        await db.orders.update_one({"id": order["id"]}, {"$set": update_doc})
        touched.append(order["id"])

    return touched


@api_router.post("/dispatch/off-order")
async def dispatch_off_order(body: OffOrderDispatchIn, user=Depends(get_current_user)):
    """Dispatch SKUs to a party — used either for walk-ins (no parent
    order) OR for an existing customer where the operator wants to short-
    circuit the regular allocation flow. Now available to BOTH admins and
    operators (matches the order-linked /dispatch/execute permission so
    the factory-floor user can record same-day off-order slips).

    The created record lives in the `dispatches` collection alongside
    order-linked dispatches, so it is included in the Daily Dispatch
    Report automatically.

    Behaviour update (Jan 2026): if the customer has matching SKUs on any
    Pending order(s), the dispatched quantity is also deducted from those
    orders FIFO. Excess dispatched quantity (beyond what is pending) is
    kept as off-order. Walk-in customers (no `customer_id`) are unchanged.
    """
    # ---- Resolve customer (existing OR walk-in) ----
    cust: Dict[str, Any] = {}
    customer_id = body.customer_id or None
    customer_name = (body.customer_name or "").strip()
    if customer_id:
        cust = await db.customers.find_one({"id": customer_id}, {"_id": 0}) or {}
        if not cust:
            raise HTTPException(status_code=404, detail="Customer not found")
        customer_name = cust.get("name") or customer_name
    if not customer_name:
        raise HTTPException(status_code=400, detail="customer_id or customer_name is required")

    # ---- Validate items ----
    if not body.items:
        raise HTTPException(status_code=400, detail="At least one item is required")
    # Per-party block-list enforcement: reject the whole dispatch if any
    # SKU in the payload is blocked for this party. The item search
    # dropdown already hides blocked SKUs; this is the backend safety net.
    if customer_id:
        blocked_ids = set((cust.get("blocked_items") or []))
        if blocked_ids:
            offending_ids = [li.item_id for li in body.items if li.item_id in blocked_ids]
            if offending_ids:
                off_docs = await db.items.find(
                    {"id": {"$in": offending_ids}}, {"_id": 0, "name": 1, "id": 1}
                ).to_list(1000)
                names = [d.get("name") or d.get("id") for d in off_docs]
                raise HTTPException(
                    status_code=400,
                    detail=f"These items are blocked for {customer_name}: {', '.join(names)}",
                )
    # Task 1 — persist operator-chosen price list on the customer BEFORE
    # pricing so the slip is priced against the freshly-chosen list and
    # next dispatch auto-suggests it. Only acts when the operator sent a
    # non-None value (sends "" to clear, sends an id to assign).
    if customer_id and body.price_list_id is not None:
        new_pl = body.price_list_id or None
        if new_pl and new_pl != cust.get("price_list_id"):
            pl_exists = await db.price_lists.count_documents({"id": new_pl})
            if pl_exists == 0:
                raise HTTPException(status_code=404, detail="Selected price list not found")
        await db.customers.update_one(
            {"id": customer_id},
            {"$set": {"price_list_id": new_pl}},
        )
        cust["price_list_id"] = new_pl
    seen_ids: set = set()
    dispatched_lines: List[Dict[str, Any]] = []
    cust_price_list = cust.get("price_list_id")
    for line in body.items:
        if line.quantity is None or int(line.quantity) <= 0:
            raise HTTPException(status_code=400, detail="Each item quantity must be > 0")
        if line.item_id in seen_ids:
            raise HTTPException(status_code=400, detail=f"Duplicate item in payload: {line.item_id}")
        seen_ids.add(line.item_id)
        sku = await db.items.find_one({"id": line.item_id}, {"_id": 0})
        if not sku:
            raise HTTPException(status_code=404, detail=f"Item {line.item_id} not found")
        pricing = await compute_line_pricing(
            cust_price_list, line.item_id, sku.get("product_name") or ""
        )
        dispatched_lines.append({
            "item_id": line.item_id,
            "item_name": sku.get("name"),
            "product_name": sku.get("product_name"),
            "variant": sku.get("variant"),
            "quantity": int(line.quantity),
            "description": (line.description or "").strip(),
            **pricing,
        })

    # NOTE: total_value (Bill Amount) is NEVER auto-computed from item
    # pricing — operator must fill it manually in Daily Report. Item pricing
    # remains on each line for printable-slip reference only.

    # Persist a freshly-typed transport name back onto the customer record
    # (Jan 2026) so the "Defaults to party's transport" placeholder uses
    # the most recent value next time.
    await _persist_customer_transport(customer_id, body.transport_name)
    if customer_id and body.transport_name:
        # Refresh local cust dict so the dispatch_doc below uses the new value.
        cust = await db.customers.find_one({"id": customer_id}, {"_id": 0}) or cust

    # Deduct dispatched quantities from the customer's matching Pending
    # orders FIFO (Jan 2026). Walk-ins and excess qty are unaffected.
    touched_order_ids = await _deduct_off_order_from_pending_orders(customer_id, dispatched_lines)
    any_order_fully_dispatched = False
    if touched_order_ids:
        cleared = await db.orders.count_documents(
            {"id": {"$in": touched_order_ids}, "status": "Dispatched"}
        )
        any_order_fully_dispatched = cleared > 0

    # MERGE into today's existing slip for this customer if one exists, so
    # multiple same-day dispatches to one party produce ONE slip.
    dispatch_ts = _resolve_dispatch_ts(body.dispatched_at)
    existing = await _find_open_dispatch_today(customer_id, customer_name, for_iso_ts=dispatch_ts)
    if existing:
        merged_items = _merge_dispatch_lines(existing.get("items") or [], dispatched_lines)
        merged_total_pcs = sum(int(it.get("quantity") or 0) for it in merged_items)
        merged_notes = (existing.get("notes") or "").strip()
        if body.notes:
            extra = body.notes.strip()
            merged_notes = (merged_notes + " | " + extra).strip(" |") if merged_notes else extra
        # Merge order_ids so traceability is preserved when an off-order
        # dispatch lands on top of an earlier slip.
        merged_order_ids = list(existing.get("order_ids") or [])
        if existing.get("order_id") and existing["order_id"] not in merged_order_ids:
            merged_order_ids.append(existing["order_id"])
        for oid in touched_order_ids:
            if oid not in merged_order_ids:
                merged_order_ids.append(oid)
        update_set: Dict[str, Any] = {
            "items": merged_items,
            "total_pcs": merged_total_pcs,
            # Task 2 — Bill Amount is NEVER auto-recomputed. It stays at
            # whatever value the operator manually entered. Adding new
            # items to a same-day slip MUST NOT silently overwrite that.
            "notes": merged_notes,
            "order_ids": merged_order_ids,
            "last_dispatched_at": now_iso(),
            "last_dispatched_by": user["email"],
        }
        if any_order_fully_dispatched:
            update_set["order_fully_dispatched"] = True
        # Persist operator-entered bag count if provided on this call. We
        # OVERWRITE (rather than sum) because the field represents the
        # total bag count for the merged slip and the operator sees the
        # merge context in the UI.
        if body.bag_count is not None:
            try:
                bc = int(body.bag_count)
                if bc < 0:
                    bc = 0
                update_set["bag_count"] = bc
            except Exception:
                pass
        await db.dispatches.update_one({"id": existing["id"]}, {"$set": update_set})
        dispatch_doc = await db.dispatches.find_one({"id": existing["id"]}, {"_id": 0})
        # Auto-consume raw materials only for the NEW lines (merging into
        # an existing slip must not double-deduct previously dispatched items).
        await _consume_bom_for_lines(dispatched_lines, dispatch_doc["id"], user.get("email") or "")
        return {
            "dispatch": dispatch_doc,
            "merged": True,
            "order_ids_touched": touched_order_ids,
        }

    dispatch_doc = {
        "id": str(uuid.uuid4()),
        "slip_no": await next_slip_no(),
        # If the off-order dispatch consumed exactly one pending order we
        # set order_id for backward-compatible single-order traceability;
        # the full list is always available in order_ids.
        "order_id": touched_order_ids[0] if len(touched_order_ids) == 1 else None,
        "order_ids": touched_order_ids,
        "off_order": True,  # convenience flag for reporting/filtering
        "customer_id": customer_id,
        "customer_name": customer_name,
        "transport_name": (body.transport_name or cust.get("transport_name") or "").strip(),
        "price_list_id": cust_price_list,
        "items": dispatched_lines,
        "total_pcs": sum(d["quantity"] for d in dispatched_lines),
        # Task 2 — Bill Amount starts EMPTY (0). The operator must
        # enter it manually from Daily Report / Dispatch Ledger.
        "total_value": 0.0,
        "notes": (body.notes or "").strip(),
        "bill_number": (body.bill_number or "").strip(),
        "bag_count": max(0, int(body.bag_count)) if (body.bag_count is not None) else 0,
        "dispatched_by": user["email"],
        "dispatched_at": dispatch_ts,
        "order_fully_dispatched": any_order_fully_dispatched,
    }
    await db.dispatches.insert_one(dispatch_doc)
    dispatch_doc.pop("_id", None)
    # Auto-consume raw materials for this off-order dispatch
    await _consume_bom_for_lines(dispatched_lines, dispatch_doc["id"], user.get("email") or "")
    return {"dispatch": dispatch_doc, "order_ids_touched": touched_order_ids}


# ======================== Party Payments ========================
# Payments are the credit side of a party's ledger — money received from a
# customer (cash, UPI, NEFT/RTGS, cheque, etc.). Combined with dispatches
# (the debit side) they drive the running account balance shown in the
# Single Party Ledger page.

PAYMENT_SOURCES = {"cash", "upi", "bank_transfer", "neft", "rtgs", "cheque", "card", "adjustment", "other"}


class PaymentIn(BaseModel):
    customer_id: str
    amount: float
    source: str = "cash"
    reference: Optional[str] = ""  # UTR, cheque #, transaction id, etc.
    paid_at: Optional[str] = None  # ISO date or datetime; defaults to "now"
    notes: Optional[str] = ""
    # "cash" = received from customer (default).
    # "supplier_on_behalf" = we paid a 3rd-party supplier on the customer's
    # behalf; that amount credits the customer's ledger AND debits the
    # supplier's ledger (via a linked supplier_payment record).
    payment_mode: Optional[str] = "cash"
    paid_to_supplier_id: Optional[str] = None  # required when payment_mode = supplier_on_behalf


class PaymentUpdate(BaseModel):
    amount: Optional[float] = None
    source: Optional[str] = None
    reference: Optional[str] = None
    paid_at: Optional[str] = None
    notes: Optional[str] = None


def _normalize_payment_source(src: str) -> str:
    s = (src or "").strip().lower().replace("-", "_").replace(" ", "_")
    return s if s in PAYMENT_SOURCES else "other"


def _normalize_payment_dt(value: Optional[str]) -> str:
    """Accept '2026-06-12', '2026-06-12T15:30:00', or ISO with TZ. Always
    persist as ISO 8601 with seconds + timezone."""
    if not value:
        return datetime.now(timezone.utc).isoformat()
    try:
        s = value.strip()
        if len(s) == 10:  # bare YYYY-MM-DD
            return datetime.fromisoformat(s + "T00:00:00").replace(tzinfo=timezone.utc).isoformat()
        # full datetime — let fromisoformat parse
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.isoformat()
    except Exception:
        return datetime.now(timezone.utc).isoformat()


@api_router.post("/payments")
async def create_payment(body: PaymentIn, user=Depends(get_current_user)):
    if not body.customer_id:
        raise HTTPException(status_code=400, detail="customer_id required")
    if body.amount is None or float(body.amount) <= 0:
        raise HTTPException(status_code=400, detail="Amount must be greater than zero")
    cust = await db.customers.find_one({"id": body.customer_id}, {"_id": 0})
    if not cust:
        raise HTTPException(status_code=404, detail="Customer not found")

    mode = (body.payment_mode or "cash").strip().lower()
    supplier = None
    if mode == "supplier_on_behalf":
        if not body.paid_to_supplier_id:
            raise HTTPException(status_code=400, detail="paid_to_supplier_id required when payment_mode is supplier_on_behalf")
        supplier = await db.suppliers.find_one({"id": body.paid_to_supplier_id}, {"_id": 0})
        if not supplier:
            raise HTTPException(status_code=404, detail="Supplier not found")
    else:
        mode = "cash"

    pid = str(uuid.uuid4())
    payment = {
        "id": pid,
        "receipt_no": await next_receipt_no(),
        "customer_id": body.customer_id,
        "customer_name": cust.get("name") or "",
        "amount": round(float(body.amount), 2),
        "source": _normalize_payment_source(body.source),
        "reference": (body.reference or "").strip(),
        "paid_at": _normalize_payment_dt(body.paid_at),
        "notes": (body.notes or "").strip(),
        "payment_mode": mode,
        "paid_to_supplier_id": (body.paid_to_supplier_id or None) if mode == "supplier_on_behalf" else None,
        "paid_to_supplier_name": (supplier.get("name") if supplier else None),
        "created_by": user.get("email") or user.get("username") or "",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.payments.insert_one(payment)
    payment.pop("_id", None)

    # If paid on behalf of customer to a supplier, mirror as a supplier_payment.
    if mode == "supplier_on_behalf" and supplier:
        sup_payment = {
            "id": str(uuid.uuid4()),
            "supplier_id": supplier["id"],
            "supplier_name": supplier.get("name") or "",
            "amount": payment["amount"],
            "source": payment["source"],
            "reference": payment["reference"],
            "paid_at": payment["paid_at"],
            "notes": payment["notes"],
            "on_behalf_of_customer_id": body.customer_id,
            "on_behalf_of_customer_name": cust.get("name") or "",
            "customer_payment_id": pid,
            "created_by": payment["created_by"],
            "created_at": payment["created_at"],
        }
        await db.supplier_payments.insert_one(sup_payment)

    return payment


@api_router.get("/payments")
async def list_payments(
    customer_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    limit: int = 500,
    skip: int = 0,
    user=Depends(get_current_user),
):
    q: Dict[str, Any] = {}
    if customer_id:
        q["customer_id"] = customer_id
    if start_date or end_date:
        rng: Dict[str, Any] = {}
        if start_date:
            rng["$gte"] = start_date
        if end_date:
            rng["$lte"] = end_date + "T23:59:59"
        q["paid_at"] = rng
    cursor = db.payments.find(q, {"_id": 0}).sort("paid_at", -1).skip(skip).limit(limit)
    items = await cursor.to_list(length=limit)
    total = await db.payments.count_documents(q)
    total_amount = round(sum(float(p.get("amount") or 0) for p in items), 2)
    return {
        "items": items,
        "total": total,
        "total_amount": total_amount,
    }


@api_router.patch("/payments/{pid}")
async def update_payment(pid: str, body: PaymentUpdate, user=Depends(require_action("edit:customerLedger"))):
    existing = await db.payments.find_one({"id": pid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Payment not found")
    upd: Dict[str, Any] = {}
    if body.amount is not None:
        if float(body.amount) <= 0:
            raise HTTPException(status_code=400, detail="Amount must be greater than zero")
        upd["amount"] = round(float(body.amount), 2)
    if body.source is not None:
        upd["source"] = _normalize_payment_source(body.source)
    if body.reference is not None:
        upd["reference"] = body.reference.strip()
    if body.paid_at is not None:
        upd["paid_at"] = _normalize_payment_dt(body.paid_at)
    if body.notes is not None:
        upd["notes"] = body.notes.strip()
    if upd:
        await db.payments.update_one({"id": pid}, {"$set": upd})
    return await db.payments.find_one({"id": pid}, {"_id": 0})


@api_router.delete("/payments/{pid}")
async def delete_payment(pid: str, admin=Depends(require_action("delete:customerLedger"))):
    # Cascade-remove the mirrored supplier_payment (if this was an on-behalf
    # payment) so the supplier ledger stays in sync.
    await db.supplier_payments.delete_many({"customer_payment_id": pid})
    res = await db.payments.delete_one({"id": pid})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Payment not found")
    return {"ok": True}


# ======================== Suppliers (Admin) ========================
# Suppliers are vendors that provide raw materials. The supplier ledger
# mirrors the customer ledger: purchases (we owe them) are debits, payments
# we make (cash or on behalf of a customer) are credits.

class RawMaterialIn(BaseModel):
    name: str
    unit: Optional[str] = "pcs"          # kg / pcs / litre / m / etc.
    default_rate: float = 0.0            # informational default for purchases
    notes: Optional[str] = ""


class RawMaterialUpdate(BaseModel):
    name: Optional[str] = None
    unit: Optional[str] = None
    default_rate: Optional[float] = None
    notes: Optional[str] = None


@api_router.get("/raw-materials")
async def list_raw_materials(user=Depends(get_current_user)):
    items = await db.raw_materials.find({}, {"_id": 0}).sort("name", 1).to_list(5000)
    return items


@api_router.post("/raw-materials")
async def create_raw_material(body: RawMaterialIn, admin=Depends(require_admin)):
    name = (body.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="name required")
    doc = {
        "id": str(uuid.uuid4()),
        "name": name,
        "unit": (body.unit or "pcs").strip(),
        "default_rate": round(float(body.default_rate or 0), 2),
        "notes": (body.notes or "").strip(),
        "stock_on_hand": 0.0,
        "created_at": now_iso(),
        "created_by": admin["email"],
    }
    await db.raw_materials.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.patch("/raw-materials/{rid}")
async def update_raw_material(rid: str, body: RawMaterialUpdate, admin=Depends(require_action("edit:rawMaterials"))):
    upd: Dict[str, Any] = {"updated_at": now_iso(), "updated_by": admin["email"]}
    if body.name is not None:
        upd["name"] = body.name.strip()
    if body.unit is not None:
        upd["unit"] = body.unit.strip()
    if body.default_rate is not None:
        upd["default_rate"] = round(float(body.default_rate), 2)
    if body.notes is not None:
        upd["notes"] = body.notes.strip()
    res = await db.raw_materials.update_one({"id": rid}, {"$set": upd})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Raw material not found")
    return await db.raw_materials.find_one({"id": rid}, {"_id": 0})


@api_router.delete("/raw-materials/{rid}")
async def delete_raw_material(rid: str, admin=Depends(require_action("delete:rawMaterials"))):
    res = await db.raw_materials.delete_one({"id": rid})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Raw material not found")
    return {"ok": True}


# ======================== Vendor Price Lists ========================
# Each vendor (supplier) can have one or many price lists. A list contains
# free-form line items (item name + unit + price). Items are NOT tied to the
# SKU master because vendors sell things we don't necessarily resell.
class VendorPriceListIn(BaseModel):
    name: str
    vendor_id: str  # supplier_id
    description: Optional[str] = ""


class VendorPriceListUpdate(BaseModel):
    name: Optional[str] = None
    vendor_id: Optional[str] = None
    description: Optional[str] = None


class VendorPriceListItemIn(BaseModel):
    name: str
    unit: Optional[str] = ""
    price: float = 0.0
    notes: Optional[str] = ""


class VendorPriceListItemUpdate(BaseModel):
    name: Optional[str] = None
    unit: Optional[str] = None
    price: Optional[float] = None
    notes: Optional[str] = None


@api_router.get("/vendor-price-lists")
async def list_vendor_price_lists(user=Depends(get_current_user)):
    lists = await db.vendor_price_lists.find({}, {"_id": 0}).sort("name", 1).to_list(1000)
    # Attach vendor_name and items_count for the index view
    vendor_ids = list({pl.get("vendor_id") for pl in lists if pl.get("vendor_id")})
    suppliers = {}
    if vendor_ids:
        async for s in db.suppliers.find({"id": {"$in": vendor_ids}}, {"_id": 0, "id": 1, "name": 1}):
            suppliers[s["id"]] = s["name"]
    out = []
    for pl in lists:
        cnt = await db.vendor_price_list_items.count_documents({"vendor_price_list_id": pl["id"]})
        out.append({**pl, "vendor_name": suppliers.get(pl.get("vendor_id"), ""), "items_count": cnt})
    return out


@api_router.post("/vendor-price-lists")
async def create_vendor_price_list(body: VendorPriceListIn, admin=Depends(require_admin)):
    name = (body.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Name required")
    sup = await db.suppliers.find_one({"id": body.vendor_id}, {"_id": 0, "id": 1, "name": 1})
    if not sup:
        raise HTTPException(status_code=404, detail="Vendor not found")
    doc = {
        "id": str(uuid.uuid4()),
        "name": name,
        "vendor_id": body.vendor_id,
        "vendor_name": sup["name"],
        "description": (body.description or "").strip(),
        "created_at": now_iso(),
        "created_by": admin["email"],
    }
    await db.vendor_price_lists.insert_one(doc)
    doc.pop("_id", None)
    return {**doc, "items_count": 0}


@api_router.get("/vendor-price-lists/{vpl_id}")
async def get_vendor_price_list(vpl_id: str, user=Depends(get_current_user)):
    pl = await db.vendor_price_lists.find_one({"id": vpl_id}, {"_id": 0})
    if not pl:
        raise HTTPException(status_code=404, detail="Vendor price list not found")
    items = await db.vendor_price_list_items.find(
        {"vendor_price_list_id": vpl_id}, {"_id": 0}
    ).sort("name", 1).to_list(5000)
    return {**pl, "items": items}


@api_router.patch("/vendor-price-lists/{vpl_id}")
async def update_vendor_price_list(vpl_id: str, body: VendorPriceListUpdate, admin=Depends(require_action("edit:vendorPriceLists"))):
    upd: Dict[str, Any] = {"updated_at": now_iso(), "updated_by": admin["email"]}
    if body.name is not None:
        n = body.name.strip()
        if not n:
            raise HTTPException(status_code=400, detail="Name cannot be empty")
        upd["name"] = n
    if body.description is not None:
        upd["description"] = body.description.strip()
    if body.vendor_id is not None:
        sup = await db.suppliers.find_one({"id": body.vendor_id}, {"_id": 0, "id": 1, "name": 1})
        if not sup:
            raise HTTPException(status_code=404, detail="Vendor not found")
        upd["vendor_id"] = body.vendor_id
        upd["vendor_name"] = sup["name"]
    res = await db.vendor_price_lists.update_one({"id": vpl_id}, {"$set": upd})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Vendor price list not found")
    return await db.vendor_price_lists.find_one({"id": vpl_id}, {"_id": 0})


@api_router.delete("/vendor-price-lists/{vpl_id}")
async def delete_vendor_price_list(vpl_id: str, admin=Depends(require_action("delete:vendorPriceLists"))):
    res = await db.vendor_price_lists.delete_one({"id": vpl_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Vendor price list not found")
    # Cascade delete its items
    await db.vendor_price_list_items.delete_many({"vendor_price_list_id": vpl_id})
    return {"ok": True}


@api_router.post("/vendor-price-lists/{vpl_id}/items")
async def add_vendor_price_list_item(vpl_id: str, body: VendorPriceListItemIn, admin=Depends(require_admin)):
    pl = await db.vendor_price_lists.find_one({"id": vpl_id}, {"_id": 0, "id": 1})
    if not pl:
        raise HTTPException(status_code=404, detail="Vendor price list not found")
    name = (body.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Item name required")
    unit = (body.unit or "").strip()
    price = round(float(body.price or 0), 2)
    # Auto-create / reuse raw material with the same name (case-insensitive)
    rm_id = await _ensure_raw_material(name, unit, price, admin)
    doc = {
        "id": str(uuid.uuid4()),
        "vendor_price_list_id": vpl_id,
        "name": name,
        "unit": unit,
        "price": price,
        "notes": (body.notes or "").strip(),
        "raw_material_id": rm_id,
        "created_at": now_iso(),
    }
    await db.vendor_price_list_items.insert_one(doc)
    doc.pop("_id", None)
    return doc


async def _ensure_raw_material(name: str, unit: str, price: float, admin: Dict[str, Any]) -> str:
    """Find a raw material by case-insensitive name; create one if missing.
    Returns the raw material id. Existing raw materials are NOT modified."""
    name = (name or "").strip()
    if not name:
        return ""
    # Case-insensitive exact match
    existing = await db.raw_materials.find_one(
        {"name": {"$regex": f"^{re.escape(name)}$", "$options": "i"}},
        {"_id": 0, "id": 1},
    )
    if existing:
        return existing["id"]
    new_id = str(uuid.uuid4())
    await db.raw_materials.insert_one({
        "id": new_id,
        "name": name,
        "unit": (unit or "pcs").strip() or "pcs",
        "default_rate": round(float(price or 0), 2),
        "notes": "Auto-added from vendor price list",
        "stock_on_hand": 0.0,
        "created_at": now_iso(),
        "created_by": admin.get("email") or admin.get("username") or "system",
    })
    return new_id


@api_router.patch("/vendor-price-lists/{vpl_id}/items/{vpi_id}")
async def update_vendor_price_list_item(
    vpl_id: str, vpi_id: str, body: VendorPriceListItemUpdate, admin=Depends(require_action("edit:vendorPriceLists"))
):
    upd: Dict[str, Any] = {"updated_at": now_iso()}
    if body.name is not None:
        n = body.name.strip()
        if not n:
            raise HTTPException(status_code=400, detail="Item name cannot be empty")
        upd["name"] = n
    if body.unit is not None:
        upd["unit"] = body.unit.strip()
    if body.price is not None:
        upd["price"] = round(float(body.price), 2)
    if body.notes is not None:
        upd["notes"] = body.notes.strip()
    # If name or unit/price changed, make sure a matching raw material exists
    if any(k in upd for k in ("name", "unit", "price")):
        cur = await db.vendor_price_list_items.find_one(
            {"id": vpi_id, "vendor_price_list_id": vpl_id}, {"_id": 0}
        ) or {}
        eff_name = upd.get("name", cur.get("name", ""))
        eff_unit = upd.get("unit", cur.get("unit", ""))
        eff_price = upd.get("price", cur.get("price", 0))
        if eff_name:
            upd["raw_material_id"] = await _ensure_raw_material(eff_name, eff_unit, eff_price, admin)
    res = await db.vendor_price_list_items.update_one(
        {"id": vpi_id, "vendor_price_list_id": vpl_id}, {"$set": upd}
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Item not found in this vendor price list")
    return await db.vendor_price_list_items.find_one({"id": vpi_id}, {"_id": 0})


@api_router.delete("/vendor-price-lists/{vpl_id}/items/{vpi_id}")
async def delete_vendor_price_list_item(vpl_id: str, vpi_id: str, admin=Depends(require_action("delete:vendorPriceLists"))):
    res = await db.vendor_price_list_items.delete_one(
        {"id": vpi_id, "vendor_price_list_id": vpl_id}
    )
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Item not found in this vendor price list")
    return {"ok": True}


@api_router.get("/vendor-price-lists/{vpl_id}/import/template")
async def vendor_price_list_import_template(vpl_id: str, admin=Depends(require_admin)):
    """Download a blank Excel template for bulk vendor-price-list item import."""
    from openpyxl import Workbook
    from fastapi.responses import StreamingResponse
    pl = await db.vendor_price_lists.find_one({"id": vpl_id}, {"_id": 0, "id": 1, "name": 1})
    if not pl:
        raise HTTPException(status_code=404, detail="Vendor price list not found")
    wb = Workbook()
    ws = wb.active
    ws.title = "Items"
    ws.append(["name", "unit", "price", "notes"])
    ws.append(["MS Sheet 1.2mm", "kg", 85.5, "example row — delete or overwrite"])
    widths = [40, 12, 14, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = "".join(c if c.isalnum() else "_" for c in pl["name"])
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="vendor_price_list_{safe_name}_template.xlsx"'},
    )


@api_router.get("/vendor-price-lists/{vpl_id}/export")
async def vendor_price_list_export(vpl_id: str, user=Depends(get_current_user)):
    """Download Excel of current vendor-price-list items (name | unit | price | notes)."""
    from openpyxl import Workbook
    from fastapi.responses import StreamingResponse
    pl = await db.vendor_price_lists.find_one({"id": vpl_id}, {"_id": 0})
    if not pl:
        raise HTTPException(status_code=404, detail="Vendor price list not found")
    items = await db.vendor_price_list_items.find(
        {"vendor_price_list_id": vpl_id}, {"_id": 0}
    ).sort("name", 1).to_list(5000)
    wb = Workbook()
    ws = wb.active
    ws.title = "Items"
    ws.append(["name", "unit", "price", "notes"])
    for it in items:
        ws.append([
            it.get("name", ""),
            it.get("unit", ""),
            float(it.get("price") or 0),
            it.get("notes", ""),
        ])
    widths = [40, 12, 14, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = "".join(c if c.isalnum() else "_" for c in pl["name"])
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="vendor_price_list_{safe_name}.xlsx"'},
    )


@api_router.post("/vendor-price-lists/{vpl_id}/import")
async def vendor_price_list_import(vpl_id: str, file: UploadFile = File(...), admin=Depends(require_admin)):
    """Bulk import items into a vendor price list from Excel.

    Columns recognised (case-insensitive header row): name, unit, price, notes.
    Behaviour:
    - Rows are upserted by name (case-insensitive) within this vendor price list:
      existing items with the same name get their unit / price / notes overwritten,
      new names are inserted.
    - Empty rows and rows without a name are skipped.
    - Invalid price values are reported as skipped rows.
    """
    from openpyxl import load_workbook
    pl = await db.vendor_price_lists.find_one({"id": vpl_id}, {"_id": 0, "id": 1})
    if not pl:
        raise HTTPException(status_code=404, detail="Vendor price list not found")
    blob = await file.read()
    if not blob:
        raise HTTPException(status_code=400, detail="Empty file")
    try:
        wb = load_workbook(io.BytesIO(blob), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel: {e}")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Empty sheet")
    header_row = rows[0]
    header_map: Dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        key = str(cell).strip().lower().replace(" ", "_")
        if key:
            header_map[key] = idx
    if "name" not in header_map:
        raise HTTPException(status_code=400, detail='Excel must have a "name" column in the first row')

    def get(r, key: str) -> Any:
        idx = header_map.get(key)
        if idx is None or idx >= len(r):
            return None
        return r[idx]

    existing_items = await db.vendor_price_list_items.find(
        {"vendor_price_list_id": vpl_id}, {"_id": 0, "id": 1, "name": 1}
    ).to_list(5000)
    by_name_lower = {(it.get("name") or "").strip().lower(): it["id"] for it in existing_items if it.get("name")}

    inserted = 0
    updated = 0
    skipped: List[Dict[str, Any]] = []
    seen_in_file: Dict[str, int] = {}

    for row_idx, r in enumerate(rows[1:], start=2):
        if not r or all(c is None or (isinstance(c, str) and not c.strip()) for c in r):
            continue
        name_raw = get(r, "name")
        name = str(name_raw).strip() if name_raw is not None else ""
        if not name:
            continue
        name_l = name.lower()
        if name_l in seen_in_file:
            skipped.append({"row": row_idx, "name": name, "reason": f"duplicate of row {seen_in_file[name_l]} in file"})
            continue
        seen_in_file[name_l] = row_idx

        unit_raw = get(r, "unit")
        unit = str(unit_raw).strip() if unit_raw is not None else ""
        notes_raw = get(r, "notes")
        notes = str(notes_raw).strip() if notes_raw is not None else ""
        price_raw = get(r, "price")
        try:
            price = float(price_raw) if price_raw is not None and str(price_raw).strip() != "" else 0.0
        except Exception:
            skipped.append({"row": row_idx, "name": name, "reason": f"invalid price '{price_raw}'"})
            continue

        existing_id = by_name_lower.get(name_l)
        rm_id = await _ensure_raw_material(name, unit, price, admin)
        if existing_id:
            await db.vendor_price_list_items.update_one(
                {"id": existing_id, "vendor_price_list_id": vpl_id},
                {"$set": {
                    "name": name,
                    "unit": unit,
                    "price": round(price, 2),
                    "notes": notes,
                    "raw_material_id": rm_id,
                    "updated_at": now_iso(),
                }},
            )
            updated += 1
        else:
            new_id = str(uuid.uuid4())
            await db.vendor_price_list_items.insert_one({
                "id": new_id,
                "vendor_price_list_id": vpl_id,
                "name": name,
                "unit": unit,
                "price": round(price, 2),
                "notes": notes,
                "raw_material_id": rm_id,
                "created_at": now_iso(),
            })
            by_name_lower[name_l] = new_id
            inserted += 1

    await db.vendor_price_lists.update_one({"id": vpl_id}, {"$set": {"updated_at": now_iso()}})
    return {
        "inserted": inserted,
        "updated": updated,
        "skipped_count": len(skipped),
        "skipped": skipped[:50],
    }


class SupplierIn(BaseModel):
    name: str
    phone: Optional[str] = ""
    address: Optional[str] = ""
    city: Optional[str] = ""
    gst_number: Optional[str] = ""
    contact_person: Optional[str] = ""
    material_category: Optional[str] = ""
    opening_balance: float = 0.0  # +ve = we owe the supplier at start
    notes: Optional[str] = ""


class SupplierUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    gst_number: Optional[str] = None
    contact_person: Optional[str] = None
    material_category: Optional[str] = None
    opening_balance: Optional[float] = None
    notes: Optional[str] = None


class SupplierPurchaseItemIn(BaseModel):
    raw_material_id: Optional[str] = None
    name: str
    unit: Optional[str] = ""
    quantity: float
    rate: float = 0.0  # per-unit price


class SupplierPurchaseIn(BaseModel):
    supplier_id: str
    amount: Optional[float] = None     # may be omitted / 0 (rate is optional)
    bill_number: Optional[str] = ""
    purchased_at: Optional[str] = None  # ISO date or datetime; default = now
    material: Optional[str] = ""        # free-text description
    notes: Optional[str] = ""
    items: Optional[List[SupplierPurchaseItemIn]] = None  # optional line-items
    bill_images: Optional[List[str]] = None  # array of base64 data URLs (bill photos)


class SupplierPaymentIn(BaseModel):
    supplier_id: str
    amount: float
    source: str = "cash"
    reference: Optional[str] = ""
    paid_at: Optional[str] = None
    notes: Optional[str] = ""
    # If set, this payment was made on behalf of a customer; the same amount
    # is credited to that customer's ledger via the linked /payments doc.
    on_behalf_of_customer_id: Optional[str] = None
    customer_payment_id: Optional[str] = None  # back-link to the customers' payment


@api_router.get("/suppliers")
async def list_suppliers(user=Depends(get_current_user)):
    items = await db.suppliers.find({}, {"_id": 0}).sort("name", 1).to_list(2000)
    return items


@api_router.post("/suppliers")
async def create_supplier(body: SupplierIn, admin=Depends(require_admin)):
    name = (body.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="name required")
    doc = {
        "id": str(uuid.uuid4()),
        "name": name,
        "phone": (body.phone or "").strip(),
        "address": (body.address or "").strip(),
        "city": (body.city or "").strip(),
        "gst_number": (body.gst_number or "").strip(),
        "contact_person": (body.contact_person or "").strip(),
        "material_category": (body.material_category or "").strip(),
        "opening_balance": round(float(body.opening_balance or 0), 2),
        "notes": (body.notes or "").strip(),
        "created_at": now_iso(),
        "created_by": admin["email"],
    }
    await db.suppliers.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.get("/suppliers/{sid}")
async def get_supplier(sid: str, user=Depends(get_current_user)):
    s = await db.suppliers.find_one({"id": sid}, {"_id": 0})
    if not s:
        raise HTTPException(status_code=404, detail="Supplier not found")
    return s


@api_router.patch("/suppliers/{sid}")
async def update_supplier(sid: str, body: SupplierUpdate, admin=Depends(require_action("edit:suppliers"))):
    upd: Dict[str, Any] = {"updated_at": now_iso(), "updated_by": admin["email"]}
    for field in ("name", "phone", "address", "city", "gst_number",
                  "contact_person", "material_category", "notes"):
        val = getattr(body, field, None)
        if val is not None:
            upd[field] = val.strip()
    if body.opening_balance is not None:
        upd["opening_balance"] = round(float(body.opening_balance), 2)
    res = await db.suppliers.update_one({"id": sid}, {"$set": upd})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Supplier not found")
    return await db.suppliers.find_one({"id": sid}, {"_id": 0})


@api_router.delete("/suppliers/{sid}")
async def delete_supplier(sid: str, admin=Depends(require_action("delete:suppliers"))):
    # Refuse if there are linked purchases or payments
    if await db.supplier_purchases.find_one({"supplier_id": sid}, {"_id": 1}):
        raise HTTPException(status_code=400, detail="Cannot delete supplier with purchase history")
    if await db.supplier_payments.find_one({"supplier_id": sid}, {"_id": 1}):
        raise HTTPException(status_code=400, detail="Cannot delete supplier with payment history")
    res = await db.suppliers.delete_one({"id": sid})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Supplier not found")
    return {"ok": True}


@api_router.post("/supplier-purchases")
async def create_supplier_purchase(body: SupplierPurchaseIn, user=Depends(get_current_user)):
    if not body.supplier_id:
        raise HTTPException(status_code=400, detail="supplier_id required")
    # Amount may be zero — purchases can be saved without a price (per user
    # request) and the rate can be added later by an admin from Vendor Ledger.
    if body.amount is not None and float(body.amount) < 0:
        raise HTTPException(status_code=400, detail="Amount cannot be negative")
    sup = await db.suppliers.find_one({"id": body.supplier_id}, {"_id": 0})
    if not sup:
        raise HTTPException(status_code=404, detail="Supplier not found")
    # ---- Vendor price-list lookup (rate auto-fill) ----------------------
    # If the material has a matching vendor-list entry, the rate is auto-
    # filled from the list (the agreed price is binding). Otherwise the
    # client-supplied rate is accepted as-is and may be zero/blank.
    vendor_prices: Dict[str, Dict[str, Any]] = {}
    vpls = await db.vendor_price_lists.find(
        {"vendor_id": body.supplier_id}, {"_id": 0, "id": 1, "name": 1},
    ).to_list(200)
    vpl_ids = [pl["id"] for pl in vpls]
    vpl_name = {pl["id"]: pl.get("name") or "" for pl in vpls}
    if vpl_ids:
        async for it in db.vendor_price_list_items.find(
            {"vendor_price_list_id": {"$in": vpl_ids}}, {"_id": 0},
        ):
            key = str(it.get("name") or "").strip().lower()
            price = float(it.get("price") or 0)
            if not key or price <= 0:
                continue
            vendor_prices[key] = {
                "price": price,
                "unit": it.get("unit") or "",
                "source_list": vpl_name.get(it.get("vendor_price_list_id"), ""),
            }
    is_admin = (user.get("role") == "admin")
    # Normalize line items if present
    items_out: List[Dict[str, Any]] = []
    if body.items:
        for it in body.items:
            qty = float(it.quantity or 0)
            name = (it.name or "").strip()
            if qty <= 0 or not name:
                continue
            vp = vendor_prices.get(name.lower())
            if is_admin:
                # Admin can override anything — accept whatever rate they
                # sent. If they didn't send one, fall back to the vendor
                # list price when available, else 0.
                if it.rate is not None and float(it.rate or 0) > 0:
                    rate = float(it.rate or 0)
                elif vp:
                    rate = float(vp["price"])
                else:
                    rate = 0.0
            else:
                # Operators are NEVER allowed to set the rate. Take the
                # vendor-list price when one exists, otherwise persist 0.
                rate = float(vp["price"]) if vp else 0.0
            if rate < 0:
                raise HTTPException(
                    status_code=400,
                    detail=f"Rate for '{name}' cannot be negative",
                )
            items_out.append({
                "raw_material_id": (it.raw_material_id or None),
                "name": name,
                "unit": (it.unit or "").strip(),
                "quantity": qty,
                "rate": round(rate, 2),
                "line_value": round(rate * qty, 2),
            })
    # Recompute total from items so the persisted amount always matches the
    # authoritative line rates.
    if items_out:
        computed_amount = round(sum(it["line_value"] for it in items_out), 2)
    else:
        computed_amount = round(float(body.amount or 0), 2)
    # If material summary wasn't provided, auto-build one from items
    material = (body.material or "").strip()
    if not material and items_out:
        material = ", ".join(
            f"{it['quantity']} {it['unit']} {it['name']}".strip() for it in items_out
        )
    doc = {
        "id": str(uuid.uuid4()),
        "supplier_id": body.supplier_id,
        "supplier_name": sup.get("name") or "",
        "amount": computed_amount,
        "bill_number": (body.bill_number or "").strip(),
        "material": material,
        "notes": (body.notes or "").strip(),
        "items": items_out,
        "bill_images": [s for s in (body.bill_images or []) if isinstance(s, str) and s.strip()],
        "purchased_at": _normalize_payment_dt(body.purchased_at),
        "created_by": user.get("email") or user.get("username") or "",
        "created_at": now_iso(),
    }
    await db.supplier_purchases.insert_one(doc)
    # Credit stock for every line item that maps to a raw material.
    for it in items_out:
        rid = it.get("raw_material_id")
        qty = float(it.get("quantity") or 0)
        if rid and qty > 0:
            await _apply_rm_movement(
                rid, qty, "purchase", doc["id"],
                user.get("email") or user.get("username") or "",
                notes=f"Purchase from {sup.get('name','')} bill {doc.get('bill_number') or '-'}",
            )
    doc.pop("_id", None)
    return doc


@api_router.delete("/supplier-purchases/{pid}")
async def delete_supplier_purchase(pid: str, admin=Depends(require_action("delete:vendorLedger"))):
    # Reverse any stock credits this purchase made before deleting
    existing = await db.supplier_purchases.find_one({"id": pid}, {"_id": 0})
    if existing:
        for it in existing.get("items") or []:
            rid = it.get("raw_material_id")
            qty = float(it.get("quantity") or 0)
            if rid and qty > 0:
                await _apply_rm_movement(
                    rid, -qty, "purchase_revert", pid, admin["email"],
                    notes=f"Reverted by purchase delete (bill {existing.get('bill_number') or '-'})",
                )
    res = await db.supplier_purchases.delete_one({"id": pid})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Purchase not found")
    return {"ok": True}


class SupplierPurchaseUpdate(BaseModel):
    """Admin-only edit of a recorded purchase from the Vendor Ledger.

    All fields are optional — only those provided are changed. When `items`
    is provided, stock movements are reconciled: previous lines are reverted
    and new lines re-applied so raw-material balances stay accurate.
    """
    supplier_id: Optional[str] = None
    amount: Optional[float] = None
    bill_number: Optional[str] = None
    purchased_at: Optional[str] = None
    material: Optional[str] = None
    notes: Optional[str] = None
    items: Optional[List[SupplierPurchaseItemIn]] = None
    bill_images: Optional[List[str]] = None


@api_router.patch("/supplier-purchases/{pid}")
async def update_supplier_purchase(
    pid: str, body: SupplierPurchaseUpdate, admin=Depends(require_action("edit:vendorLedger")),
):
    existing = await db.supplier_purchases.find_one({"id": pid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Purchase not found")
    upd: Dict[str, Any] = {
        "updated_at": now_iso(),
        "updated_by": admin.get("email") or admin.get("username") or "",
    }

    # Vendor swap (rare but supported)
    new_supplier_id = body.supplier_id or existing.get("supplier_id")
    if new_supplier_id != existing.get("supplier_id"):
        sup = await db.suppliers.find_one({"id": new_supplier_id}, {"_id": 0})
        if not sup:
            raise HTTPException(status_code=404, detail="Supplier not found")
        upd["supplier_id"] = new_supplier_id
        upd["supplier_name"] = sup.get("name") or ""

    # Header fields
    if body.bill_number is not None:
        upd["bill_number"] = body.bill_number.strip()
    if body.notes is not None:
        upd["notes"] = body.notes.strip()
    if body.purchased_at is not None:
        upd["purchased_at"] = _normalize_payment_dt(body.purchased_at)
    if body.material is not None:
        upd["material"] = body.material.strip()
    if body.bill_images is not None:
        upd["bill_images"] = [
            s for s in body.bill_images if isinstance(s, str) and s.strip()
        ]

    # Items + amount reconciliation
    if body.items is not None:
        # Lookup vendor price list for the (possibly new) vendor
        target_supplier_id = upd.get("supplier_id", existing.get("supplier_id"))
        vendor_prices: Dict[str, Dict[str, Any]] = {}
        vpls = await db.vendor_price_lists.find(
            {"vendor_id": target_supplier_id}, {"_id": 0, "id": 1, "name": 1},
        ).to_list(200)
        vpl_ids = [pl["id"] for pl in vpls]
        if vpl_ids:
            async for it in db.vendor_price_list_items.find(
                {"vendor_price_list_id": {"$in": vpl_ids}}, {"_id": 0},
            ):
                key = str(it.get("name") or "").strip().lower()
                price = float(it.get("price") or 0)
                if not key or price <= 0:
                    continue
                vendor_prices[key] = {"price": price, "unit": it.get("unit") or ""}

        new_items: List[Dict[str, Any]] = []
        for it in body.items:
            qty = float(it.quantity or 0)
            name = (it.name or "").strip()
            if qty <= 0 or not name:
                continue
            vp = vendor_prices.get(name.lower())
            # Admin-only endpoint — full override: respect the rate the
            # admin typed when > 0, else fall back to vendor list, else 0.
            if it.rate is not None and float(it.rate or 0) > 0:
                rate = float(it.rate or 0)
            elif vp:
                rate = float(vp["price"])
            else:
                rate = 0.0
            if rate < 0:
                raise HTTPException(
                    status_code=400,
                    detail=f"Rate for '{name}' cannot be negative",
                )
            new_items.append({
                "raw_material_id": (it.raw_material_id or None),
                "name": name,
                "unit": (it.unit or "").strip(),
                "quantity": qty,
                "rate": round(rate, 2),
                "line_value": round(rate * qty, 2),
            })

        # Reverse previous stock movements
        for it in (existing.get("items") or []):
            rid = it.get("raw_material_id")
            qty = float(it.get("quantity") or 0)
            if rid and qty > 0:
                await _apply_rm_movement(
                    rid, -qty, "purchase_revert", pid,
                    admin.get("email") or admin.get("username") or "",
                    notes=f"Edited purchase (bill {existing.get('bill_number') or '-'})",
                )
        # Apply new stock movements
        for it in new_items:
            rid = it.get("raw_material_id")
            qty = float(it.get("quantity") or 0)
            if rid and qty > 0:
                await _apply_rm_movement(
                    rid, qty, "purchase", pid,
                    admin.get("email") or admin.get("username") or "",
                    notes=f"Purchase edited (bill {upd.get('bill_number') or existing.get('bill_number') or '-'})",
                )
        upd["items"] = new_items
        # Recompute amount from items unless the client also supplied an amount
        if body.amount is not None:
            upd["amount"] = round(float(body.amount or 0), 2)
        else:
            upd["amount"] = round(sum(it["line_value"] for it in new_items), 2)
        # Refresh material summary if not explicitly provided
        if body.material is None and new_items:
            upd["material"] = ", ".join(
                f"{it['quantity']} {it['unit']} {it['name']}".strip() for it in new_items
            )
    elif body.amount is not None:
        # Amount-only adjustment (no items change) — just persist it.
        upd["amount"] = round(float(body.amount or 0), 2)

    await db.supplier_purchases.update_one({"id": pid}, {"$set": upd})
    return await db.supplier_purchases.find_one({"id": pid}, {"_id": 0})


@api_router.post("/supplier-payments")
async def create_supplier_payment(body: SupplierPaymentIn, user=Depends(get_current_user)):
    if not body.supplier_id:
        raise HTTPException(status_code=400, detail="supplier_id required")
    if body.amount is None or float(body.amount) <= 0:
        raise HTTPException(status_code=400, detail="Amount must be greater than zero")
    sup = await db.suppliers.find_one({"id": body.supplier_id}, {"_id": 0})
    if not sup:
        raise HTTPException(status_code=404, detail="Supplier not found")
    customer_id = (body.on_behalf_of_customer_id or "").strip() or None
    customer_name = ""
    customer_payment_id = (body.customer_payment_id or "").strip() or None
    if customer_id:
        cust = await db.customers.find_one({"id": customer_id}, {"_id": 0})
        if not cust:
            raise HTTPException(status_code=404, detail="Customer not found")
        customer_name = cust.get("name") or ""
    doc = {
        "id": str(uuid.uuid4()),
        "supplier_id": body.supplier_id,
        "supplier_name": sup.get("name") or "",
        "amount": round(float(body.amount), 2),
        "source": _normalize_payment_source(body.source),
        "reference": (body.reference or "").strip(),
        "paid_at": _normalize_payment_dt(body.paid_at),
        "notes": (body.notes or "").strip(),
        "on_behalf_of_customer_id": customer_id,
        "on_behalf_of_customer_name": customer_name,
        "customer_payment_id": customer_payment_id,
        "created_by": user.get("email") or user.get("username") or "",
        "created_at": now_iso(),
    }
    await db.supplier_payments.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.delete("/supplier-payments/{pid}")
async def delete_supplier_payment(pid: str, admin=Depends(require_action("delete:vendorLedger"))):
    res = await db.supplier_payments.delete_one({"id": pid})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Supplier payment not found")
    return {"ok": True}


@api_router.get("/supplier-ledger/{sid}")
async def supplier_ledger(sid: str, user=Depends(get_current_user)):
    """Combined ledger for a single supplier: purchases (Dr) and payments
    (Cr) interleaved by date with a running balance. Balance > 0 means we
    owe the supplier; balance < 0 means the supplier owes us."""
    sup = await db.suppliers.find_one({"id": sid}, {"_id": 0})
    if not sup:
        raise HTTPException(status_code=404, detail="Supplier not found")
    purchases = await db.supplier_purchases.find({"supplier_id": sid}, {"_id": 0}).to_list(5000)
    payments = await db.supplier_payments.find({"supplier_id": sid}, {"_id": 0}).to_list(5000)
    returns = await db.purchase_returns.find({"supplier_id": sid}, {"_id": 0}).to_list(5000)
    rows: List[Dict[str, Any]] = []
    for p in purchases:
        rows.append({
            "kind": "purchase", "id": p["id"],
            "when": p.get("purchased_at"),
            "debit": float(p.get("amount") or 0), "credit": 0.0,
            "particulars": p.get("material") or "Material purchase",
            "reference": p.get("bill_number") or "",
            "notes": p.get("notes") or "",
            "raw": p,
        })
    for p in payments:
        particulars = "Payment"
        if p.get("on_behalf_of_customer_name"):
            particulars = f"On behalf of {p['on_behalf_of_customer_name']}"
        rows.append({
            "kind": "payment", "id": p["id"],
            "when": p.get("paid_at"),
            "debit": 0.0, "credit": float(p.get("amount") or 0),
            "particulars": particulars,
            "reference": p.get("reference") or "",
            "notes": p.get("notes") or "",
            "raw": p,
        })
    for r in returns:
        rows.append({
            "kind": "purchase_return", "id": r["id"],
            "when": r.get("returned_at"),
            "debit": 0.0, "credit": float(r.get("amount") or 0),
            "particulars": f"Purchase return — {r.get('material') or 'goods returned'}",
            "reference": r.get("reference") or f"PR#{r.get('return_no') or ''}",
            "notes": (r.get("reason") + (" · " + r.get("notes") if r.get("notes") else "")).strip(" ·") if r.get("reason") else (r.get("notes") or ""),
            "raw": r,
        })
    rows.sort(key=lambda r: (r.get("when") or "", r["kind"] == "payment"))
    opening = float(sup.get("opening_balance") or 0)
    bal = opening
    for r in rows:
        bal += r["debit"] - r["credit"]
        r["balance"] = round(bal, 2)
    total_debit = round(sum(r["debit"] for r in rows), 2)
    total_credit = round(sum(r["credit"] for r in rows), 2)
    return {
        "supplier": sup,
        "rows": rows,
        "opening_balance": round(opening, 2),
        "total_debit": total_debit,
        "total_credit": total_credit,
        "closing_balance": round(bal, 2),
    }


# ======================== Sale Returns (Customer) ========================
# A sale return is goods returned by a customer. It is a CREDIT entry on
# the customer's ledger (reduces what they owe us). Stored separately so
# the audit trail is clean.

class SaleReturnIn(BaseModel):
    customer_id: str
    amount: float
    returned_at: Optional[str] = None      # ISO date string; defaults to now
    reference: Optional[str] = ""          # CN / return note number from operator
    reason: Optional[str] = ""             # e.g. "wrong item", "damaged"
    notes: Optional[str] = ""


class SaleReturnUpdate(BaseModel):
    amount: Optional[float] = None
    returned_at: Optional[str] = None
    reference: Optional[str] = None
    reason: Optional[str] = None
    notes: Optional[str] = None


def _normalize_return_dt(s: Optional[str]) -> str:
    """Same lenient parser used for payments / purchases."""
    if not s:
        return now_iso()
    s = s.strip()
    if not s:
        return now_iso()
    if "T" not in s and len(s) == 10:
        s = f"{s}T00:00:00+00:00"
    return s


@api_router.post("/sale-returns")
async def create_sale_return(body: SaleReturnIn, user=Depends(get_current_user)):
    if not body.customer_id:
        raise HTTPException(status_code=400, detail="customer_id required")
    if body.amount is None or float(body.amount) <= 0:
        raise HTTPException(status_code=400, detail="Amount must be greater than zero")
    cust = await db.customers.find_one({"id": body.customer_id}, {"_id": 0})
    if not cust:
        raise HTTPException(status_code=404, detail="Customer not found")
    # Auto-increment return_no (per-customer not global — keep it simple, global is fine)
    last = await db.sale_returns.find_one({}, sort=[("return_no", -1)], projection={"return_no": 1})
    next_no = int((last or {}).get("return_no") or 0) + 1
    doc = {
        "id": str(uuid.uuid4()),
        "return_no": next_no,
        "customer_id": body.customer_id,
        "customer_name": cust.get("name") or "",
        "amount": round(float(body.amount), 2),
        "returned_at": _normalize_return_dt(body.returned_at),
        "reference": (body.reference or "").strip(),
        "reason": (body.reason or "").strip(),
        "notes": (body.notes or "").strip(),
        "created_by": user.get("email") or user.get("username") or "",
        "created_at": now_iso(),
    }
    await db.sale_returns.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.get("/sale-returns")
async def list_sale_returns(
    customer_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    limit: int = 500,
    user=Depends(get_current_user),
):
    q: Dict[str, Any] = {}
    if customer_id:
        q["customer_id"] = customer_id
    if start_date or end_date:
        rng: Dict[str, Any] = {}
        if start_date:
            rng["$gte"] = start_date
        if end_date:
            rng["$lte"] = end_date + "T23:59:59"
        q["returned_at"] = rng
    items = await db.sale_returns.find(q, {"_id": 0}).sort("returned_at", -1).to_list(int(limit))
    total = await db.sale_returns.count_documents(q)
    total_amount = round(sum(float(r.get("amount") or 0) for r in items), 2)
    return {"items": items, "total": total, "total_amount": total_amount}


@api_router.patch("/sale-returns/{rid}")
async def update_sale_return(rid: str, body: SaleReturnUpdate, admin=Depends(require_action("edit:customerLedger"))):
    upd: Dict[str, Any] = {}
    if body.amount is not None:
        if float(body.amount) <= 0:
            raise HTTPException(status_code=400, detail="Amount must be greater than zero")
        upd["amount"] = round(float(body.amount), 2)
    if body.returned_at is not None:
        upd["returned_at"] = _normalize_return_dt(body.returned_at)
    if body.reference is not None:
        upd["reference"] = body.reference.strip()
    if body.reason is not None:
        upd["reason"] = body.reason.strip()
    if body.notes is not None:
        upd["notes"] = body.notes.strip()
    if not upd:
        raise HTTPException(status_code=400, detail="Nothing to update")
    upd["updated_at"] = now_iso()
    res = await db.sale_returns.update_one({"id": rid}, {"$set": upd})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Sale return not found")
    doc = await db.sale_returns.find_one({"id": rid}, {"_id": 0})
    return doc


@api_router.delete("/sale-returns/{rid}")
async def delete_sale_return(rid: str, admin=Depends(require_action("delete:customerLedger"))):
    res = await db.sale_returns.delete_one({"id": rid})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Sale return not found")
    return {"ok": True}


# ======================== Purchase Returns (Vendor) ========================
# A purchase return is goods returned to the vendor. It is a CREDIT entry
# on the vendor's ledger (reduces what we owe them).

class PurchaseReturnIn(BaseModel):
    supplier_id: str
    amount: float
    returned_at: Optional[str] = None
    reference: Optional[str] = ""          # Debit Note number / vendor's ref
    material: Optional[str] = ""           # free text describing what was returned
    reason: Optional[str] = ""             # e.g. "defective", "wrong material"
    notes: Optional[str] = ""


class PurchaseReturnUpdate(BaseModel):
    amount: Optional[float] = None
    returned_at: Optional[str] = None
    reference: Optional[str] = None
    material: Optional[str] = None
    reason: Optional[str] = None
    notes: Optional[str] = None


@api_router.post("/purchase-returns")
async def create_purchase_return(body: PurchaseReturnIn, user=Depends(get_current_user)):
    if not body.supplier_id:
        raise HTTPException(status_code=400, detail="supplier_id required")
    if body.amount is None or float(body.amount) <= 0:
        raise HTTPException(status_code=400, detail="Amount must be greater than zero")
    sup = await db.suppliers.find_one({"id": body.supplier_id}, {"_id": 0})
    if not sup:
        raise HTTPException(status_code=404, detail="Supplier not found")
    last = await db.purchase_returns.find_one({}, sort=[("return_no", -1)], projection={"return_no": 1})
    next_no = int((last or {}).get("return_no") or 0) + 1
    doc = {
        "id": str(uuid.uuid4()),
        "return_no": next_no,
        "supplier_id": body.supplier_id,
        "supplier_name": sup.get("name") or "",
        "amount": round(float(body.amount), 2),
        "returned_at": _normalize_return_dt(body.returned_at),
        "reference": (body.reference or "").strip(),
        "material": (body.material or "").strip(),
        "reason": (body.reason or "").strip(),
        "notes": (body.notes or "").strip(),
        "created_by": user.get("email") or user.get("username") or "",
        "created_at": now_iso(),
    }
    await db.purchase_returns.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.get("/purchase-returns")
async def list_purchase_returns(
    supplier_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    limit: int = 500,
    user=Depends(get_current_user),
):
    q: Dict[str, Any] = {}
    if supplier_id:
        q["supplier_id"] = supplier_id
    if start_date or end_date:
        rng: Dict[str, Any] = {}
        if start_date:
            rng["$gte"] = start_date
        if end_date:
            rng["$lte"] = end_date + "T23:59:59"
        q["returned_at"] = rng
    items = await db.purchase_returns.find(q, {"_id": 0}).sort("returned_at", -1).to_list(int(limit))
    total = await db.purchase_returns.count_documents(q)
    total_amount = round(sum(float(r.get("amount") or 0) for r in items), 2)
    return {"items": items, "total": total, "total_amount": total_amount}


@api_router.patch("/purchase-returns/{rid}")
async def update_purchase_return(rid: str, body: PurchaseReturnUpdate, admin=Depends(require_action("edit:vendorLedger"))):
    upd: Dict[str, Any] = {}
    if body.amount is not None:
        if float(body.amount) <= 0:
            raise HTTPException(status_code=400, detail="Amount must be greater than zero")
        upd["amount"] = round(float(body.amount), 2)
    if body.returned_at is not None:
        upd["returned_at"] = _normalize_return_dt(body.returned_at)
    if body.reference is not None:
        upd["reference"] = body.reference.strip()
    if body.material is not None:
        upd["material"] = body.material.strip()
    if body.reason is not None:
        upd["reason"] = body.reason.strip()
    if body.notes is not None:
        upd["notes"] = body.notes.strip()
    if not upd:
        raise HTTPException(status_code=400, detail="Nothing to update")
    upd["updated_at"] = now_iso()
    res = await db.purchase_returns.update_one({"id": rid}, {"$set": upd})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Purchase return not found")
    doc = await db.purchase_returns.find_one({"id": rid}, {"_id": 0})
    return doc


@api_router.delete("/purchase-returns/{rid}")
async def delete_purchase_return(rid: str, admin=Depends(require_action("delete:vendorLedger"))):
    res = await db.purchase_returns.delete_one({"id": rid})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Purchase return not found")
    return {"ok": True}


# ======================== Health ========================
# ======================== Raw Material Inventory: BOM + Stock Movements ========================
# Each Item (SKU) can have a Bill of Materials (BOM) — the list of raw
# materials and per-unit quantities required to manufacture one piece of
# that specific SKU. When a dispatch goes out, the system multiplies each
# line's quantity by the BOM of that SKU and deducts the matching raw
# materials from stock. Supplier purchases (with raw_material_id on their
# line items) credit stock back. All movements are recorded in
# `raw_material_movements` so the operator can audit every change.

class BomComponentIn(BaseModel):
    raw_material_id: str
    qty_per_unit: float  # how much of this raw material is needed to make ONE piece of the SKU


class ItemBomUpdate(BaseModel):
    components: List[BomComponentIn] = []


class RawMaterialAdjustIn(BaseModel):
    delta: float  # signed adjustment (+ adds stock, - removes)
    notes: Optional[str] = ""


async def _apply_rm_movement(
    raw_material_id: str,
    delta: float,
    kind: str,                  # "purchase" | "dispatch" | "adjust" | "purchase_revert"
    reference_id: Optional[str],
    actor: Optional[str],
    notes: str = "",
) -> Optional[float]:
    """Atomically apply a stock delta to a raw material and log the movement.

    Returns the new balance, or None if the raw material doesn't exist."""
    if not raw_material_id or delta == 0:
        return None
    rm = await db.raw_materials.find_one_and_update(
        {"id": raw_material_id},
        {"$inc": {"stock_on_hand": float(delta)}, "$set": {"updated_at": now_iso()}},
        return_document=True,
        projection={"_id": 0, "id": 1, "name": 1, "unit": 1, "stock_on_hand": 1},
    )
    if not rm:
        logger.warning("RM movement skipped: raw_material %s not found", raw_material_id)
        return None
    await db.raw_material_movements.insert_one({
        "id": str(uuid.uuid4()),
        "raw_material_id": raw_material_id,
        "raw_material_name": rm.get("name"),
        "unit": rm.get("unit"),
        "delta": round(float(delta), 4),
        "balance_after": round(float(rm.get("stock_on_hand") or 0), 4),
        "kind": kind,
        "reference_id": reference_id,
        "actor": actor or "",
        "notes": (notes or "").strip(),
        "at": now_iso(),
    })
    return float(rm.get("stock_on_hand") or 0)


async def _compute_bom_consumption(lines: List[Dict[str, Any]]) -> Dict[str, float]:
    """Aggregate the raw-material qty that WOULD be consumed by these dispatch
    lines, given each line's SKU BOM. Returns {raw_material_id: total_qty}.
    Lines without an item_id or whose SKU has no BOM contribute nothing."""
    if not lines:
        return {}
    item_ids = sorted({l.get("item_id") for l in lines if l.get("item_id")})
    if not item_ids:
        return {}
    items = await db.items.find({"id": {"$in": item_ids}}, {"_id": 0}).to_list(2000)
    imap = {it["id"]: it for it in items}
    consumed: Dict[str, float] = {}
    for line in lines:
        iid = line.get("item_id")
        qty = int(line.get("quantity") or 0)
        if not iid or qty <= 0:
            continue
        sku = imap.get(iid)
        if not sku:
            continue
        for comp in (sku.get("bom") or []):
            rid = comp.get("raw_material_id")
            per = float(comp.get("qty_per_unit") or 0)
            if not rid or per <= 0:
                continue
            consumed[rid] = consumed.get(rid, 0.0) + (per * qty)
    return consumed


async def _assert_bom_stock_available(lines: List[Dict[str, Any]]) -> None:
    """Refuse to proceed when any raw material in the BOM would be driven
    below zero by this dispatch. Raises HTTPException(400) with a human-
    readable list of every RM that's short, so the operator knows exactly
    what to purchase before re-attempting. Purchases (which only add stock)
    are unaffected — only dispatches/consumption are gated.
    """
    consumed = await _compute_bom_consumption(lines)
    if not consumed:
        return
    rm_docs = await db.raw_materials.find(
        {"id": {"$in": list(consumed.keys())}}, {"_id": 0}
    ).to_list(2000)
    rmap = {r["id"]: r for r in rm_docs}
    shortages: List[str] = []
    for rid, need in consumed.items():
        rm = rmap.get(rid)
        if not rm:
            # Unknown RM in BOM — fail loud rather than silently skip.
            shortages.append(f"unknown raw material (id={rid}) referenced by BOM")
            continue
        have = float(rm.get("stock_on_hand") or 0)
        if have + 1e-9 < need:
            short = need - have
            shortages.append(
                f"{rm.get('name')}: need {round(need, 3)} {rm.get('unit') or ''}, "
                f"only {round(have, 3)} in stock (short by {round(short, 3)})"
            )
    if shortages:
        raise HTTPException(
            status_code=400,
            detail="Insufficient raw material stock. Add a purchase first: "
            + "; ".join(shortages),
        )


async def _consume_bom_for_lines(
    lines: List[Dict[str, Any]],
    dispatch_id: str,
    actor: str,
) -> List[Dict[str, Any]]:
    """Look up the BOM for each dispatched line's SKU (item_id), multiply by
    the dispatched quantity, and deduct stock. Stock is **CLAMPED AT ZERO** —
    if a deduction would drive a raw material below zero, only the available
    amount is consumed and a `clamped=True` flag is recorded on the movement
    so the audit trail is honest about what physically happened.

    This lets operators dispatch even when not every historical purchase has
    been keyed in yet; over time, as more purchases get recorded, the
    software's RM balances converge on the physical reality.
    """
    consumed = await _compute_bom_consumption(lines)
    if not consumed:
        return []
    out = []
    for rid, total in consumed.items():
        requested = float(total)
        if requested <= 0:
            continue
        # Read current stock, clamp the deduction at zero, then apply.
        # Concurrent same-RM dispatches are not expected in this factory
        # workflow (a single foreman keys them in), so a simple
        # read-then-write is acceptable and keeps the audit trail accurate.
        rm = await db.raw_materials.find_one(
            {"id": rid},
            {"_id": 0, "id": 1, "name": 1, "unit": 1, "stock_on_hand": 1},
        )
        if not rm:
            logger.warning("RM consumption skipped: raw_material %s not found", rid)
            continue
        current = float(rm.get("stock_on_hand") or 0)
        available = max(0.0, current)
        applied = min(available, requested)
        clamped = applied < requested - 1e-9
        new_stock = max(0.0, current - applied)  # never < 0
        # Atomic update + audit-trail entry. We hand-roll the movement here
        # (rather than calling `_apply_rm_movement`) so we can record
        # `requested` and `clamped` alongside the actual `delta`.
        await db.raw_materials.update_one(
            {"id": rid},
            {"$set": {"stock_on_hand": new_stock, "updated_at": now_iso()}},
        )
        await db.raw_material_movements.insert_one({
            "id": str(uuid.uuid4()),
            "raw_material_id": rid,
            "raw_material_name": rm.get("name"),
            "unit": rm.get("unit"),
            "delta": round(-applied, 4),
            "balance_after": round(new_stock, 4),
            "requested": round(requested, 4),
            "clamped": clamped,
            "kind": "dispatch",
            "reference_id": dispatch_id,
            "actor": actor or "",
            "notes": (
                f"Auto-consumed by dispatch {dispatch_id}"
                + (
                    f" (clamped from {round(requested,4)} → {round(applied,4)} to keep stock ≥ 0)"
                    if clamped else ""
                )
            ),
            "at": now_iso(),
        })
        out.append({
            "raw_material_id": rid,
            "requested": round(requested, 4),
            "consumed": round(applied, 4),
            "balance_after": round(new_stock, 4),
            "clamped": clamped,
        })
    return out


# ---- Endpoints ----
@api_router.get("/items/{iid}/bom")
async def get_item_bom(iid: str, user=Depends(get_current_user)):
    """Return the SKU's BOM with each component enriched with the
    raw-material name, unit, and current stock — for display in the editor."""
    sku = await db.items.find_one({"id": iid}, {"_id": 0})
    if not sku:
        raise HTTPException(status_code=404, detail="Item (SKU) not found")
    bom = list(sku.get("bom") or [])
    rm_ids = [c.get("raw_material_id") for c in bom if c.get("raw_material_id")]
    rms = []
    if rm_ids:
        rms = await db.raw_materials.find(
            {"id": {"$in": rm_ids}}, {"_id": 0}
        ).to_list(1000)
    rmap = {rm["id"]: rm for rm in rms}
    out = []
    for c in bom:
        rid = c.get("raw_material_id")
        rm = rmap.get(rid) or {}
        out.append({
            "raw_material_id": rid,
            "raw_material_name": rm.get("name") or "(deleted)",
            "unit": rm.get("unit") or "",
            "qty_per_unit": float(c.get("qty_per_unit") or 0),
            "stock_on_hand": float(rm.get("stock_on_hand") or 0),
        })
    return {
        "item_id": iid,
        "item_name": sku.get("name"),
        "product_name": sku.get("product_name"),
        "components": out,
    }


@api_router.put("/items/{iid}/bom")
async def set_item_bom(iid: str, body: ItemBomUpdate, admin=Depends(require_action("edit:products"))):
    """Replace the entire BOM for a SKU. Components are de-duplicated by
    raw_material_id (later entries win)."""
    sku = await db.items.find_one({"id": iid}, {"_id": 0, "id": 1, "name": 1})
    if not sku:
        raise HTTPException(status_code=404, detail="Item (SKU) not found")
    rids = [c.raw_material_id for c in body.components]
    if rids:
        existing = await db.raw_materials.find(
            {"id": {"$in": rids}}, {"_id": 0, "id": 1}
        ).to_list(1000)
        existing_ids = {r["id"] for r in existing}
        missing = [r for r in rids if r not in existing_ids]
        if missing:
            raise HTTPException(status_code=400, detail=f"Unknown raw_material_id(s): {missing}")
    seen: Dict[str, float] = {}
    for c in body.components:
        if c.qty_per_unit <= 0:
            raise HTTPException(status_code=400, detail="qty_per_unit must be > 0")
        seen[c.raw_material_id] = float(c.qty_per_unit)
    bom_doc = [{"raw_material_id": rid, "qty_per_unit": round(q, 4)} for rid, q in seen.items()]
    await db.items.update_one(
        {"id": iid},
        {"$set": {"bom": bom_doc, "bom_updated_at": now_iso(), "bom_updated_by": admin["email"]}},
    )
    return await get_item_bom(iid, user=admin)


@api_router.get("/raw-materials/{rid}/movements")
async def list_rm_movements(rid: str, limit: int = 100, user=Depends(get_current_user)):
    """Most recent stock movements for one raw material, newest first."""
    rm = await db.raw_materials.find_one({"id": rid}, {"_id": 0})
    if not rm:
        raise HTTPException(status_code=404, detail="Raw material not found")
    if limit <= 0 or limit > 500:
        limit = 100
    rows = await db.raw_material_movements.find(
        {"raw_material_id": rid}, {"_id": 0}
    ).sort("at", -1).to_list(limit)
    return {
        "raw_material": rm,
        "current_stock": float(rm.get("stock_on_hand") or 0),
        "rows": rows,
    }


@api_router.post("/raw-materials/{rid}/adjust")
async def adjust_rm_stock(rid: str, body: RawMaterialAdjustIn, admin=Depends(require_admin)):
    """Manual ± stock adjustment. Logged in `raw_material_movements`."""
    if body.delta == 0:
        raise HTTPException(status_code=400, detail="delta cannot be zero")
    rm = await db.raw_materials.find_one({"id": rid}, {"_id": 0, "id": 1})
    if not rm:
        raise HTTPException(status_code=404, detail="Raw material not found")
    bal = await _apply_rm_movement(
        rid, float(body.delta), "adjust", None, admin["email"], body.notes or ""
    )
    return {"ok": True, "balance_after": bal}


# ======================== Health ========================
@api_router.get("/")
async def root():
    return {"message": "Factory Order Management API"}


@api_router.get("/pwa/health")
async def pwa_health():
    """Server-side PWA & security check.

    Confirms that the assets a phone needs to install / verify the app are
    actually served and contain the expected hardening markers.
    """
    base = Path(__file__).parent.parent / "frontend" / "public"
    manifest_path = base / "manifest.json"
    sw_path = base / "service-worker.js"

    manifest_ok = False
    manifest_info: Dict[str, Any] = {}
    if manifest_path.exists():
        try:
            mf = json.loads(manifest_path.read_text())
            need = ["name", "short_name", "start_url", "display", "icons"]
            missing = [k for k in need if not mf.get(k)]
            manifest_ok = (not missing) and isinstance(mf.get("icons"), list) and len(mf["icons"]) >= 2
            manifest_info = {
                "name": mf.get("name"),
                "start_url": mf.get("start_url"),
                "display": mf.get("display"),
                "icons_count": len(mf.get("icons") or []),
                "missing_fields": missing,
            }
        except Exception as e:
            manifest_info = {"error": str(e)}

    sw_text = sw_path.read_text() if sw_path.exists() else ""
    sw_checks = {
        "exists": sw_path.exists(),
        "api_bypass": "/api/" in sw_text and "network" in sw_text.lower(),
        "logout_listener": "jk-logout" in sw_text,
        "skip_waiting": "skipWaiting" in sw_text,
    }

    icons = {
        name: (base / name).exists() and (base / name).stat().st_size > 0
        for name in (
            "logo192.png", "logo512.png",
            "logo192-maskable.png", "logo512-maskable.png",
            "apple-touch-icon.png", "favicon.png",
        )
    }

    return {
        "ok": manifest_ok and sw_checks["exists"] and sw_checks["api_bypass"] and sw_checks["logout_listener"] and all(icons.values()),
        "manifest": {"valid": manifest_ok, **manifest_info},
        "service_worker": sw_checks,
        "icons": icons,
        "security": {
            "https_enforced_by_platform": True,
            "api_responses_cached": False,  # SW skips /api/* entirely
            "caches_cleared_on_logout": sw_checks["logout_listener"],
            "jwt_storage": "localStorage (cleared on logout)",
        },
        "checked_at": now_iso(),
    }


# =====================================================================
# Ledger sharing — email PDF / WhatsApp PDF
# =====================================================================
import ledger_share as _lshare


class LedgerRow(BaseModel):
    id: Optional[str] = None
    when: Optional[str] = None
    particulars: Optional[str] = None
    reference: Optional[str] = None
    debit: Optional[float] = 0
    credit: Optional[float] = 0
    balance: Optional[float] = 0
    notes: Optional[str] = None
    model_config = ConfigDict(extra="ignore", coerce_numbers_to_str=True)


class LedgerPartyIn(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    city: Optional[str] = None
    location: Optional[str] = None
    address: Optional[str] = None
    gst_number: Optional[str] = None
    material_category: Optional[str] = None
    model_config = ConfigDict(extra="ignore", coerce_numbers_to_str=True)


class LedgerPeriodIn(BaseModel):
    startDate: Optional[str] = None
    endDate: Optional[str] = None
    model_config = ConfigDict(extra="ignore")


class LedgerSharePayload(BaseModel):
    title: str
    party: Optional[LedgerPartyIn] = None
    period: Optional[LedgerPeriodIn] = None
    opening: float = 0
    closing: float = 0
    total_debit: float = 0
    total_credit: float = 0
    rows: List[LedgerRow] = Field(default_factory=list)
    pcs_total: Optional[float] = None
    model_config = ConfigDict(extra="ignore")


class LedgerEmailRequest(BaseModel):
    recipient_email: str
    payload: LedgerSharePayload


class LedgerWhatsAppRequest(BaseModel):
    recipient_phone: str
    payload: LedgerSharePayload


@api_router.post("/ledger/email")
async def ledger_email(
    body: LedgerEmailRequest,
    user: Dict[str, Any] = Depends(get_current_user),
):
    """Render the ledger PDF and email it to `recipient_email`."""
    try:
        result = await _lshare.share_ledger_email(
            db,
            body.payload.model_dump(),
            body.recipient_email,
        )
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        logger.exception("ledger_email failed")
        raise HTTPException(status_code=500, detail=f"Email failed: {e}")


@api_router.post("/ledger/whatsapp")
async def ledger_whatsapp(
    body: LedgerWhatsAppRequest,
    user: Dict[str, Any] = Depends(get_current_user),
):
    """Render the ledger PDF and deliver it via Meta WhatsApp Cloud API."""
    try:
        result = await _lshare.share_ledger_whatsapp(
            body.payload.model_dump(),
            body.recipient_phone,
        )
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        logger.exception("ledger_whatsapp failed")
        raise HTTPException(status_code=500, detail=f"WhatsApp send failed: {e}")


@api_router.post("/ledger/preview-pdf")
async def ledger_preview_pdf(
    body: LedgerSharePayload,
    user: Dict[str, Any] = Depends(get_current_user),
):
    """Return the rendered PDF bytes for the given payload — used for
    debugging / letting the user download the same PDF that would be
    emailed / sent on WhatsApp."""
    try:
        pdf = _lshare.render_ledger_pdf(body.model_dump())
        headers = {
            "Content-Disposition": (
                'attachment; filename="ledger.pdf"'
            ),
        }
        return Response(content=pdf, media_type="application/pdf", headers=headers)
    except Exception as e:
        logger.exception("ledger_preview_pdf failed")
        raise HTTPException(status_code=500, detail=f"PDF render failed: {e}")


# ---------------------------------------------------------------------
# Dispatch Slip sharing — email PDF / WhatsApp PDF
# ---------------------------------------------------------------------
class SlipItemIn(BaseModel):
    item_name: Optional[str] = None
    description: Optional[str] = None
    quantity: Optional[float] = 0
    unit_price: Optional[float] = 0
    net_unit_price: Optional[float] = 0
    line_value: Optional[float] = 0
    model_config = ConfigDict(extra="ignore", coerce_numbers_to_str=True)


class SlipSharePayload(BaseModel):
    slip_no: Optional[str] = None
    date: Optional[str] = None
    gr_number: Optional[str] = None
    party: Optional[LedgerPartyIn] = None
    transport_name: Optional[str] = None
    dispatched_by: Optional[str] = None
    items: List[SlipItemIn] = Field(default_factory=list)
    bill_amount: float = 0
    cash_amount: float = 0
    grand_total: float = 0
    line_amount: float = 0
    gst: float = 0
    total_pcs: float = 0
    private_mark: Optional[str] = None
    bag_count: float = 0
    notes: Optional[str] = None
    model_config = ConfigDict(extra="ignore", coerce_numbers_to_str=True)


class SlipEmailRequest(BaseModel):
    recipient_email: str
    payload: SlipSharePayload


class SlipWhatsAppRequest(BaseModel):
    recipient_phone: str
    payload: SlipSharePayload


@api_router.post("/slip/email")
async def slip_email(
    body: SlipEmailRequest,
    user: Dict[str, Any] = Depends(get_current_user),
):
    """Render the dispatch slip PDF and email it to `recipient_email`."""
    try:
        result = await _lshare.share_slip_email(
            db,
            body.payload.model_dump(),
            body.recipient_email,
        )
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        logger.exception("slip_email failed")
        raise HTTPException(status_code=500, detail=f"Email failed: {e}")


@api_router.post("/slip/whatsapp")
async def slip_whatsapp(
    body: SlipWhatsAppRequest,
    user: Dict[str, Any] = Depends(get_current_user),
):
    """Render the dispatch slip PDF and deliver it via Meta WhatsApp Cloud API."""
    try:
        result = await _lshare.share_slip_whatsapp(
            body.payload.model_dump(),
            body.recipient_phone,
        )
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        logger.exception("slip_whatsapp failed")
        raise HTTPException(status_code=500, detail=f"WhatsApp send failed: {e}")


# =====================================================================
# AI Assistant — floating chatbot + on-demand summaries
# =====================================================================
import ai_service as _ai
from fastapi.responses import StreamingResponse


class AIChatRequest(BaseModel):
    session_id: Optional[str] = None
    message: str
    confirm_action_id: Optional[str] = None


class AICancelRequest(BaseModel):
    action_id: str


class AISummaryRequest(BaseModel):
    payload: Dict[str, Any] = Field(default_factory=dict)


@api_router.post("/ai/chat/stream")
async def ai_chat_stream(
    body: AIChatRequest,
    user: Dict[str, Any] = Depends(get_current_user),
):
    """SSE stream endpoint for the floating chatbot agent. Emits JSON
    events including tool-calling + confirmation for write actions:
      data: {"type":"session","session_id":"..."}
      data: {"type":"tool_call","tool":"...","args":{...}}
      data: {"type":"tool_result","tool":"...","result":{...}}
      data: {"type":"confirm","action_id":"...","tool":"...","args":{...},"intent":"..."}
      data: {"type":"reply_delta","text":"..."}
      data: {"type":"done","assistant_message_id":"..."}
      data: {"type":"error","message":"..."}
    """
    async def gen():
        try:
            async for evt in _ai.agent_stream(
                db, user, body.session_id, body.message,
                confirm_action_id=body.confirm_action_id,
            ):
                yield f"data: {json.dumps(evt, ensure_ascii=False)}\n\n"
        except Exception as e:
            logger.exception("ai_chat_stream failed")
            yield f"data: {json.dumps({'type':'error','message':str(e)})}\n\n"

    return StreamingResponse(
        gen(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        },
    )


@api_router.post("/ai/chat/cancel")
async def ai_chat_cancel(
    body: AICancelRequest,
    user: Dict[str, Any] = Depends(get_current_user),
):
    user_id = str(user.get("id") or user.get("email") or "anon")
    ok = await _ai.cancel_pending(body.action_id, user_id)
    return {"ok": ok}


@api_router.get("/ai/chat/sessions")
async def ai_list_sessions(user: Dict[str, Any] = Depends(get_current_user)):
    user_id = str(user.get("id") or user.get("email") or "anon")
    sessions = await _ai.list_sessions(db, user_id)
    return {"sessions": sessions}


@api_router.get("/ai/chat/sessions/{session_id}")
async def ai_get_session(
    session_id: str,
    user: Dict[str, Any] = Depends(get_current_user),
):
    user_id = str(user.get("id") or user.get("email") or "anon")
    messages = await _ai.get_session_messages(db, session_id, user_id)
    return {"session_id": session_id, "messages": messages}


@api_router.delete("/ai/chat/sessions/{session_id}")
async def ai_delete_session(
    session_id: str,
    user: Dict[str, Any] = Depends(get_current_user),
):
    user_id = str(user.get("id") or user.get("email") or "anon")
    await _ai.delete_session(db, session_id, user_id)
    return {"ok": True}


@api_router.post("/ai/summary/ledger")
async def ai_summary_ledger(
    body: AISummaryRequest,
    user: Dict[str, Any] = Depends(get_current_user),
):
    try:
        summary = await _ai.summarize_ledger(body.payload or {})
        return {"summary": summary}
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        logger.exception("ai_summary_ledger failed")
        raise HTTPException(status_code=500, detail=f"Summary failed: {e}")


@api_router.post("/ai/summary/dispatch")
async def ai_summary_dispatch(
    body: AISummaryRequest,
    user: Dict[str, Any] = Depends(get_current_user),
):
    try:
        summary = await _ai.summarize_dispatch(body.payload or {})
        return {"summary": summary}
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        logger.exception("ai_summary_dispatch failed")
        raise HTTPException(status_code=500, detail=f"Summary failed: {e}")


app.include_router(api_router)


# ---- Kubernetes / deployment health probes ----
# The platform's liveness/readiness probe calls `GET /health` at the ROOT
# (no /api prefix). Without this route the probe gets a 404 and the pod is
# marked unhealthy, which blocks the deployment from going live.
@app.get("/health")
async def health():
    return {"status": "ok"}


@app.get("/healthz")
async def healthz():
    return {"status": "ok"}


@api_router.get("/health")
async def api_health():
    return {"status": "ok"}


app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
async def on_startup():
    await seed_db()
    # Backfill: ensure every raw material has a numeric `stock_on_hand`.
    await db.raw_materials.update_many(
        {"stock_on_hand": {"$exists": False}},
        {"$set": {"stock_on_hand": 0.0}},
    )
    # One-time BOM migration (Feb 2026): BOMs originally lived on
    # products. They now live on items (SKUs). Copy any product-level BOM
    # to every SKU under that product that does NOT yet have its own BOM,
    # then drop the product-level field so it can't drift.
    legacy_products = await db.products.find(
        {"bom": {"$exists": True, "$ne": []}}, {"_id": 0, "id": 1, "bom": 1}
    ).to_list(1000)
    for lp in legacy_products:
        pid = lp.get("id")
        bom = lp.get("bom") or []
        if pid and bom:
            await db.items.update_many(
                {"product_id": pid, "bom": {"$exists": False}},
                {"$set": {"bom": bom, "bom_updated_at": now_iso(),
                          "bom_updated_by": "migration"}},
            )
    if legacy_products:
        await db.products.update_many(
            {"bom": {"$exists": True}},
            {"$unset": {"bom": "", "bom_updated_at": "", "bom_updated_by": ""}},
        )
        logger.info("BOM migration: moved BOM from %d product(s) to their SKUs", len(legacy_products))

    # One-time clamp (Jun 2026): under the new "stock never goes below zero"
    # policy, any raw_material whose stock_on_hand is currently negative
    # (from older dispatches that consumed unrecorded inventory) is bumped
    # back to 0. A clamp-correction movement is logged so the audit trail
    # explains the adjustment.
    negatives = await db.raw_materials.find(
        {"stock_on_hand": {"$lt": 0}}, {"_id": 0, "id": 1, "name": 1, "unit": 1, "stock_on_hand": 1}
    ).to_list(10000)
    if negatives:
        ts = now_iso()
        bulk_movements = []
        for rm in negatives:
            prev = float(rm.get("stock_on_hand") or 0)
            await db.raw_materials.update_one(
                {"id": rm["id"]},
                {"$set": {"stock_on_hand": 0.0, "updated_at": ts}},
            )
            bulk_movements.append({
                "id": str(uuid.uuid4()),
                "raw_material_id": rm["id"],
                "raw_material_name": rm.get("name"),
                "unit": rm.get("unit"),
                "delta": round(-prev, 4),     # +ve delta brings -X back to 0
                "balance_after": 0.0,
                "kind": "adjust",
                "reference_id": "policy-clamp-zero",
                "actor": "migration",
                "notes": "Stock-never-negative policy: clamped from "
                          f"{round(prev, 4)} to 0.",
                "at": ts,
            })
        if bulk_movements:
            await db.raw_material_movements.insert_many(bulk_movements)
        logger.info("Clamped %d raw material(s) from negative to 0 under new policy", len(negatives))
    # One-time migration: legacy "Delivered" status was removed (Feb 2026).
    # Existing Delivered orders collapse to the new terminal status "Cleared".
    migrated = await db.orders.update_many(
        {"status": "Delivered"}, {"$set": {"status": "Cleared"}}
    )
    if migrated.modified_count:
        logger.info("Migrated %s Delivered → Cleared", migrated.modified_count)

    # One-time backfill: assign sequential numeric slip_no to existing
    # dispatches in chronological order. New dispatches get one via the
    # `next_slip_no()` helper at insert time.
    cursor = db.dispatches.find(
        {"slip_no": {"$exists": False}}, {"_id": 0, "id": 1}
    ).sort("dispatched_at", 1)
    docs = await cursor.to_list(100000)
    if docs:
        # Seed the counter to the current max so we continue from there.
        existing_max = await db.dispatches.find_one(
            {"slip_no": {"$exists": True}}, sort=[("slip_no", -1)], projection={"slip_no": 1, "_id": 0}
        )
        seq = int((existing_max or {}).get("slip_no") or 0)
        for d in docs:
            seq += 1
            await db.dispatches.update_one({"id": d["id"]}, {"$set": {"slip_no": seq}})
        # Sync counter to seq
        await db.counters.update_one(
            {"_id": "dispatch_slip"},
            {"$max": {"seq": seq}},
            upsert=True,
        )
        logger.info("Backfilled slip_no on %s existing dispatches", len(docs))

    # Dedupe migration (Jun 2026): if any historical dispatches share the
    # same slip_no (legacy data, restored backups, etc.), reassign fresh
    # slip_nos to all but the oldest of each duplicate set so slip_no is
    # globally unique. The global atomic counter (`next_slip_no`) only
    # increments, so future numbers continue from the new max.
    dup_groups = await db.dispatches.aggregate([
        {"$match": {"slip_no": {"$exists": True, "$ne": None}}},
        {"$group": {"_id": "$slip_no", "ids": {"$push": "$id"}, "count": {"$sum": 1}}},
        {"$match": {"count": {"$gt": 1}}},
    ]).to_list(100000)
    if dup_groups:
        # Find current max to start reassigning above it.
        max_doc = await db.dispatches.find_one(
            {"slip_no": {"$exists": True}}, sort=[("slip_no", -1)], projection={"slip_no": 1, "_id": 0}
        )
        seq = int((max_doc or {}).get("slip_no") or 0)
        reassigned = 0
        for g in dup_groups:
            # Keep the first id (oldest by insertion order), reassign the rest.
            for did in g["ids"][1:]:
                seq += 1
                await db.dispatches.update_one({"id": did}, {"$set": {"slip_no": seq}})
                reassigned += 1
        await db.counters.update_one(
            {"_id": "dispatch_slip"},
            {"$max": {"seq": seq}},
            upsert=True,
        )
        logger.info(
            "Slip dedupe migration: reassigned slip_no on %d dispatches across %d duplicate group(s)",
            reassigned, len(dup_groups),
        )

    # Enforce uniqueness going forward — a sparse unique index allows
    # documents without slip_no (defensive) but rejects any future
    # duplicate insert / update.
    try:
        await db.dispatches.create_index("slip_no", unique=True, sparse=True, name="slip_no_unique")
    except Exception as e:
        logger.warning("Could not create unique index on dispatches.slip_no: %s", e)
    try:
        await db.deleted_slips.create_index("slip_no", unique=True, name="deleted_slip_no_unique")
    except Exception as e:
        logger.warning("Could not create unique index on deleted_slips.slip_no: %s", e)

    # Unconditional counter sync — fixes the "Dispatch failed / duplicate
    # key error on slip_no_unique" bug that surfaces when the seed data or
    # a restored backup leaves ``db.dispatches`` populated but the
    # ``db.counters.dispatch_slip`` document behind (or missing entirely).
    # The `$max` update advances the counter to at least the highest
    # existing slip_no without ever moving it backwards.
    highest_slip = await db.dispatches.find_one(
        {"slip_no": {"$exists": True, "$ne": None}},
        sort=[("slip_no", -1)],
        projection={"_id": 0, "slip_no": 1},
    )
    max_slip_no = int((highest_slip or {}).get("slip_no") or 0)
    if max_slip_no:
        await db.counters.update_one(
            {"_id": "dispatch_slip"},
            {"$max": {"seq": max_slip_no}},
            upsert=True,
        )
        logger.info("Synced dispatch_slip counter to at least %d", max_slip_no)

    # Same backfill for payments → receipt_no
    pcursor = db.payments.find(
        {"receipt_no": {"$exists": False}}, {"_id": 0, "id": 1}
    ).sort("paid_at", 1)
    pdocs = await pcursor.to_list(100000)
    if pdocs:
        existing_max = await db.payments.find_one(
            {"receipt_no": {"$exists": True}}, sort=[("receipt_no", -1)], projection={"receipt_no": 1, "_id": 0}
        )
        seq = int((existing_max or {}).get("receipt_no") or 0)
        for p in pdocs:
            seq += 1
            await db.payments.update_one({"id": p["id"]}, {"$set": {"receipt_no": seq}})
        await db.counters.update_one(
            {"_id": "payment_receipt"},
            {"$max": {"seq": seq}},
            upsert=True,
        )
        logger.info("Backfilled receipt_no on %s existing payments", len(pdocs))

    # Start the Gmail backup scheduler (daily cron, defaults to 9 PM IST).
    try:
        await backup_mod.init_scheduler(db)
    except Exception as e:
        logger.exception("Failed to init backup scheduler: %s", e)


@app.on_event("shutdown")
async def shutdown_db_client():
    try:
        await backup_mod.shutdown_scheduler()
    except Exception:
        pass
    client.close()
